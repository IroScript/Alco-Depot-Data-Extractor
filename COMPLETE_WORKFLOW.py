"""
================================================================================
COMPLETE DEPOT DATA PROCESSING WORKFLOW
================================================================================
This script combines all steps from data extraction to final target vs achievement report.

WORKFLOW:
1. Extract product-level sales data from depot databases (uses process_product_level_FAST.py output)
2. Create achievement pivot table with concatenated keys
3. Extract target data with correct product names
4. Merge targets into MPO field data with fuzzy matching
5. Fix achievement pivot case sensitivity
6. Create final target vs achievement report with alternating columns

INPUT FILES REQUIRED:
- Product_Level_Net_Sales_*.csv (from process_product_level_FAST.py)
- Unit Target of April-2026 (2).xlsx
- MPO_CODE_AND_FIELD.xlsx

OUTPUT FILES:
- Achievement_Pivot_*.xlsx
- Target_Data_With_Products_*.xlsx
- MPO_Field_With_Targets_*.xlsx
- Achievement_Pivot_Fixed_*.xlsx
- MPO_Target_vs_Achievement_*.xlsx (FINAL OUTPUT with formulas)
- MPO_Target_vs_Achievement_Values_*.xlsx (FINAL OUTPUT with values)

Author: Auto-generated
Date: 2026-05-23
================================================================================
"""

import sys
if hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='backslashreplace')
    except:
        pass

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from difflib import SequenceMatcher
import re
import os
import glob
import warnings
warnings.filterwarnings('ignore')

# ============================================================================
# CONFIGURATION & PRE-PROCESSING
# ============================================================================

# Input files
MPO_FIELD_FILE = r"C:\Users\Irak\Desktop\Barishal April Data\MPO_CODE_AND_FIELD.xlsx"
TARGET_FILE = r"C:\Users\Irak\Desktop\Barishal April Data\Unit Target of April-2026 (2).xlsx"

def clean_excel_file_inplace(filepath):
    """
    Permanently cleans the Excel file in-place:
    1. Replaces CHAR(160) non-breaking spaces with normal spaces.
    2. Removes all control/non-printable characters (ASCII 0-31 and 127-159).
    3. Trims leading/trailing spaces and collapses multiple spaces.
    Only touches text/string cells, preserving formulas, numbers, dates, styles, and sheet structure.
    """
    if not os.path.exists(filepath):
        print(f"WARNING: File {filepath} not found for cleaning.")
        return
    
    print(f"Permanently cleaning '{os.path.basename(filepath)}' in-place...")
    try:
        wb = load_workbook(filepath)
        cleaned_cells_count = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        orig_val = cell.value
                        # Replace CHAR(160) (non-breaking spaces) with normal space
                        val = orig_val.replace('\xa0', ' ')
                        # CLEAN: Remove control/non-printable characters
                        val = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', val)
                        # TRIM: Strip and collapse multiple spaces
                        val = ' '.join(val.strip().split())
                        
                        if val != orig_val:
                            cell.value = val
                            cleaned_cells_count += 1
        if cleaned_cells_count > 0:
            wb.save(filepath)
            print(f"  [OK] Cleaned {cleaned_cells_count} text cells in '{os.path.basename(filepath)}'.")
        else:
            print(f"  [OK] '{os.path.basename(filepath)}' was already 100% clean.")
        wb.close()
    except PermissionError:
        print(f"\nERROR: Permission denied while writing to '{os.path.basename(filepath)}'.")
        print("Please close the file if it is open in Excel and run this script again!")
        sys.exit(1)
    except Exception as e:
        print(f"WARNING: Failed to clean '{os.path.basename(filepath)}' in-place. Error: {e}")

print("=" * 80)
print("PERMANENT IN-PLACE DATA CLEANING")
print("=" * 80)
clean_excel_file_inplace(MPO_FIELD_FILE)
clean_excel_file_inplace(TARGET_FILE)
print("=" * 80 + "\n")

# Find the latest Product_Level_Net_Sales CSV
csv_files = glob.glob(r"C:\Users\Irak\Desktop\Barishal April Data\Product_Level_Net_Sales_*.csv")
if not csv_files:
    print("ERROR: No Product_Level_Net_Sales CSV file found!")
    print("Please run process_product_level_FAST.py first to generate the sales data.")
    exit(1)

ACHIEVEMENT_CSV = max(csv_files, key=os.path.getctime)  # Get the latest file

print("=" * 80)
print("COMPLETE DEPOT DATA PROCESSING WORKFLOW")
print("=" * 80)
print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
print(f"Using achievement file: {os.path.basename(ACHIEVEMENT_CSV)}\n")

# ============================================================================
# STEP 1: CREATE ACHIEVEMENT PIVOT TABLE
# ============================================================================

print("=" * 80)
print("STEP 1: CREATE ACHIEVEMENT PIVOT TABLE")
print("=" * 80)

print("\n1.1: Loading achievement CSV...")
df_sales = pd.read_csv(ACHIEVEMENT_CSV)
print(f"  - Total records: {len(df_sales)}")

print("\n1.2: Adding concatenated columns...")
df_sales['DepotMPO_CodeProduct_Code'] = df_sales['Depot'] + '_' + df_sales['MPO_Code'] + '_' + df_sales['Product_Code']
df_sales['DepotMPO_Code'] = df_sales['Depot'] + '_' + df_sales['MPO_Code']
print(f"  - Added DepotMPO_CodeProduct_Code and DepotMPO_Code")

print("\n1.3: Creating pivot table...")
pivot = pd.pivot_table(
    df_sales,
    values='ACTUAL_SALE_QTY',
    index=['DepotMPO_CodeProduct_Code', 'Depot', 'MPO_Code', 'DepotMPO_Code', 'Product_Code', 'Product_Name'],
    columns='Month',
    aggfunc='sum',
    fill_value=0,
    margins=True,
    margins_name='Grand Total'
)
pivot_df = pivot.reset_index()
print(f"  - Pivot table shape: {pivot_df.shape}")

print("\n1.4: Extracting April achievement...")
april_col = None
for col in pivot_df.columns:
    if '04' in str(col) or 'April' in str(col):
        april_col = col
        break

if april_col:
    april_summary = pivot_df[['DepotMPO_CodeProduct_Code', 'Depot', 'MPO_Code', 'DepotMPO_Code', 
                               'Product_Code', 'Product_Name', april_col]].copy()
    april_summary.rename(columns={april_col: 'April_Achievement'}, inplace=True)
    april_summary = april_summary[april_summary['DepotMPO_CodeProduct_Code'] != 'Grand Total']
    print(f"  - April records: {len(april_summary)}")
    print(f"  - Total April achievement: {april_summary['April_Achievement'].sum():,.0f}")
else:
    print("  ERROR: April column not found!")
    exit(1)

# Add uppercase lookup column
april_summary.insert(0, 'LOOKUP_KEY_UPPER', april_summary['DepotMPO_CodeProduct_Code'].str.upper())
print(f"  - Added LOOKUP_KEY_UPPER column")

timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
achievement_pivot_file = f"Achievement_Pivot_{timestamp}.xlsx"

with pd.ExcelWriter(achievement_pivot_file, engine='openpyxl') as writer:
    pivot_df.to_excel(writer, sheet_name='Pivot_All_Months', index=False)
    april_summary.to_excel(writer, sheet_name='April_Achievement', index=False)

print(f"\n[OK] Achievement pivot saved: {achievement_pivot_file}")

# ============================================================================
# STEP 2: EXTRACT TARGET DATA WITH PRODUCT NAMES
# ============================================================================

print("\n" + "=" * 80)
print("STEP 2: EXTRACT TARGET DATA WITH PRODUCT NAMES")
print("=" * 80)

print("\n2.1: Reading product names from row 3...")
df_raw = pd.read_excel(TARGET_FILE, header=None)
product_names_row = df_raw.iloc[2]

product_names = []
# There are 17 unique products starting at index 8 (columns 8 to 24)
for i in range(8, 25):
    prod_name = product_names_row.iloc[i]
    if pd.notna(prod_name):
        product_names.append(str(prod_name).strip())
    else:
        product_names.append(f"Product_{i}")

print(f"  - Total unique product columns: {len(product_names)}")

print("\n2.2: Reading data and combining General Party and Type-1 & PP targets...")
df_target_raw = pd.read_excel(TARGET_FILE, header=1)

# Assign metadata columns
df_target_raw.columns.values[0] = 'Sl_No'
df_target_raw.columns.values[1] = 'Name'
df_target_raw.columns.values[2] = 'Designation'
df_target_raw.columns.values[3] = 'Market'

# Extract target for MPO/SMPO (General Party) only, completely skipping FM/AM (Type-1 & PP) Targets in columns AA to AQ (index 26 to 42)
for i in range(8, 25):
    prod_name = product_names[i - 8]
    val_gp = pd.to_numeric(df_target_raw.iloc[:, i], errors='coerce').fillna(0)
    df_target_raw.iloc[:, i] = val_gp.round(0).astype(int)
    df_target_raw.columns.values[i] = prod_name

# Keep only columns 0 to 24 (inclusive) which are metadata + the 17 combined target columns
# Keep only columns 0 to 24 (inclusive) which are metadata + the 17 combined target columns
df_target_raw = df_target_raw.iloc[:, :25].copy()

# Extract zones for every single row from df_raw (header=None) by scanning column 0
zones_list = []
current_zone = None
for idx, row in df_raw.iterrows():
    if pd.notna(row.iloc[0]) and 'Zone' in str(row.iloc[0]):
        current_zone = str(row.iloc[0]).replace('Zone :', '').replace('Zone', '').strip()
    zones_list.append(current_zone)

# Assign Zone using our exact row index mapping
df_target_raw['Zone'] = [zones_list[i + 2] for i in range(len(df_target_raw))]

print("\n2.3: Filtering target markets using blacklist (no designation filtering)...")
df_target_raw = df_target_raw[df_target_raw['Market'].notna()]
blacklist = ['total', 'zone', 'present market', 'confirm sales', 'unit target', 'designation', 'name of field forces', 'sl. no.']
df_target_filtered = df_target_raw[~df_target_raw['Market'].astype(str).str.lower().str.contains('|'.join(blacklist), na=False)].reset_index(drop=True)

print(f"  - Total target markets extracted: {len(df_target_filtered)}")
print(f"  - Unique zones found: {df_target_filtered['Zone'].nunique()}")

print("\n2.5: Processing product targets...")
product_cols = product_names.copy()

for col in product_cols:
    if col in df_target_filtered.columns:
        try:
            df_target_filtered.loc[:, col] = pd.to_numeric(df_target_filtered[col], errors='coerce').fillna(0).round(0).astype(int)
        except:
            df_target_filtered.loc[:, col] = 0

df_target_filtered['Total_Target'] = df_target_filtered[product_cols].sum(axis=1)
print(f"  - Product columns: {len(product_cols)}")
print(f"  - Total target quantity: {df_target_filtered['Total_Target'].sum():,.0f}")

target_data_file = f"Target_Data_With_Products_{timestamp}.xlsx"
with pd.ExcelWriter(target_data_file, engine='openpyxl') as writer:
    df_target_filtered[['Zone', 'Market', 'Designation', 'Total_Target']].to_excel(writer, sheet_name='Summary', index=False)
    df_target_filtered.to_excel(writer, sheet_name='Full_Data', index=False)

print(f"\n[OK] Target data saved: {target_data_file}")

# ============================================================================
# STEP 3: MERGE TARGETS INTO MPO FIELD DATA
# ============================================================================

print("\n" + "=" * 80)
print("STEP 3: MERGE TARGETS INTO MPO FIELD DATA")
print("=" * 80)

print("\n3.1: Loading MPO_CODE_AND_FIELD file...")
df_mpo = pd.read_excel(MPO_FIELD_FILE)
print(f"  - Total records: {len(df_mpo)}")

print("\n3.2: Adding DEPOT_MPO_CODE concatenation...")
df_mpo['DEPOT'] = df_mpo['DEPOT'].astype(str)
df_mpo['MPO CODE'] = df_mpo['MPO CODE'].astype(str)
df_mpo['DEPOT_MPO_CODE'] = df_mpo['DEPOT'] + '_' + df_mpo['MPO CODE']
print(f"  - Added DEPOT_MPO_CODE")

print("\n3.3: Normalizing market names for matching...")

def strip_depot_code(name):
    """Remove trailing depot codes like (BARI), (DK.B), (CTG.A), (RNG) etc."""
    import re
    cleaned = re.sub(r'\s*\([A-Z]{2,5}(?:[\.\-][A-Z0-9])?\)\s*$', '', name.strip().upper())
    return cleaned.strip()

def normalize_core(name):
    """Normalize keeping numbers and hyphens - removes only depot suffix"""
    if pd.isna(name):
        return ""
    n = strip_depot_code(str(name))
    n = re.sub(r'\s*-\s*', '-', n)
    n = re.sub(r'\s*\+\s*', '+', n)
    return ' '.join(n.split())

def remove_all_parens(n):
    """Strip ALL parentheticals for fuzzy matching"""
    return ' '.join(re.sub(r'\s*\([^)]*\)\s*', ' ', n).split())

def zone_matches(mpo_zone, target_zone):
    if pd.isna(mpo_zone) or pd.isna(target_zone):
        return True
    m = str(mpo_zone).upper().strip()
    t = str(target_zone).upper().strip()
    if m in t:
        return True
    mappings = {
        'SLT': 'SYLHET',
        'HOBI': 'HOBIGONJ',
        'COM': 'CUMILLA',
        'TANG': 'TANGAIL',
        'NOR': 'NORSHINGDI',
        'NORS': 'NORSHINGDI',
        'DIN': 'DINAJPUR',
        'THAK': 'THAKURGAON',
        'RNG': 'RANGPUR',
        'CTG': 'CHITTAGONG',
        'DK': 'DHAKA',
        'MYM': 'MYMENSINGH',
        'JSR': 'JESSORE',
        'BARI': 'RAJSHAHI',
        'GAIB': 'HATIABANDHA'
    }
    for k, v in mappings.items():
        if k in m and v in t:
            return True
    return False

# Build target list and mappings
# Since target markets can have duplicates (e.g., Sherpur under different zones), we map normalized names to a list of target rows
target_norm_dict = {}
for idx, row in df_target_filtered.iterrows():
    market = row['Market']
    norm = normalize_core(market)
    no_paren = remove_all_parens(norm)
    
    if norm not in target_norm_dict:
        target_norm_dict[norm] = []
    target_norm_dict[norm].append(row)
    
    if no_paren != norm:
        if no_paren not in target_norm_dict:
            target_norm_dict[no_paren] = []
        target_norm_dict[no_paren].append(row)

print("\n3.4: Matching markets (zone-aware two-pass with split-combine support)...")

def safe_num(val):
    v = pd.to_numeric(val, errors='coerce')
    return 0 if pd.isna(v) else v

matches = []      # List of (mpo_market, target_market, score, type)
no_matches = []

for idx, row in df_mpo.iterrows():
    mpo_market = row['MARKET']
    mpo_zone = row['ZONE']
    
    norm = normalize_core(mpo_market)
    no_paren = remove_all_parens(norm)
    
    # Try whole-name match first!
    whole_match_row = None
    whole_score = 1.0
    whole_match_type = 'Exact'
    
    candidates = []
    if norm in target_norm_dict:
        candidates = target_norm_dict[norm]
    elif no_paren in target_norm_dict:
        candidates = target_norm_dict[no_paren]
        
    if candidates:
        zone_matched = [c for c in candidates if zone_matches(mpo_zone, c['Zone'])]
        if zone_matched:
            whole_match_row = zone_matched[0]
        else:
            whole_match_row = candidates[0]
            
    if whole_match_row is not None:
        matches.append((mpo_market, whole_match_row['Market'], whole_score, whole_match_type))
        df_mpo.at[idx, '_matched_to'] = whole_match_row['Market']
        for col in product_cols:
            df_mpo.at[idx, col] = safe_num(whole_match_row[col])
            
    elif '+' in mpo_market:
        # Fallback to split-combine matching if no whole-name match
        parts = [p.strip() for p in mpo_market.split('+')]
        matched_rows_for_parts = []
        for part in parts:
            part_norm = normalize_core(part)
            part_no_paren = remove_all_parens(part_norm)
            
            best_t_row = None
            if part_norm in target_norm_dict:
                part_cands = target_norm_dict[part_norm]
                zone_cands = [c for c in part_cands if zone_matches(mpo_zone, c['Zone'])]
                best_t_row = zone_cands[0] if zone_cands else part_cands[0]
            elif part_no_paren in target_norm_dict:
                part_cands = target_norm_dict[part_no_paren]
                zone_cands = [c for c in part_cands if zone_matches(mpo_zone, c['Zone'])]
                best_t_row = zone_cands[0] if zone_cands else part_cands[0]
                    
            if best_t_row is not None:
                matched_rows_for_parts.append(best_t_row)
                
        if len(matched_rows_for_parts) == len(parts):
            combined_target_name = ' + '.join([str(r['Market']) for r in matched_rows_for_parts])
            matches.append((mpo_market, combined_target_name, 1.0, 'Combined'))
            df_mpo.at[idx, '_matched_to'] = combined_target_name
            for col in product_cols:
                df_mpo.at[idx, col] = sum(safe_num(r[col]) for r in matched_rows_for_parts)
        else:
            no_matches.append((mpo_market, 0.0))
            df_mpo.at[idx, '_matched_to'] = None
            for col in product_cols:
                df_mpo.at[idx, col] = 0
    else:
        no_matches.append((mpo_market, 0.0))
        df_mpo.at[idx, '_matched_to'] = None
        for col in product_cols:
            df_mpo.at[idx, col] = 0

print(f"  - Total matches: {len(matches)}")
print(f"  - Exact matches: {sum(1 for m in matches if m[3] == 'Exact')}")
print(f"  - Combined matches: {sum(1 for m in matches if m[3] == 'Combined')}")
print(f"  - Fuzzy matches: {sum(1 for m in matches if m[3] == 'Fuzzy')}")
print(f"  - No matches: {len(no_matches)}")

base_cols = ['DEPOT', 'ZONE', 'FM/AM, ZONE', 'MARKET', 'MPO CODE', 'DEPOT_MPO_CODE']
final_cols = base_cols + product_cols
# Filter df_mpo to only keep matched rows (exclude unmatched markets from Master Target sheet and final report)
df_mpo_matched = df_mpo[df_mpo['_matched_to'].notna()].copy()
df_mpo_final = df_mpo_matched[final_cols].copy()
df_mpo_final['Total_Target'] = df_mpo_final[product_cols].sum(axis=1)

mpo_field_targets_file = f"MPO_Field_With_Targets_{timestamp}.xlsx"
with pd.ExcelWriter(mpo_field_targets_file, engine='openpyxl') as writer:
    df_mpo_final.to_excel(writer, sheet_name='MPO_Field_Targets', index=False)
    df_mpo_final[base_cols + ['Total_Target']].to_excel(writer, sheet_name='Summary', index=False)
    
    # Match Report
    match_report = pd.DataFrame({
        'MPO_Market': [m[0] for m in matches] + [m[0] for m in no_matches],
        'Target_Market': [m[1] for m in matches] + ['NO MATCH'] * len(no_matches),
        'Match_Score': [m[2] for m in matches] + [m[1] for m in no_matches],
        'Match_Type': [m[3] for m in matches] + ['No Match'] * len(no_matches),
        'MPO_Normalized': [normalize_core(m[0]) for m in matches] + [normalize_core(m[0]) for m in no_matches],
        'Target_Normalized': [normalize_core(m[1]) for m in matches] + [''] * len(no_matches)
    })
    match_report.to_excel(writer, sheet_name='Match_Report', index=False)
    
    # Markets Taken (Successfully Matched)
    markets_taken = pd.DataFrame({
        'Market': [m[0] for m in matches],
        'Matched_To': [m[1] for m in matches],
        'Match_Type': [m[3] for m in matches],
        'Match_Score': [m[2] for m in matches],
        'Total_Target': [df_mpo_final[df_mpo_final['MARKET'] == m[0]]['Total_Target'].iloc[0] if len(df_mpo_final[df_mpo_final['MARKET'] == m[0]]) > 0 else 0 for m in matches]
    })
    markets_taken = markets_taken.sort_values('Market')
    markets_taken.to_excel(writer, sheet_name='Markets_Taken', index=False)
    
    # Markets Skipped (No Match Found)
    markets_skipped = pd.DataFrame({
        'Market': [m[0] for m in no_matches],
        'Best_Score': [m[1] for m in no_matches],
        'Reason': ['No match found'] * len(no_matches),
        'MPO_Normalized': [normalize_core(m[0]) for m in no_matches]
    })
    markets_skipped = markets_skipped.sort_values('Market')
    markets_skipped.to_excel(writer, sheet_name='Markets_Skipped', index=False)
    
    # Target Markets Not Used
    used_targets = set()
    for m in matches:
        if m[3] == 'Combined':
            for t_part in m[1].split(' + '):
                used_targets.add(t_part.strip())
        else:
            used_targets.add(m[1])
            
    all_target_markets = set(df_target_filtered['Market'].unique())
    unused_target_markets = all_target_markets - used_targets
    
    if unused_target_markets:
        unused_df = pd.DataFrame({
            'Target_Market': sorted(list(unused_target_markets)),
            'Reason': ['Not matched to any MPO market'] * len(unused_target_markets),
            'Target_Normalized': [normalize_core(m) for m in sorted(list(unused_target_markets))]
        })
        unused_df.to_excel(writer, sheet_name='Target_Markets_Not_Used', index=False)

print(f"\n[OK] MPO field with targets saved: {mpo_field_targets_file}")
print(f"  - Sheet 1: MPO_Field_Targets (full data)")
print(f"  - Sheet 2: Summary (without products)")
print(f"  - Sheet 3: Match_Report (all matches)")
print(f"  - Sheet 4: Markets_Taken ({len(matches)} markets)")
print(f"  - Sheet 5: Markets_Skipped ({len(no_matches)} markets)")
if unused_target_markets:
    print(f"  - Sheet 6: Target_Markets_Not_Used ({len(unused_target_markets)} markets)")

# ============================================================================
# STEP 4: CREATE FINAL TARGET VS ACHIEVEMENT REPORT
# ============================================================================

print("\n" + "=" * 80)
print("STEP 4: CREATE FINAL TARGET VS ACHIEVEMENT REPORT")
print("=" * 80)

print("\n4.1: Creating product code mapping...")
product_mapping = april_summary[['Product_Code', 'Product_Name']].drop_duplicates()
product_code_dict = dict(zip(product_mapping['Product_Name'], product_mapping['Product_Code']))
print(f"  - Unique products in achievement: {len(product_code_dict)}")

print("\n4.2: Matching target products to codes...")

def normalize_product_name(name):
    if pd.isna(name):
        return ""
    name = str(name).upper().strip()
    name = re.sub(r'\s*(TAB|TABLET|CAP|CAPSULE|SYP|SYRUP|SUSP|SUSPENSION)\.?\s*', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\([^)]*\)', '', name)
    name = re.sub(r'\d+\s*MG', '', name, flags=re.IGNORECASE)
    name = name.replace('-', '').replace('.', '').replace(',', '').replace('+', '')
    name = ' '.join(name.split())
    return name.replace(' ', '')

def find_product_code(target_product, product_code_dict):
    target_norm = normalize_product_name(target_product)
    best_match = None
    best_score = 0
    
    for ach_product, code in product_code_dict.items():
        ach_norm = normalize_product_name(ach_product)
        score = SequenceMatcher(None, target_norm, ach_norm).ratio()
        if score > best_score:
            best_score = score
            best_match = code
    
    if best_score >= 0.75:
        return best_match
    else:
        return None

target_to_code = {}
for prod in product_cols:
    code = find_product_code(prod, product_code_dict)
    target_to_code[prod] = code if code else ""

matched = sum(1 for v in target_to_code.values() if v)
print(f"  - Matched: {matched}/{len(product_cols)}")

print("\n4.3: Creating dataframe with alternating columns...")
new_columns = base_cols.copy()
new_col_mapping = {}

for prod in product_cols:
    if target_to_code.get(prod):
        target_col = f"{prod}"
        ach_col = f"{prod}_ACH"
        new_columns.append(target_col)
        new_columns.append(ach_col)
        new_col_mapping[target_col] = (prod, 'target')
        new_col_mapping[ach_col] = (prod, 'ach')

df_new = pd.DataFrame(columns=new_columns)
for col in base_cols:
    df_new[col] = df_mpo_final[col]

for new_col, (orig_col, col_type) in new_col_mapping.items():
    if col_type == 'target':
        df_new[new_col] = df_mpo_final[orig_col]
    else:
        df_new[new_col] = 0

print(f"  - New shape: {df_new.shape}")

print("\n4.4: Adding product codes and labels rows...")
row1_data = {}
for col in base_cols:
    row1_data[col] = ""
for new_col, (orig_col, col_type) in new_col_mapping.items():
    if col_type == 'target':
        row1_data[new_col] = ""
    else:
        row1_data[new_col] = target_to_code.get(orig_col, "")

row2_data = {}
for col in base_cols:
    row2_data[col] = ""
for new_col, (orig_col, col_type) in new_col_mapping.items():
    if col_type == 'target':
        row2_data[new_col] = "TARGET"
    else:
        row2_data[new_col] = "ACH."

df_row1 = pd.DataFrame([row1_data])
df_row2 = pd.DataFrame([row2_data])
df_final = pd.concat([df_row1, df_row2, df_new], ignore_index=True)

print(f"  - Final shape: {df_final.shape}")

print("\n4.5: Saving with VLOOKUP formulas...")
output_file = f"MPO_Target_vs_Achievement_{timestamp}.xlsx"
df_final.to_excel(output_file, sheet_name='Target_vs_Achievement', index=False)

wb = load_workbook(output_file)
ws = wb['Target_vs_Achievement']

for excel_col_idx, col_name in enumerate(df_final.columns, start=1):
    if col_name in new_col_mapping:
        orig_col, col_type = new_col_mapping[col_name]
        
        if col_type == 'ach':
            col_letter = get_column_letter(excel_col_idx)
            product_code_cell = f"{col_letter}$2"
            depot_mpo_cell = "$F"
            
            for row_idx in range(4, len(df_final) + 2):
                formula = f'=IFERROR(VLOOKUP(UPPER({depot_mpo_cell}{row_idx}&"_"&{product_code_cell}),\'[{achievement_pivot_file}]April_Achievement\'!$A:$H,8,FALSE),0)'
                ws[f"{col_letter}{row_idx}"] = formula

wb.save(output_file)
wb.close()
print(f"  - Formulas added for {matched} products")

print("\n4.6: Creating version with calculated values...")
achievement_lookup = {}
for idx, row in april_summary.iterrows():
    key = row['LOOKUP_KEY_UPPER']
    achievement_lookup[key] = row['April_Achievement']

df_with_values = pd.read_excel(output_file, sheet_name='Target_vs_Achievement')
for col in df_with_values.columns:
    df_with_values[col] = df_with_values[col].astype(object)

for new_col, (orig_col, col_type) in new_col_mapping.items():
    if col_type == 'ach':
        code = target_to_code.get(orig_col)
        if code:
            for row_idx in range(2, len(df_with_values)):
                depot_mpo = df_with_values.iloc[row_idx]['DEPOT_MPO_CODE']
                if pd.notna(depot_mpo):
                    key = f"{depot_mpo}_{code}".upper()
                    ach_value = achievement_lookup.get(key, 0)
                    if isinstance(ach_value, (int, float)) and not pd.isna(ach_value):
                        df_with_values.at[row_idx, new_col] = int(ach_value) if ach_value == int(ach_value) else ach_value
                    else:
                        df_with_values.at[row_idx, new_col] = 0

values_file = f"MPO_Target_vs_Achievement_Values_{timestamp}.xlsx"
df_with_values.to_excel(values_file, sheet_name='Target_vs_Achievement', index=False)

print(f"\n[OK] Target vs Achievement saved: {output_file}")
print(f"[OK] Values version saved: {values_file}")

# ============================================================================
# FINAL SUMMARY
# ============================================================================

print("\n" + "=" * 80)
print("WORKFLOW COMPLETED SUCCESSFULLY!")
print("=" * 80)

print(f"\nFILES CREATED:")
print(f"  1. {achievement_pivot_file}")
print(f"  2. {target_data_file}")
print(f"  3. {mpo_field_targets_file} * (with verification sheets)")
print(f"  4. {output_file} (FINAL - with VLOOKUP formulas)")
print(f"  5. {values_file} (FINAL - with calculated values)")

print(f"\nSTATISTICS:")
print(f"  - Total MPO records: {len(df_mpo_final)}")
print(f"  - Product columns: {len(product_cols)}")
print(f"  - Products matched: {matched}/{len(product_cols)}")
print(f"  - Total target: {df_mpo_final['Total_Target'].sum():,.0f}")
print(f"  - Total achievement: {april_summary['April_Achievement'].sum():,.0f}")

print(f"\nMARKET MATCHING:")
exact_count = sum(1 for m in matches if m[3] == 'Exact')
combined_count = sum(1 for m in matches if m[3] == 'Combined')
fuzzy_count = sum(1 for m in matches if m[3] == 'Fuzzy')
no_match_count = len(no_matches)
total_matched = len(matches)
print(f"  - Exact matches: {exact_count} ({exact_count/len(df_mpo)*100:.1f}%)")
print(f"  - Combined matches: {combined_count} ({combined_count/len(df_mpo)*100:.1f}%)")
print(f"  - Fuzzy matches: {fuzzy_count} ({fuzzy_count/len(df_mpo)*100:.1f}%)")
print(f"  - No matches (ZERO targets): {no_match_count} ({no_match_count/len(df_mpo)*100:.1f}%)")
print(f"  - Total matched: {total_matched} ({total_matched/len(df_mpo)*100:.1f}%)")

if no_matches:
    print(f"\n[WARNING] MARKETS WITH ZERO TARGETS (first 10):")
    for market, score in no_matches[:10]:
        print(f"    - {market} (best score: {score:.2f})")
    if len(no_matches) > 10:
        print(f"    ... and {len(no_matches) - 10} more (see Markets_Skipped sheet)")

print("\n" + "=" * 80)
print(f"Completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print("=" * 80)
print("\nVERIFICATION SHEETS:")
print(f"  [OK] Open '{mpo_field_targets_file}'")
print(f"  [OK] Check 'Markets_Taken' sheet - All successfully matched markets")
print(f"  [OK] Check 'Markets_Skipped' sheet - Markets with ZERO targets")
print(f"  [OK] Check 'Target_Markets_Not_Used' sheet - Target markets not matched to any MPO")
print(f"  [OK] Check 'Match_Report' sheet - Detailed matching information")
print("\n" + "=" * 80)
