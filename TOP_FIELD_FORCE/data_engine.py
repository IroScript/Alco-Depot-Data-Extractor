import os
import sys
import sqlite3
import json
from datetime import datetime

# Define base directories
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(BASE_DIR)
DEFAULT_DB_PATH = os.path.join(PARENT_DIR, "sales.db")
DATA_OUT_DIR = os.path.join(BASE_DIR, "data")

os.makedirs(DATA_OUT_DIR, exist_ok=True)

# 6 Strategic Product Groups defined by user
STRATEGIC_6_MAPPING = {
    'ALAGRA 120 TAB': ['ALK1', 'ALM1', 'ZA04', 'ZA05'],
    'ALAGRA 180 TAB': ['ALN1', 'ALP1'],
    'AMDIN PLUS TAB': ['AMK3', 'AMM3', 'ZA11'],
    'DERMA CAP': ['DEJ1', 'DEK1', 'DEM1', 'DEN1', 'ZD01'],
    'MOKAST 10 TAB': ['MON1', 'MOO1', 'MOP1'],
    'TOLEC TAB': ['TOL2']
}

def load_excel_mappings():
    excel_path = os.path.join(PARENT_DIR, "PRODUCT_CODE_AND_SUBGROUP_OF_PRODUCTS.xlsx")
    code_to_subgroup = {}
    if os.path.exists(excel_path):
        try:
            import pandas as pd
            df_map = pd.read_excel(excel_path)
            df_map["Product_Code_ffill"] = df_map["Product_Code"].ffill()
            for _, row in df_map.iterrows():
                code = str(row["Product_Code_ffill"]).strip().upper()
                subg = str(row["SUB_GROUP_STANDARD"]).strip() if pd.notna(row["SUB_GROUP_STANDARD"]) else ""
                if code and code != "NAN" and subg and subg != "nan":
                    code_to_subgroup[code] = subg
        except Exception as e:
            print(f"Note: Could not read Excel mapping via pandas: {e}")
            
    # Always ensure strategic 6 mappings exist
    for grp_name, codes in STRATEGIC_6_MAPPING.items():
        for c in codes:
            code_to_subgroup[c] = grp_name
            
    return code_to_subgroup

def _clean_cell(val):
    """Normalise a cell value to a trimmed string; '' for None/blank/'None'."""
    if val is None:
        return ''
    s = str(val).strip()
    return '' if s in ('None', 'nan', 'NaN') else s


def _clean_market_name(name):
    """Human-level market name formatting (mirrors FieldEdit's clean_market_name)."""
    if not name:
        return ''
    name = _clean_cell(name).upper()
    import re
    name = re.sub(r'\s+', ' ', name)
    if name == 'HATIBANDHA (HATIB.)':
        return 'HATIBANDHA (HATIBANDHA-1)'
    return name


def _add_from_excel(mpo_map, fpath, header_row, code_cols, market_col, skip_if_present=True):
    """Generic helper: read one workbook and union code->market entries.

    code_cols / market_col are 1-based column numbers as seen in Excel.
    Multiple code columns (aliases) all map to the same market.
    """
    try:
        import openpyxl
        if not os.path.exists(fpath):
            return
        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws = wb.active
        for r in range(header_row + 1, ws.max_row + 1):
            mkt = _clean_market_name(ws.cell(row=r, column=market_col).value)
            if not mkt:
                continue
            for cc in code_cols:
                code = _clean_cell(ws.cell(row=r, column=cc).value)
                if code and (not skip_if_present or code not in mpo_map):
                    mpo_map[code] = mkt
    except Exception as e:
        print(f"Note: Could not read {os.path.basename(fpath)}: {e}")


def _try_google_sheet_market_map(mpo_map):
    """Attempt the live FieldEdit Google Sheet (complete source of truth).

    Replicates FieldEdit/field_formatter_updated_gui_active_data.py:process_excel
    auth flow + cutoff logic + clean_market_name. Silent no-op if credentials
    are unavailable or the service-account key is revoked.
    """
    try:
        import json
        fe_dir = os.path.join(PARENT_DIR, "FieldEdit")
        config_path = os.path.join(fe_dir, "config.json")
        if not os.path.exists(config_path):
            return
        with open(config_path, 'r') as f:
            cfg = json.load(f)
        creds_path = cfg.get('credentials_file', 'alco-pharma-cf4b49e394bb.json')
        if not os.path.isabs(creds_path):
            creds_path = os.path.join(fe_dir, creds_path)
        if not os.path.exists(creds_path):
            return

        import gspread
        from google.oauth2.service_account import Credentials
        scopes = [
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive'
        ]
        creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
        client = gspread.authorize(creds)
        sheet = client.open_by_key(cfg['spreadsheet_id'])

        worksheet = None
        target_gid = str(cfg.get('gid', '1918615875'))
        for ws in sheet.worksheets():
            if str(ws.id) == target_gid:
                worksheet = ws
                break
        if not worksheet:
            worksheet = sheet.get_worksheet(0)

        all_values = worksheet.get_all_values()

        # cutoff at the secondary data block marker (FieldEdit logic)
        cutoff_idx = len(all_values)
        for r_idx in range(1, len(all_values)):
            row_str = " ".join([str(c) for c in all_values[r_idx] if c is not None])
            if "FM (SELF APP CODE)" in row_str or "FM (SELF" in row_str:
                cutoff_idx = r_idx
                break
        all_values = all_values[:cutoff_idx]
        while len(all_values) > 1:
            if all(c is None or str(c).strip() == "" for c in all_values[-1]):
                all_values.pop()
            else:
                break

        def gv(row, idx):
            return row[idx] if idx < len(row) else ""

        added = 0
        for row in all_values[1:]:
            mkt = _clean_market_name(gv(row, 3))  # MARKET (idx3)
            if not mkt:
                continue
            # All code aliases present in the sheet schema
            for ci in (2, 4, 12, 13, 15):  # NEW CODE, OLD CODE, DREAM, DEPOTMPO, APP CODE FINAL
                code = _clean_cell(gv(row, ci))
                if code and code not in mpo_map:
                    mpo_map[code] = mkt
                    added += 1
        print(f"Live Google Sheet: enriched {added} additional MPO->market entries.")
    except Exception as e:
        print(f"Note: Live Google Sheet unavailable ({str(e)[:80]}). Using local Excel union + depot fallback.")


def load_mpo_market_lookup():
    """
    Build the canonical MPO_CODE -> MARKET_NAME map.

    Field-force understanding (source priority):
      1. Live FieldEdit Google Sheet (complete source of truth) — attempted first;
         silently falls back if the service account key is revoked/unavailable.
      2. FieldEdit/Archive/FIELD.xlsx  — flat export of that same Google Sheet
         (DEPOT, ZONE, NEW CODE, MARKET, OLD CODE, ..., DREAM APPS MPO CODE,
          DEPOTMPO CODE, APP CODE (FINAL)). Multiple code columns all map to MARKET.
      3. 02E_FINAL_MPO_Target_vs_Achievement (col5 code -> col4 market).
      4. archive/03_Zone_Wise_Sales_Grouped_Report (FIXED: col9 DREAM APPS MPO CODE
         -> col3 MARKET). NOTE: the previous code wrongly read col4 (FM/AM,ZONE text)
         as the code; the real MPO code lives in col9.
      5. archive/recent/mpo_code.xlsx (the file init_db.py joins against).

    Market names are normalised via clean_market_name (uppercase / whitespace /
    HATIBANDHA override) to match the human-level standard formatting used by
    the field-force reports.
    """
    mpo_map = {}

    # 1. Live Google Sheet (best-effort; enriches on top of local union)
    _try_google_sheet_market_map(mpo_map)

    # 2. FIELD.xlsx — flat export of the live Google Sheet (richest local source)
    _add_from_excel(
        mpo_map,
        os.path.join(PARENT_DIR, "FieldEdit", "Archive", "FIELD.xlsx"),
        header_row=1,
        code_cols=[3, 5, 13, 14, 16],   # NEW CODE, OLD CODE, DREAM, DEPOTMPO, APP CODE FINAL
        market_col=4,                    # MARKET
    )

    # 3. 02E achievement report (code col5 -> market col4)
    _add_from_excel(
        mpo_map,
        os.path.join(PARENT_DIR, "02E_FINAL_MPO_Target_vs_Achievement_Values_24_Jun_2026_02.13_PM.xlsx"),
        header_row=3,
        code_cols=[5],
        market_col=4,
    )

    # 4. 03_Zone_Wise — FIXED: real MPO code is in col9 (DREAM APPS MPO CODE), market in col3
    _add_from_excel(
        mpo_map,
        os.path.join(PARENT_DIR, "archive", "03_Zone_Wise_Sales_Grouped_Report_23_Jun_2026_09.30_AM.xlsx"),
        header_row=1,
        code_cols=[9],   # was wrongly 4
        market_col=3,
    )

    # 5. mpo_code.xlsx — the file the init pipeline LEFT JOINs against
    #    Schema: col1 DEPOT, col2 ZONE, col3 MARKET, col4 FM/AM, col5 MPO CODE
    _add_from_excel(
        mpo_map,
        os.path.join(PARENT_DIR, "archive", "recent", "mpo_code.xlsx"),
        header_row=1,
        code_cols=[5],   # MPO CODE
        market_col=3,     # MARKET
    )

    return mpo_map

class DataEngine:
    def __init__(self, db_path=DEFAULT_DB_PATH):
        self.db_path = db_path
        if not os.path.exists(self.db_path):
            print(f"Warning: Database not found at {self.db_path}")

    def get_connection(self):
        return sqlite3.connect(self.db_path)

    def generate_all_data(self):
        """
        Executes all SQL calculations emphasizing PRODUCT CODE as primary identifier,
        grouping by SUB_GROUP_STANDARD, and calculating the 6 Strategic Products Top 50 MPO by UNIT.
        """
        conn = self.get_connection()
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        
        mpo_market_map = load_mpo_market_lookup()

        # CRITICAL: Enrich with the DB's own market data. For MPO codes where some
        # rows have a non-null market, the most common market is the real answer
        # (the DB already got it from init_db.py's LEFT JOIN against mpo_code.xlsx).
        # Only codes with ALL-null rows need the external lookup.
        cur.execute("""
            SELECT mpo_code, market, COUNT(*) as cnt
            FROM sales
            WHERE market IS NOT NULL AND TRIM(market) != ''
            GROUP BY mpo_code, market
            ORDER BY mpo_code, cnt DESC
        """)
        for row in cur.fetchall():
            code = _clean_cell(row[0])
            mkt = _clean_market_name(row[1])
            if code and mkt and code not in mpo_market_map:
                mpo_market_map[code] = mkt

        data = {
            "meta": {
                "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "db_path": self.db_path,
                "note": "Product calculations are anchored on product_code and grouped by SUB_GROUP_STANDARD."
            },
            "kpis": {},
            "top_50_products": [],
            "top_5_products_deep": [],
            "top_50_mpos": [],
            "top_20_fms": [],
            "top_5_sector_heads": [],
            "monthly_trends": [],
            "strategic_6_products": {}
        }

        try:
            # 1. OVERVIEW KPIS
            cur.execute("""
                SELECT 
                    SUM(line_amount) as total_sales,
                    COUNT(DISTINCT invoice_no) as total_invoices,
                    COUNT(DISTINCT customer_id) as total_parties,
                    COUNT(DISTINCT mpo_code) as total_mpos,
                    COUNT(DISTINCT product_code) as total_products,
                    COUNT(DISTINCT fm_am) as total_fms
                FROM sales
            """)
            row = cur.fetchone()
            total_sales = float(row["total_sales"] or 0)
            total_invoices = int(row["total_invoices"] or 0)
            total_parties = int(row["total_parties"] or 0)
            
            data["kpis"] = {
                "total_sales": total_sales,
                "total_invoices": total_invoices,
                "total_parties": total_parties,
                "total_mpos": int(row["total_mpos"] or 0),
                "total_products": int(row["total_products"] or 0),
                "total_fms": int(row["total_fms"] or 0),
                "avg_invoice_val": round(total_sales / total_invoices, 2) if total_invoices > 0 else 0
            }

            # 2. MONTHLY TRENDS (Overall)
            cur.execute("""
                SELECT 
                    month,
                    SUM(line_amount) as sales,
                    COUNT(DISTINCT invoice_no) as invoices,
                    COUNT(DISTINCT customer_id) as parties,
                    SUM(quantity) as quantity
                FROM sales
                GROUP BY month
                ORDER BY month ASC
            """)
            for r in cur.fetchall():
                data["monthly_trends"].append({
                    "month": r["month"],
                    "sales": round(float(r["sales"] or 0), 2),
                    "invoices": int(r["invoices"] or 0),
                    "parties": int(r["parties"] or 0),
                    "quantity": round(float(r["quantity"] or 0), 2)
                })

            # 3. TOP 50 PRODUCTS CALCULATION (Grouped/merged by SUB_GROUP_STANDARD as per user rule)
            code_to_subgroup = load_excel_mappings()

            cur.execute("""
                SELECT 
                    product_code,
                    MAX(product_name) as product_name,
                    SUM(line_amount) as total_sales,
                    SUM(quantity) as total_quantity,
                    COUNT(DISTINCT invoice_no) as total_invoices,
                    COUNT(DISTINCT customer_id) as total_parties
                FROM sales
                WHERE product_code IS NOT NULL AND product_code != ''
                GROUP BY product_code
            """)
            raw_prod_rows = cur.fetchall()
            
            # Group/merge by SUB_GROUP_STANDARD
            grouped_prods = {}
            for r in raw_prod_rows:
                p_code = str(r["product_code"]).strip().upper()
                grp_name = code_to_subgroup.get(p_code, r["product_name"] or p_code)
                if grp_name not in grouped_prods:
                    grouped_prods[grp_name] = {
                        "group_name": grp_name,
                        "codes": [],
                        "total_sales": 0.0,
                        "total_quantity": 0.0,
                        "total_invoices": 0,
                        "total_parties": 0
                    }
                grouped_prods[grp_name]["codes"].append(p_code)
                grouped_prods[grp_name]["total_sales"] += float(r["total_sales"] or 0)
                grouped_prods[grp_name]["total_quantity"] += float(r["total_quantity"] or 0)
                grouped_prods[grp_name]["total_invoices"] += int(r["total_invoices"] or 0)
                grouped_prods[grp_name]["total_parties"] += int(r["total_parties"] or 0)

            # Sort descending by total_sales and take Top 50
            sorted_groups = sorted(grouped_prods.values(), key=lambda x: x["total_sales"], reverse=True)[:50]
            
            for idx, g in enumerate(sorted_groups, 1):
                placeholders = ",".join(["?"] * len(g["codes"]))
                cur.execute(f"""
                    SELECT 
                        month,
                        SUM(line_amount) as m_sales,
                        SUM(quantity) as m_qty,
                        COUNT(DISTINCT invoice_no) as m_invoices,
                        COUNT(DISTINCT customer_id) as m_parties
                    FROM sales
                    WHERE product_code IN ({placeholders})
                    GROUP BY month
                    ORDER BY month ASC
                """, g["codes"])
                m_breakdown = []
                for mb in cur.fetchall():
                    m_breakdown.append({
                        "month": mb["month"],
                        "sales": round(float(mb["m_sales"] or 0), 2),
                        "quantity": round(float(mb["m_qty"] or 0), 2),
                        "invoices": int(mb["m_invoices"] or 0),
                        "parties": int(mb["m_parties"] or 0)
                    })

                prod_item = {
                    "rank": idx,
                    "product_code": ", ".join(g["codes"]),
                    "product_name": g["group_name"],
                    "total_sales": round(g["total_sales"], 2),
                    "total_quantity": round(g["total_quantity"], 2),
                    "total_invoices": g["total_invoices"],
                    "total_parties": g["total_parties"],
                    "contribution_pct": round((g["total_sales"] / total_sales) * 100, 2) if total_sales > 0 else 0,
                    "monthly_breakdown": m_breakdown
                }
                data["top_50_products"].append(prod_item)
                if idx <= 5:
                    data["top_5_products_deep"].append(prod_item)

            # 4. TOP 50 MPO CALCULATION
            cur.execute("""
                SELECT 
                    mpo_code,
                    MAX(zone) as zone,
                    MAX(depot) as depot,
                    MAX(market) as market,
                    SUM(line_amount) as total_sales,
                    SUM(quantity) as total_quantity,
                    COUNT(DISTINCT invoice_no) as total_invoices,
                    COUNT(DISTINCT customer_id) as total_parties
                FROM sales
                WHERE mpo_code IS NOT NULL AND mpo_code != ''
                GROUP BY mpo_code
                ORDER BY SUM(line_amount) DESC
                LIMIT 50
            """)
            top_50_mpo_rows = cur.fetchall()
            
            for idx, r in enumerate(top_50_mpo_rows, 1):
                m_code = r["mpo_code"]
                m_sales = float(r["total_sales"] or 0)
                
                cur.execute("""
                    SELECT 
                        month,
                        SUM(line_amount) as m_sales,
                        COUNT(DISTINCT invoice_no) as m_invoices,
                        COUNT(DISTINCT customer_id) as m_parties
                    FROM sales
                    WHERE mpo_code = ?
                    GROUP BY month
                    ORDER BY month ASC
                """, (m_code,))
                m_breakdown = []
                for mb in cur.fetchall():
                    m_breakdown.append({
                        "month": mb["month"],
                        "sales": round(float(mb["m_sales"] or 0), 2),
                        "invoices": int(mb["m_invoices"] or 0),
                        "parties": int(mb["m_parties"] or 0)
                    })

                raw_mkt = r["market"]
                if not raw_mkt or str(raw_mkt).strip() in ['', 'None', 'Unknown'] or str(raw_mkt).strip().startswith('DK.'):
                    # Zone codes like 'DK.A'/'DK.B' are NOT market names (they mean Dhaka-A/B);
                    # fall back to the mapped market, then to the real DEPOT name (never the zone).
                    raw_mkt = mpo_market_map.get(m_code) or _clean_cell(r["depot"]) or "Unknown"

                data["top_50_mpos"].append({
                    "rank": idx,
                    "mpo_code": m_code,
                    "market": raw_mkt,
                    "zone": r["zone"] or "Unknown",
                    "depot": r["depot"] or "Unknown",
                    "total_sales": round(m_sales, 2),
                    "total_quantity": round(float(r["total_quantity"] or 0), 2),
                    "total_invoices": int(r["total_invoices"] or 0),
                    "total_parties": int(r["total_parties"] or 0),
                    "contribution_pct": round((m_sales / total_sales) * 100, 2) if total_sales > 0 else 0,
                    "monthly_breakdown": m_breakdown
                })

            # 5. TOP 20 FM (Field Managers)
            cur.execute("""
                SELECT 
                    fm_am as fm_name,
                    SUM(line_amount) as total_sales,
                    COUNT(DISTINCT mpo_code) as active_mpos,
                    COUNT(DISTINCT invoice_no) as total_invoices,
                    COUNT(DISTINCT customer_id) as total_parties
                FROM sales
                WHERE fm_am IS NOT NULL AND fm_am != ''
                GROUP BY fm_am
                ORDER BY SUM(line_amount) DESC
                LIMIT 20
            """)
            for idx, r in enumerate(cur.fetchall(), 1):
                fm_sales = float(r["total_sales"] or 0)
                data["top_20_fms"].append({
                    "rank": idx,
                    "fm_name": r["fm_name"],
                    "total_sales": round(fm_sales, 2),
                    "active_mpos": int(r["active_mpos"] or 0),
                    "total_invoices": int(r["total_invoices"] or 0),
                    "total_parties": int(r["total_parties"] or 0),
                    "contribution_pct": round((fm_sales / total_sales) * 100, 2) if total_sales > 0 else 0
                })

            # 6. TOP 5 SECTOR HEADS / ZONES
            cur.execute("""
                SELECT 
                    zone as sector_name,
                    SUM(line_amount) as total_sales,
                    COUNT(DISTINCT mpo_code) as active_mpos,
                    COUNT(DISTINCT invoice_no) as total_invoices,
                    COUNT(DISTINCT customer_id) as total_parties
                FROM sales
                WHERE zone IS NOT NULL AND zone != ''
                GROUP BY zone
                ORDER BY SUM(line_amount) DESC
                LIMIT 5
            """)
            for idx, r in enumerate(cur.fetchall(), 1):
                sec_sales = float(r["total_sales"] or 0)
                data["top_5_sector_heads"].append({
                    "rank": idx,
                    "sector_name": r["sector_name"],
                    "total_sales": round(sec_sales, 2),
                    "active_mpos": int(r["active_mpos"] or 0),
                    "total_invoices": int(r["total_invoices"] or 0),
                    "total_parties": int(r["total_parties"] or 0),
                    "contribution_pct": round((sec_sales / total_sales) * 100, 2) if total_sales > 0 else 0
                })

            # 7. STRATEGIC 6 PRODUCTS (TOP 50 MPO HIERARCHY BY UNIT WITH MONTH-WISE DATA)
            distinct_months = [t["month"] for t in data["monthly_trends"]]
            
            for prod_name, codes in STRATEGIC_6_MAPPING.items():
                placeholders = ",".join(["?"] * len(codes))
                
                # Overall stats
                cur.execute(f"""
                    SELECT 
                        SUM(quantity) as tot_units,
                        SUM(line_amount) as tot_sales,
                        COUNT(DISTINCT customer_id) as tot_parties,
                        COUNT(DISTINCT invoice_no) as tot_invoices
                    FROM sales
                    WHERE product_code IN ({placeholders})
                """, codes)
                stat_row = cur.fetchone()
                
                # Top 50 MPOs overall sorted by UNIT DESC (Hierarchy by unit)
                cur.execute(f"""
                    SELECT 
                        mpo_code,
                        MAX(market) as market,
                        MAX(zone) as zone,
                        MAX(depot) as depot,
                        SUM(quantity) as units,
                        COUNT(DISTINCT customer_id) as parties,
                        COUNT(DISTINCT invoice_no) as invoices,
                        SUM(line_amount) as sales
                    FROM sales
                    WHERE product_code IN ({placeholders}) AND mpo_code IS NOT NULL AND mpo_code != ''
                    GROUP BY mpo_code
                    ORDER BY SUM(quantity) DESC
                    LIMIT 50
                """, codes)
                
                mpo_list_all = []
                for idx, mr in enumerate(cur.fetchall(), 1):
                    m_code = mr["mpo_code"]
                    
                    cur.execute(f"""
                        SELECT 
                            month,
                            SUM(quantity) as m_units,
                            COUNT(DISTINCT customer_id) as m_parties,
                            COUNT(DISTINCT invoice_no) as m_invoices,
                            SUM(line_amount) as m_sales
                        FROM sales
                        WHERE product_code IN ({placeholders}) AND mpo_code = ?
                        GROUP BY month
                        ORDER BY month ASC
                    """, (*codes, m_code))
                    m_break = []
                    for mb in cur.fetchall():
                        m_break.append({
                            "month": mb["month"],
                            "units": round(float(mb["m_units"] or 0), 2),
                            "parties": int(mb["m_parties"] or 0),
                            "invoices": int(mb["m_invoices"] or 0),
                            "sales": round(float(mb["m_sales"] or 0), 2)
                        })
                        
                    raw_mkt = mr["market"]
                    if not raw_mkt or str(raw_mkt).strip() in ['', 'None', 'Unknown'] or str(raw_mkt).strip().startswith('DK.'):
                        raw_mkt = mpo_market_map.get(m_code) or _clean_cell(mr["depot"]) or "Unknown"

                    mpo_list_all.append({
                        "rank": idx,
                        "mpo_code": m_code,
                        "market": raw_mkt,
                        "zone": mr["zone"] or "Unknown",
                        "depot": mr["depot"] or "Unknown",
                        "units": round(float(mr["units"] or 0), 2),
                        "parties": int(mr["parties"] or 0),
                        "invoices": int(mr["invoices"] or 0),
                        "sales": round(float(mr["sales"] or 0), 2),
                        "monthly_breakdown": m_break
                    })
                    
                # Month-wise Top 50 MPOs (by unit)
                mpo_by_month = {}
                for m_val in distinct_months:
                    cur.execute(f"""
                        SELECT 
                            mpo_code,
                            MAX(market) as market,
                            MAX(zone) as zone,
                            MAX(depot) as depot,
                            SUM(quantity) as units,
                            COUNT(DISTINCT customer_id) as parties,
                            COUNT(DISTINCT invoice_no) as invoices,
                            SUM(line_amount) as sales
                        FROM sales
                        WHERE product_code IN ({placeholders}) AND month = ? AND mpo_code IS NOT NULL AND mpo_code != ''
                        GROUP BY mpo_code
                        ORDER BY SUM(quantity) DESC
                        LIMIT 50
                    """, (*codes, m_val))
                    mpo_list_month = []
                    for idx, mr in enumerate(cur.fetchall(), 1):
                        raw_mkt = mr["market"]
                        if not raw_mkt or str(raw_mkt).strip() in ['', 'None', 'Unknown'] or str(raw_mkt).strip().startswith('DK.'):
                            raw_mkt = mpo_market_map.get(mr["mpo_code"]) or _clean_cell(mr["depot"]) or "Unknown"

                        mpo_list_month.append({
                            "rank": idx,
                            "mpo_code": mr["mpo_code"],
                            "market": raw_mkt,
                            "zone": mr["zone"] or "Unknown",
                            "depot": mr["depot"] or "Unknown",
                            "units": round(float(mr["units"] or 0), 2),
                            "parties": int(mr["parties"] or 0),
                            "invoices": int(mr["invoices"] or 0),
                            "sales": round(float(mr["sales"] or 0), 2)
                        })
                    mpo_by_month[m_val] = mpo_list_month
                    
                data["strategic_6_products"][prod_name] = {
                    "product_name": prod_name,
                    "merged_codes": codes,
                    "total_units": round(float(stat_row["tot_units"] or 0), 2),
                    "total_sales": round(float(stat_row["tot_sales"] or 0), 2),
                    "total_parties": int(stat_row["tot_parties"] or 0),
                    "total_invoices": int(stat_row["tot_invoices"] or 0),
                    "mpo_top50_all": mpo_list_all,
                    "mpo_top50_by_month": mpo_by_month
                }

        except Exception as e:
            print(f"Error executing SQL calculations: {e}")
            raise
        finally:
            conn.close()

        # Save to JSON for static / instant dashboard load
        out_file = os.path.join(DATA_OUT_DIR, "api_data.json")
        with open(out_file, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
            
        print(f"✓ Data generation complete! Saved snapshot to: {out_file}")
        return data

if __name__ == "__main__":
    engine = DataEngine()
    engine.generate_all_data()
