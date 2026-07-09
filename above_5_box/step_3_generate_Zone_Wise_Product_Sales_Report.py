import os
import re
import json
import glob
import threading
import math
import random
import gspread
import pandas as pd
from datetime import datetime
from google.oauth2.service_account import Credentials
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════
#  Design tokens & Constants
# ══════════════════════════════════════════════════════════════════
_C = {
    'void'   : '#03040A',
    'panel'  : '#060816',
    'panel2' : '#050912',
    'cyan'   : '#00F2FE',
    'mag'    : '#FF00AA',
    'violet' : '#9D00FF',
    'green'  : '#00FF88',
    'red'    : '#FF2244',
    'text'   : '#D8F0F8',
    'muted'  : '#5F98A7',
    'border' : '#102A45',
    'hot'    : '#0D2A3F',
}

W_WIN, H_WIN = 1180, 680

# ══════════════════════════════════════════════════════════════════
#  Helper functions
# ══════════════════════════════════════════════════════════════════
def get_brand_name(name):
    """Extract the first word as the brand name"""
    n = name.strip()
    match = re.match(r'^([A-Za-z]+)', n)
    if match:
        return match.group(1).upper()
    return n.upper()

def get_subgroup_name(name):
    """Clean name by stripping parentheticals, common dosage suffixes, and strengths"""
    n = name.strip()
    # Remove parentheses and brackets content (e.g. (F), (B), (NEW), [A])
    n = re.sub(r'\s*\([^)]*\)', '', n)
    n = re.sub(r'\s*\[[^\]]*\]', '', n)
    
    # Suffixes to strip
    suffixes = [
        r"'s\b",
        r"’s\b",
        r'\btab(?:let)?s?\b',
        r'\bcap(?:sule)?s?\b',
        r'\bsyrup\b',
        r'\bsusp(?:ension)?\b',
        r'\bpfs\b',
        r'\binj(?:ection)?\b',
        r'\bcream\b',
        r'\boint(?:ment)?\b',
        r'\bgel\b',
        r'\bmg\b',
        r'\bmcg\b',
        r'\bml\b',
        r'\bgm\b',
        r'\bg\b'
    ]
    for suffix in suffixes:
        n = re.sub(suffix, '', n, flags=re.IGNORECASE).strip()
    
    # Normalize spaces
    n = re.sub(r'\s+', ' ', n).strip()
    return n

def clean_subgroup_key(s):
    if not s:
        return ""
    # Strip dots, dashes, and spaces from start/end, and uppercase
    return s.strip().strip('.').strip('-').strip().upper()

# ══════════════════════════════════════════════════════════════════
#  Tkinter Scrollable Frame Component
# ══════════════════════════════════════════════════════════════════
class ScrollableFrame(tk.Frame):
    def __init__(self, container, *args, **kwargs):
        super().__init__(container, *args, **kwargs)

        self.canvas = tk.Canvas(self, bg=_C['void'], highlightthickness=0)
        self.scrollbar = tk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas, bg=_C['void'])

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(
                scrollregion=self.canvas.bbox("all")
            )
        )

        self.canvas_window = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        
        # Make the scrollable frame fill the canvas width
        self.canvas.bind('<Configure>', self._on_canvas_configure)

        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        
        # Bind mousewheel when mouse enters/leaves the scroll region
        self.canvas.bind("<Enter>", self._bind_all_wheel)
        self.canvas.bind("<Leave>", self._unbind_all_wheel)
        self.scrollable_frame.bind("<Enter>", self._bind_all_wheel)
        self.scrollable_frame.bind("<Leave>", self._unbind_all_wheel)

    def _bind_all_wheel(self, event):
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)

    def _unbind_all_wheel(self, event):
        self.canvas.unbind_all("<MouseWheel>")

    def _on_canvas_configure(self, event):
        self.canvas.itemconfig(self.canvas_window, width=event.width)

    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")

    def bind_widget_scroll(self, widget):
        """Recursively bind mousewheel event to a widget and all its children so scrolling works over Entry, Checkbutton, etc."""
        widget.bind("<MouseWheel>", self._on_mousewheel)
        widget.bind("<Enter>", self._bind_all_wheel)
        widget.bind("<Leave>", self._unbind_all_wheel)
        for child in widget.winfo_children():
            self.bind_widget_scroll(child)

# ══════════════════════════════════════════════════════════════════
#  Main GUI Application Class
# ══════════════════════════════════════════════════════════════════
class ZoneReportApp:
    def __init__(self, root):
        self.root = root
        self.root.title("ALCO ZONE-WISE REPORT GENERATOR")
        self.root.geometry(f"{W_WIN}x{H_WIN}")
        self.root.resizable(False, False)
        self.root.configure(bg=_C['void'])

        # State Variables
        self.input_file = tk.StringVar()
        self.output_dir = tk.StringVar()
        self.search_var = tk.StringVar()
        
        self.opt_group_brand = tk.BooleanVar(value=False)
        self.opt_group_subgroup = tk.BooleanVar(value=True)
        self.opt_summary_end = tk.BooleanVar(value=False)
        self.opt_only_summary = tk.BooleanVar(value=True)
        self.opt_exclude_vacant = tk.BooleanVar(value=True)
        
        self.products_data = [] # List of dicts: {'code': str, 'default_name': str, 'var_select': BooleanVar, 'var_name': StringVar, 'var_subgroup': StringVar, 'var_is_subgrouped': BooleanVar, 'row_frame': Frame}
        self.unique_months = []
        
        self._t = 0.0
        self._scanline_y = 0.0
        self._stars = []
        self._after_id = None

        self._build()
        self._seed_stars()
        self._tick()
        
        # Bind search box modification
        self.search_var.trace_add("write", self.filter_products)
        
        # Auto-detect latest Product_Level_Net_Sales CSV on startup
        self.auto_detect_latest_csv()

    def _build(self):
        W, H = W_WIN, H_WIN

        # Canvas for background decoration
        self.cv = tk.Canvas(self.root, width=W, height=H, bg=_C['void'], highlightthickness=0, bd=0)
        self.cv.place(x=0, y=0)

        self._paint_gradient(W, H)
        self._paint_grid(W, H)
        self._paint_brackets(W, H)

        # Scanline item
        self._sl = self.cv.create_line(0, 0, W, 0, fill='#00F2FE', width=1, stipple='gray12')

        # Title
        self.cv.create_text(W//2, 24, text='ALCO ZONE-WISE SALES ENGINE', font=('Courier New', 14, 'bold'), fill=_C['cyan'])
        self.cv.create_text(W//2, 41, text='∞  MPO & BRAND SALES ANALYSIS SYSTEM  ∞', font=('Courier New', 7), fill=_C['muted'])
        self.cv.create_line(28, 54, W-28, 54, fill=_C['border'], width=1)
        self.cv.create_oval(23, 49, 33, 59, fill=_C['cyan'], outline='')
        self.cv.create_oval(W-33, 49, W-23, 59, fill=_C['violet'], outline='')

        # Panel 1: Left Frame (inputs and options)
        self.left_frame = tk.Frame(self.root, bg=_C['panel'], bd=1, relief='flat')
        self.left_frame.place(x=28, y=70, width=350, height=560)
        self._draw_panel_border(28, 70, 378, 630)
        self._build_left_panel()

        # Panel 2: Middle Frame (product selection list)
        self.middle_frame = tk.Frame(self.root, bg=_C['panel'], bd=1, relief='flat')
        self.middle_frame.place(x=398, y=70, width=390, height=560)
        self._draw_panel_border(398, 70, 788, 630)
        self._build_middle_panel()

        # Panel 3: Right Frame (selected products summary)
        self.right_frame = tk.Frame(self.root, bg=_C['panel'], bd=1, relief='flat')
        self.right_frame.place(x=808, y=70, width=344, height=560)
        self._draw_panel_border(808, 70, 1152, 630)
        self._build_right_panel()

    def _draw_panel_border(self, x1, y1, x2, y2):
        self.cv.create_rectangle(x1, y1, x2, y2, outline=_C['border'], width=1)

    def _build_left_panel(self):
        # Header inside panel
        lbl_cfg = tk.Label(self.left_frame, text="[ CONFIGURATION & OPERATION ]", font=('Courier New', 9, 'bold'), fg=_C['cyan'], bg=_C['panel'])
        lbl_cfg.pack(anchor="w", padx=15, pady=(15, 5))

        # File selection message label
        self.lbl_file_msg = tk.Label(self.left_frame, text="", font=('Courier New', 7, 'bold'), fg=_C['green'], bg=_C['panel'], wraplength=310, justify="left")
        self.lbl_file_msg.pack(anchor="w", padx=15, pady=(5, 0))

        # Input File
        lbl_in = tk.Label(self.left_frame, text="INPUT SALES CSV", font=('Courier New', 8, 'bold'), fg=_C['muted'], bg=_C['panel'])
        lbl_in.pack(anchor="w", padx=15, pady=(10, 2))
        
        in_entry_frame = tk.Frame(self.left_frame, bg=_C['panel2'])
        in_entry_frame.pack(fill="x", padx=15, pady=2)
        
        self.ent_in = tk.Entry(in_entry_frame, textvariable=self.input_file, bg=_C['panel2'], fg=_C['text'],
                               insertbackground=_C['cyan'], relief='flat', font=('Courier New', 8), bd=1)
        self.ent_in.pack(side="left", fill="x", expand=True, padx=5, pady=4)
        
        btn_in = tk.Button(in_entry_frame, text="BROWSE", command=self.browse_input, font=('Courier New', 7, 'bold'),
                           fg=_C['cyan'], bg='#07152B', relief='flat', cursor='hand2', activebackground='#0B2A4A')
        btn_in.pack(side="right", padx=3, pady=3)

        # Output Folder
        lbl_out = tk.Label(self.left_frame, text="OUTPUT FOLDER", font=('Courier New', 8, 'bold'), fg=_C['muted'], bg=_C['panel'])
        lbl_out.pack(anchor="w", padx=15, pady=(10, 2))
        
        out_entry_frame = tk.Frame(self.left_frame, bg=_C['panel2'])
        out_entry_frame.pack(fill="x", padx=15, pady=2)
        
        self.ent_out = tk.Entry(out_entry_frame, textvariable=self.output_dir, bg=_C['panel2'], fg=_C['text'],
                                insertbackground=_C['cyan'], relief='flat', font=('Courier New', 8), bd=1)
        self.ent_out.pack(side="left", fill="x", expand=True, padx=5, pady=4)
        
        btn_out = tk.Button(out_entry_frame, text="BROWSE", command=self.browse_output, font=('Courier New', 7, 'bold'),
                            fg=_C['cyan'], bg='#07152B', relief='flat', cursor='hand2', activebackground='#0B2A4A')
        btn_out.pack(side="right", padx=3, pady=3)

        # Divider
        div = tk.Frame(self.left_frame, height=1, bg=_C['border'])
        div.pack(fill="x", padx=15, pady=15)

        # Switches Header
        lbl_sw = tk.Label(self.left_frame, text="[ EXTRACTION SWITCHES ]", font=('Courier New', 9, 'bold'), fg=_C['cyan'], bg=_C['panel'])
        lbl_sw.pack(anchor="w", padx=15, pady=(5, 5))

        # Checkbox 1: Group by Brand Family
        self.cb_group = tk.Checkbutton(self.left_frame, text="Group by Brand Family (Sum Values)",
                                       variable=self.opt_group_brand, onvalue=True, offvalue=False,
                                       font=('Courier New', 8), fg=_C['text'], bg=_C['panel'],
                                       activebackground=_C['panel'], activeforeground=_C['cyan'],
                                       selectcolor=_C['void'], command=lambda: self.toggle_grouping_mode('brand'))
        self.cb_group.pack(anchor="w", padx=20, pady=4)

        # Checkbox 1b: Group by Sub-Group
        self.cb_subgroup = tk.Checkbutton(self.left_frame, text="Group by Sub-Group (Sum Values)",
                                          variable=self.opt_group_subgroup, onvalue=True, offvalue=False,
                                          font=('Courier New', 8), fg=_C['text'], bg=_C['panel'],
                                          activebackground=_C['panel'], activeforeground=_C['cyan'],
                                          selectcolor=_C['void'], command=lambda: self.toggle_grouping_mode('subgroup'))
        self.cb_subgroup.pack(anchor="w", padx=20, pady=4)

        # Checkbox 2: Summary Column at the End
        self.cb_end = tk.Checkbutton(self.left_frame, text="Add Group Summary Column at End",
                                     variable=self.opt_summary_end, onvalue=True, offvalue=False,
                                     font=('Courier New', 8), fg=_C['text'], bg=_C['panel'],
                                     activebackground=_C['panel'], activeforeground=_C['cyan'],
                                     selectcolor=_C['void'])
        self.cb_end.pack(anchor="w", padx=20, pady=4)

        # Checkbox 3: Show Only Summarized Columns
        self.cb_only = tk.Checkbutton(self.left_frame, text="Show ONLY Summarized Group Columns",
                                      variable=self.opt_only_summary, onvalue=True, offvalue=False,
                                      font=('Courier New', 8), fg=_C['text'], bg=_C['panel'],
                                      activebackground=_C['panel'], activeforeground=_C['cyan'],
                                      selectcolor=_C['void'])
        self.cb_only.pack(anchor="w", padx=20, pady=4)

        # Checkbox 4: Exclude Vacant Markets
        self.cb_vacant = tk.Checkbutton(self.left_frame, text="Exclude Vacant Markets (Y in VACANT column)",
                                      variable=self.opt_exclude_vacant, onvalue=True, offvalue=False,
                                      font=('Courier New', 8), fg=_C['text'], bg=_C['panel'],
                                      activebackground=_C['panel'], activeforeground=_C['cyan'],
                                      selectcolor=_C['void'])
        self.cb_vacant.pack(anchor="w", padx=20, pady=4)

        # Spacer
        spacer = tk.Frame(self.left_frame, bg=_C['panel'])
        spacer.pack(fill="both", expand=True)

        # Status and Progress Bar inside Left Panel
        self.lbl_status = tk.Label(self.left_frame, text="◉ SYSTEM READY", font=('Courier New', 8), fg=_C['cyan'], bg=_C['panel'])
        self.lbl_status.pack(anchor="w", padx=15, pady=(5, 2))
        
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(self.left_frame, variable=self.progress_var, maximum=100, mode='determinate')
        self.progress_bar.pack(fill="x", padx=15, pady=(2, 15))

        # Execute Button
        self.exec_btn = tk.Button(self.left_frame, text="⟨  AUTOMATE FORMATTING  ⟩", command=self.run_process,
                                  font=('Courier New', 10, 'bold'), fg=_C['cyan'], bg='#030A1C', relief='flat',
                                  cursor='hand2', activebackground='#061C34', activeforeground='#FFFFFF', bd=1,
                                  highlightbackground=_C['cyan'])
        self.exec_btn.pack(fill="x", padx=15, pady=(0, 20))

    def _build_middle_panel(self):
        # Header
        lbl_p = tk.Label(self.middle_frame, text="[ PRODUCT CODES SELECTION & RENAMING ]", font=('Courier New', 9, 'bold'), fg=_C['cyan'], bg=_C['panel'])
        lbl_p.pack(anchor="w", padx=15, pady=(15, 5))

        # Search Bar
        search_frame = tk.Frame(self.middle_frame, bg=_C['panel2'])
        search_frame.pack(fill="x", padx=15, pady=5)
        
        lbl_search = tk.Label(search_frame, text="🔍", font=('Segoe UI', 9), fg=_C['muted'], bg=_C['panel2'])
        lbl_search.pack(side="left", padx=5)
        
        self.ent_search = tk.Entry(search_frame, textvariable=self.search_var, bg=_C['panel2'], fg=_C['text'],
                                   insertbackground=_C['cyan'], relief='flat', font=('Courier New', 8), bd=1)
        self.ent_search.pack(side="left", fill="x", expand=True, padx=5, pady=4)

        # Select All / Deselect All
        ctrl_frame = tk.Frame(self.middle_frame, bg=_C['panel'])
        ctrl_frame.pack(fill="x", padx=15, pady=5)
        
        btn_sel_all = tk.Button(ctrl_frame, text="SELECT ALL", command=self.select_all_products, font=('Courier New', 7, 'bold'),
                                fg=_C['cyan'], bg='#07152B', relief='flat', cursor='hand2')
        btn_sel_all.pack(side="left", padx=(0, 10))

        btn_desel_all = tk.Button(ctrl_frame, text="DESELECT ALL", command=self.deselect_all_products, font=('Courier New', 7, 'bold'),
                                  fg=_C['cyan'], bg='#07152B', relief='flat', cursor='hand2')
        btn_desel_all.pack(side="left")

        # Column Labels
        labels_frame = tk.Frame(self.middle_frame, bg=_C['panel'])
        labels_frame.pack(fill="x", padx=15, pady=(5, 2))
        lbl_c = tk.Label(labels_frame, text="CODE", font=('Courier New', 7, 'bold'), fg=_C['muted'], bg=_C['panel'], width=10, anchor='w')
        lbl_c.pack(side="left", padx=(5, 2))
        lbl_n = tk.Label(labels_frame, text="NAME", font=('Courier New', 7, 'bold'), fg=_C['muted'], bg=_C['panel'], width=18, anchor='w')
        lbl_n.pack(side="left", padx=5)
        lbl_s = tk.Label(labels_frame, text="SUB-GROUP", font=('Courier New', 7, 'bold'), fg=_C['muted'], bg=_C['panel'], width=12, anchor='w')
        lbl_s.pack(side="left", padx=5)
        lbl_g = tk.Label(labels_frame, text="GRP?", font=('Courier New', 7, 'bold'), fg=_C['muted'], bg=_C['panel'], width=5, anchor='w')
        lbl_g.pack(side="left", padx=(2, 5))

        # Scrollable Product Container
        self.prod_scroll = ScrollableFrame(self.middle_frame)
        self.prod_scroll.pack(fill="both", expand=True, padx=15, pady=(0, 15))

    def _build_right_panel(self):
        # Header
        lbl_selected = tk.Label(self.right_frame, text="[ SELECTED PRODUCTS AT A GLANCE ]", font=('Courier New', 9, 'bold'), fg=_C['cyan'], bg=_C['panel'])
        lbl_selected.pack(anchor="w", padx=15, pady=(15, 12))

        # Scrollable summary container
        self.selected_scroll = ScrollableFrame(self.right_frame)
        self.selected_scroll.pack(fill="both", expand=True, padx=15, pady=(0, 15))

    def update_switch_states(self):
        """Enable/Disable sub-options based on Grouping switches"""
        is_grouped = self.opt_group_brand.get() or self.opt_group_subgroup.get()
        if is_grouped:
            self.cb_end.configure(state='normal')
            self.cb_only.configure(state='normal')
        else:
            self.cb_end.configure(state='disabled')
            self.cb_only.configure(state='disabled')
            self.opt_summary_end.set(False)
            self.opt_only_summary.set(False)

    def toggle_grouping_mode(self, mode):
        if mode == 'brand':
            if self.opt_group_brand.get():
                self.opt_group_subgroup.set(False)
        elif mode == 'subgroup':
            if self.opt_group_subgroup.get():
                self.opt_group_brand.set(False)
        self.update_switch_states()
        self.update_selected_summary()

    # ── Background styling ───────────────────────────────────────────
    def _paint_gradient(self, W, H):
        for y in range(H):
            t = y / H
            r = int(3  + t * 4)
            g = int(4  + t * 5)
            b = int(10 + t * 16)
            self.cv.create_line(0, y, W, y, fill=f'#{r:02x}{g:02x}{b:02x}', width=1)

    def _paint_grid(self, W, H):
        for x in range(0, W, 40):
            self.cv.create_line(x, 0, x, H, fill='#07091A', width=1)
        for y in range(0, H, 40):
            self.cv.create_line(0, y, W, y, fill='#07091A', width=1)

    def _paint_brackets(self, W, H, pad=10, sz=18):
        c = _C['cyan']
        # TL
        self.cv.create_line(pad, pad, pad+sz, pad, fill=c, width=1)
        self.cv.create_line(pad, pad, pad, pad+sz, fill=c, width=1)
        # TR
        self.cv.create_line(W-pad-sz, pad, W-pad, pad, fill=c, width=1)
        self.cv.create_line(W-pad, pad, W-pad, pad+sz, fill=c, width=1)
        # BL
        self.cv.create_line(pad, H-pad, pad+sz, H-pad, fill=c, width=1)
        self.cv.create_line(pad, H-pad-sz, pad, H-pad, fill=c, width=1)
        # BR
        self.cv.create_line(W-pad-sz, H-pad, W-pad, H-pad, fill=c, width=1)
        self.cv.create_line(W-pad, H-pad-sz, W-pad, H-pad, fill=c, width=1)

    def _tick(self):
        self._t += 0.045
        self._scanline_y = (self._scanline_y + 1.8) % H_WIN
        self.cv.coords(self._sl, 0, self._scanline_y, W_WIN, self._scanline_y)

        # Pulse status light
        status_text = self.lbl_status.cget('text')
        if 'READY' in status_text:
            v = int(184 + 56 * math.sin(self._t))
            self.lbl_status.configure(fg=f'#00{v:02x}{v//2:02x}')

        self._after_id = self.root.after(50, self._tick)

    # ── Star decoration ──────────────────────────────────────────────
    def _seed_stars(self):
        colors = [_C['cyan'], '#FFFFFF', _C['violet'], '#4FACFE']
        for _ in range(35):
            x = random.randint(42, W_WIN-42)
            y = random.randint(58, H_WIN-30)
            sz = random.choice([1, 1, 2])
            sid = self.cv.create_oval(x, y, x+sz, y+sz, fill=random.choice(colors), outline='', state='hidden')
            self._stars.append({
                'id'  : sid,
                'ph'  : random.uniform(0, 2*math.pi),
                'spd' : random.uniform(0.025, 0.065),
            })

    # ── Auto detect latest CSV ──────────────────────────────────────────
    def auto_detect_latest_csv(self):
        csv_files = glob.glob(r'c:\Users\Irak\Desktop\Barishal April Data\*Product_Level_Net_Sales*.csv')
        if csv_files:
            latest_csv = max(csv_files, key=os.path.getmtime)
            self.input_file.set(latest_csv)
            self.output_dir.set(os.path.dirname(latest_csv))
            self.lbl_file_msg.configure(text="✓ LATEST WORKFLOW FILE AUTO-SELECTED", fg=_C['green'])
            self.load_products_from_csv(latest_csv)

    # ── Browse File Functions ───────────────────────────────────────────
    def browse_input(self):
        f = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
        if f:
            self.input_file.set(f)
            if not self.output_dir.get():
                self.output_dir.set(os.path.dirname(f))
            self.lbl_file_msg.configure(text="⚠ MANUAL SELECTION ACTIVE", fg=_C['mag'])
            self.load_products_from_csv(f)

    def browse_output(self):
        d = filedialog.askdirectory()
        if d:
            self.output_dir.set(d)

    # ── Load Products from CSV ──────────────────────────────────────────
    def load_products_from_csv(self, csv_path):
        # Clear existing middle pane
        for widget in self.prod_scroll.scrollable_frame.winfo_children():
            widget.destroy()
        self.products_data.clear()
        self.search_var.set("") # Reset search query

        try:
            self._updating_batch = True
            df = pd.read_csv(csv_path)
            if 'Product_Code' not in df.columns or 'Product_Name' not in df.columns:
                messagebox.showerror("Format Error", "CSV file must contain 'Product_Code' and 'Product_Name' columns.")
                return

            # Clean and get unique codes and their most frequent names
            df['Product_Code'] = df['Product_Code'].astype(str).str.strip().str.upper()
            df['Product_Name'] = df['Product_Name'].astype(str).str.strip()
            
            prod_name_counts = df.groupby(['Product_Code', 'Product_Name']).size().reset_index(name='count')
            idx_max = prod_name_counts.groupby('Product_Code')['count'].idxmax()
            standard_names = prod_name_counts.loc[idx_max, ['Product_Code', 'Product_Name']]
            
            # Load subgroup mapping and standard name mapping from Google Sheet (gid=1219133636) or fallback to excel file
            subgroup_mapping = {}
            standard_name_mapping = {}
            try:
                print("Loading Product Code mapping from Google Sheet (gid=1219133636)...")
                import sys, os
                sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..'))
                from googleDrive.credentials_loader import get_sheet_service_account_credentials, get_spreadsheet_id
                scopes = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
                creds = get_sheet_service_account_credentials(scopes=scopes)
                client = gspread.authorize(creds)
                sheet = client.open_by_key(get_spreadsheet_id('master_field_force_sheet') or '1Q4utivZ5OpgDznqlqElYU-HWNnZYI71YYpcZKcSM3xY')
                ws_prod = sheet.get_worksheet_by_id(1219133636)
                if ws_prod:
                    rows_prod = ws_prod.get_all_records()
                    for r in rows_prod:
                        pcode_all = str(r.get('PRODUCT_CODE_ALL_ROW', '')).strip().upper()
                        subg = str(r.get('SUB_GROUP_STANDARD', '')).strip()
                        if pcode_all and pcode_all != 'NAN' and subg and subg != 'NAN':
                            subgroup_mapping[pcode_all] = subg
                        pcode = str(r.get('Product_Code', '')).strip().upper()
                        pname = str(r.get('Product_Name', '')).strip()
                        if pcode and pcode != 'NAN' and pname and pname != 'NAN':
                            standard_name_mapping[pcode] = pname
                    print(f"Loaded {len(subgroup_mapping)} subgroups and {len(standard_name_mapping)} product names from Google Sheet.")
            except Exception as ex:
                print(f"Note: Could not load Product mapping from Google Sheet ({str(ex)[:80]}). Using resilient fallback...")

            if not subgroup_mapping:
                excel_path = None
                paths_to_try = [
                    os.path.join(os.path.dirname(os.path.abspath(__file__)), 'PRODUCT_CODE_AND_SUBGROUP_OF_PRODUCTS.xlsx'),
                    os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'PRODUCT_CODE_AND_SUBGROUP_OF_PRODUCTS.xlsx'),
                    r'c:\Users\Irak\Desktop\Barishal April Data\PRODUCT_CODE_AND_SUBGROUP_OF_PRODUCTS.xlsx'
                ]
                for p in paths_to_try:
                    if os.path.exists(p):
                        excel_path = p
                        break
                if excel_path:
                    try:
                        df_excel = pd.read_excel(excel_path)
                        
                        # 1. Map PRODUCT_CODE_ALL_ROW -> SUB_GROUP_STANDARD (takes the first unique match)
                        df_sub = df_excel.dropna(subset=['PRODUCT_CODE_ALL_ROW', 'SUB_GROUP_STANDARD']).copy()
                        df_sub['PRODUCT_CODE_ALL_ROW'] = df_sub['PRODUCT_CODE_ALL_ROW'].astype(str).str.strip().str.upper()
                        df_sub['SUB_GROUP_STANDARD'] = df_sub['SUB_GROUP_STANDARD'].astype(str).str.strip()
                        subgroup_mapping = dict(zip(df_sub['PRODUCT_CODE_ALL_ROW'], df_sub['SUB_GROUP_STANDARD']))
                        
                        # 2. Map Product_Code -> Product_Name (from rows where Product_Code is NOT null)
                        df_name = df_excel.dropna(subset=['Product_Code', 'Product_Name']).copy()
                        df_name['Product_Code'] = df_name['Product_Code'].astype(str).str.strip().str.upper()
                        df_name['Product_Name'] = df_name['Product_Name'].astype(str).str.strip()
                        standard_name_mapping = dict(zip(df_name['Product_Code'], df_name['Product_Name']))
                    except Exception as ex:
                        print(f"Error loading Excel mapping: {ex}")

            # Sort products by code
            sorted_prods = standard_names.sort_values(by='Product_Code')
            
            # Populate scroll frame
            for idx, row in sorted_prods.iterrows():
                code = row['Product_Code']
                def_name = row['Product_Name']
                
                # Check Excel mappings first, fallback to CSV values
                standard_name_val = standard_name_mapping.get(code, def_name)
                subgroup_val = subgroup_mapping.get(code, get_subgroup_name(def_name))
                
                var_select = tk.BooleanVar(value=True) # Selected by default
                var_name = tk.StringVar(value=standard_name_val)
                var_subgroup = tk.StringVar(value=subgroup_val)
                var_is_subgrouped = tk.BooleanVar(value=True)
                
                # Draw row frame (hidden by default until packed)
                row_frame = tk.Frame(self.prod_scroll.scrollable_frame, bg=_C['void'])
                row_frame.pack(fill="x", pady=2, padx=5)
                
                cb = tk.Checkbutton(row_frame, text=code, variable=var_select,
                                    font=('Courier New', 8, 'bold'), fg=_C['cyan'], bg=_C['void'],
                                    activebackground=_C['void'], activeforeground=_C['cyan'],
                                    selectcolor=_C['panel'])
                cb.pack(side="left", padx=(5, 2))
                
                ent_name = tk.Entry(row_frame, textvariable=var_name, font=('Segoe UI', 8), width=18,
                               bg=_C['panel'], fg=_C['text'], insertbackground=_C['cyan'],
                               relief='flat', bd=1, highlightthickness=1, highlightcolor=_C['border'],
                               highlightbackground='#09182C')
                ent_name.pack(side="left", padx=5, pady=1)

                ent_sub = tk.Entry(row_frame, textvariable=var_subgroup, font=('Segoe UI', 8), width=12,
                               bg=_C['panel'], fg=_C['text'], insertbackground=_C['cyan'],
                               relief='flat', bd=1, highlightthickness=1, highlightcolor=_C['border'],
                               highlightbackground='#09182C')
                ent_sub.pack(side="left", padx=5, pady=1)

                cb_grp = tk.Checkbutton(row_frame, text="Grp", variable=var_is_subgrouped,
                                        font=('Courier New', 7), fg=_C['muted'], bg=_C['void'],
                                        activebackground=_C['void'], activeforeground=_C['cyan'],
                                        selectcolor=_C['panel'])
                cb_grp.pack(side="left", padx=(2, 5))

                self.products_data.append({
                    'code': code,
                    'default_name': def_name,
                    'var_select': var_select,
                    'var_name': var_name,
                    'var_subgroup': var_subgroup,
                    'var_is_subgrouped': var_is_subgrouped,
                    'row_frame': row_frame
                })

                # Bind scroll recursively for this row
                self.prod_scroll.bind_widget_scroll(row_frame)

                # Trace changes to update selected summary automatically
                var_select.trace_add("write", lambda *args: self.update_selected_summary())
                var_name.trace_add("write", lambda *args: self.update_selected_summary())
                var_subgroup.trace_add("write", lambda *args: self.update_selected_summary())
                var_is_subgrouped.trace_add("write", lambda *args: self.update_selected_summary())
                
            self._updating_batch = False
            self.update_selected_summary()

        except Exception as e:
            messagebox.showerror("Error Reading CSV", f"Failed to load products:\n{e}")

    # ── Filter / Search Products ───────────────────────────────────────
    def filter_products(self, *args):
        query = self.search_var.get().strip().lower()
        for item in self.products_data:
            code = item['code'].lower()
            name = item['var_name'].get().lower()
            subgroup = item['var_subgroup'].get().lower()
            if not query or query in code or query in name or query in subgroup:
                item['row_frame'].pack(fill="x", pady=2, padx=5)
            else:
                item['row_frame'].pack_forget()

    # ── Update Selected Summary at a Glance ────────────────────────────
    def update_selected_summary(self):
        if getattr(self, '_updating_batch', False):
            return
        # Clear selected summary container
        for widget in self.selected_scroll.scrollable_frame.winfo_children():
            widget.destroy()

        checked = [p for p in self.products_data if p['var_select'].get()]

        lbl_total = tk.Label(self.selected_scroll.scrollable_frame, text=f"Total Selected: {len(checked)} / {len(self.products_data)}",
                             font=('Courier New', 8, 'bold'), fg=_C['green'], bg=_C['void'])
        lbl_total.pack(anchor="w", padx=10, pady=(5, 10))

        if self.opt_group_brand.get():
            # Group by Brand Family
            from collections import defaultdict
            groups = defaultdict(list)
            for p in checked:
                brand = get_brand_name(p['var_name'].get())
                groups[brand].append(p)
                
            for grp_name in sorted(groups.keys()):
                # Draw group header
                lbl_h = tk.Label(self.selected_scroll.scrollable_frame, text=f"■ {grp_name}",
                                 font=('Courier New', 8, 'bold'), fg=_C['cyan'], bg=_C['void'])
                lbl_h.pack(anchor="w", padx=10, pady=(5, 2))
                self.selected_scroll.bind_widget_scroll(lbl_h)
                
                for p in groups[grp_name]:
                    item_frame = tk.Frame(self.selected_scroll.scrollable_frame, bg=_C['void'])
                    item_frame.pack(fill="x", pady=1, padx=20)
                    
                    lbl_c = tk.Label(item_frame, text=f"{p['code']}:", font=('Courier New', 8), fg=_C['muted'], bg=_C['void'])
                    lbl_c.pack(side="left")
                    
                    lbl_n = tk.Label(item_frame, textvariable=p['var_name'], font=('Segoe UI', 8), fg=_C['text'], bg=_C['void'])
                    lbl_n.pack(side="left", fill="x", expand=True, padx=5, anchor="w")
                    self.selected_scroll.bind_widget_scroll(item_frame)

        elif self.opt_group_subgroup.get():
            # Group by Sub-Group
            from collections import defaultdict
            groups = defaultdict(list)
            ungrouped = []
            for p in checked:
                if p['var_is_subgrouped'].get():
                    sub_name = clean_subgroup_key(p['var_subgroup'].get())
                    if not sub_name:
                        sub_name = "UNNAMED SUBGROUP"
                    groups[sub_name].append(p)
                else:
                    ungrouped.append(p)
            
            # Draw grouped ones
            for grp_name in sorted(groups.keys()):
                lbl_h = tk.Label(self.selected_scroll.scrollable_frame, text=f"⧉ {grp_name}",
                                 font=('Courier New', 8, 'bold'), fg=_C['cyan'], bg=_C['void'])
                lbl_h.pack(anchor="w", padx=10, pady=(5, 2))
                self.selected_scroll.bind_widget_scroll(lbl_h)
                
                for p in groups[grp_name]:
                    item_frame = tk.Frame(self.selected_scroll.scrollable_frame, bg=_C['void'])
                    item_frame.pack(fill="x", pady=1, padx=20)
                    
                    lbl_c = tk.Label(item_frame, text=f"{p['code']}:", font=('Courier New', 8), fg=_C['muted'], bg=_C['void'])
                    lbl_c.pack(side="left")
                    
                    lbl_n = tk.Label(item_frame, textvariable=p['var_name'], font=('Segoe UI', 8), fg=_C['text'], bg=_C['void'])
                    lbl_n.pack(side="left", fill="x", expand=True, padx=5, anchor="w")
                    self.selected_scroll.bind_widget_scroll(item_frame)
            
            # Draw ungrouped ones
            if ungrouped:
                lbl_h = tk.Label(self.selected_scroll.scrollable_frame, text="⧉ [UNGROUPED]",
                                 font=('Courier New', 8, 'bold'), fg=_C['mag'], bg=_C['void'])
                lbl_h.pack(anchor="w", padx=10, pady=(5, 2))
                self.selected_scroll.bind_widget_scroll(lbl_h)
                
                for p in ungrouped:
                    item_frame = tk.Frame(self.selected_scroll.scrollable_frame, bg=_C['void'])
                    item_frame.pack(fill="x", pady=1, padx=20)
                    
                    lbl_c = tk.Label(item_frame, text=f"{p['code']}:", font=('Courier New', 8), fg=_C['muted'], bg=_C['void'])
                    lbl_c.pack(side="left")
                    
                    lbl_n = tk.Label(item_frame, textvariable=p['var_name'], font=('Segoe UI', 8), fg=_C['text'], bg=_C['void'])
                    lbl_n.pack(side="left", fill="x", expand=True, padx=5, anchor="w")
                    self.selected_scroll.bind_widget_scroll(item_frame)
        else:
            # Individual list
            for p in checked:
                item_frame = tk.Frame(self.selected_scroll.scrollable_frame, bg=_C['void'])
                item_frame.pack(fill="x", pady=1, padx=10)

                lbl_c = tk.Label(item_frame, text=f"[{p['code']}]", font=('Courier New', 8, 'bold'), fg=_C['cyan'], bg=_C['void'])
                lbl_c.pack(side="left")

                lbl_n = tk.Label(item_frame, textvariable=p['var_name'], font=('Segoe UI', 8), fg=_C['text'], bg=_C['void'], anchor="w", justify="left")
                lbl_n.pack(side="left", fill="x", expand=True, padx=8)
                self.selected_scroll.bind_widget_scroll(item_frame)

    # ── Select / Deselect All ──────────────────────────────────────────
    def select_all_products(self):
        self._updating_batch = True
        for p in self.products_data:
            # Only affect visible products to align with user filters
            if p['row_frame'].winfo_manager() == 'pack':
                p['var_select'].set(True)
        self._updating_batch = False
        self.update_selected_summary()

    def deselect_all_products(self):
        self._updating_batch = True
        for p in self.products_data:
            # Only affect visible products to align with user filters
            if p['row_frame'].winfo_manager() == 'pack':
                p['var_select'].set(False)
        self._updating_batch = False
        self.update_selected_summary()

    # ── Progress setter ────────────────────────────────────────────────
    def set_progress(self, pct: float, status: str = 'PROCESSING'):
        self.progress_var.set(pct)
        if pct == 0:
            self.lbl_status.configure(text='◉  SYSTEM READY', fg=_C['cyan'])
        elif pct >= 100:
            self.lbl_status.configure(text='◉  SEQUENCE COMPLETE', fg=_C['green'])
        else:
            self.lbl_status.configure(text=f'◉  {status}', fg=_C['cyan'])
        self.root.update_idletasks()

    # ── Success Dialog ────────────────────────────────────────────────
    def show_success_dialog(self, out_path):
        dialog = tk.Toplevel(self.root)
        dialog.title("Success")
        dialog.geometry("420x200")
        dialog.resizable(False, False)
        dialog.configure(bg=_C['void'])
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center relative to parent
        parent_x = self.root.winfo_rootx()
        parent_y = self.root.winfo_rooty()
        dialog.geometry(f"+{parent_x + 380}+{parent_y + 240}")

        W_D, H_D = 420, 200
        cv = tk.Canvas(dialog, width=W_D, height=H_D, bg=_C['void'], highlightthickness=0, bd=0)
        cv.place(x=0, y=0)

        # Background decoration
        for y in range(H_D):
            t = y / H_D
            r = int(3  + t * 4)
            g = int(4  + t * 5)
            b = int(10 + t * 16)
            cv.create_line(0, y, W_D, y, fill=f'#{r:02x}{g:02x}{b:02x}', width=1)

        # Brackets
        c = _C['cyan']
        pad, sz = 10, 12
        cv.create_line(pad, pad, pad+sz, pad, fill=c, width=1)
        cv.create_line(pad, pad, pad, pad+sz, fill=c, width=1)
        cv.create_line(W_D-pad-sz, pad, W_D-pad, pad, fill=c, width=1)
        cv.create_line(W_D-pad, pad, W_D-pad, pad+sz, fill=c, width=1)
        cv.create_line(pad, H_D-pad, pad+sz, H_D-pad, fill=c, width=1)
        cv.create_line(pad, H_D-pad-sz, pad, H_D-pad, fill=c, width=1)
        cv.create_line(W_D-pad-sz, H_D-pad, W_D-pad, H_D-pad, fill=c, width=1)
        cv.create_line(W_D-pad, H_D-pad-sz, W_D-pad, H_D-pad, fill=c, width=1)

        # Circle tick
        cv.create_oval(25, 30, 55, 60, fill='#051A18', outline=_C['green'], width=1.5)
        cv.create_text(40, 45, text='✓', font=('Courier New', 13, 'bold'), fill=_C['green'])

        # Title
        cv.create_text(70, 45, text="PROCESSING COMPLETE", font=('Courier New', 10, 'bold'), fill=_C['green'], anchor='w')

        norm_path = os.path.normpath(out_path)
        wrapped_path = norm_path
        if len(norm_path) > 45:
            chunks = []
            for i in range(0, len(norm_path), 45):
                chunks.append(norm_path[i:i+45])
            wrapped_path = "\n".join(chunks)

        cv.create_text(25, 80, text=f"Saved to:\n{wrapped_path}", font=('Courier New', 8), fill=_C['text'], anchor='nw')

        # Action handlers
        def open_file():
            try:
                os.startfile(norm_path)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to open file:\n{e}", parent=dialog)

        def open_folder():
            try:
                os.startfile(os.path.dirname(norm_path))
            except Exception as e:
                messagebox.showerror("Error", f"Failed to open folder:\n{e}", parent=dialog)

        # Draw buttons using Canvas items and Tkinter Buttons
        def make_btn(cx, cy, label, cmd, w=100, h=24, border_color=_C['cyan']):
            x1, y1, x2, y2 = cx-w//2, cy-h//2, cx+w//2, cy+h//2
            r = 4
            poly = [x1+r, y1, x2-r, y1, x2, y1, x2, y1+r, x2, y2-r, x2, y2, x2-r, y2, x1+r, y2, x1, y2, x1, y2-r, x1, y1+r, x1, y1]
            bg_id = cv.create_polygon(poly, smooth=True, fill='#07152B', outline=border_color, width=1)
            
            btn = tk.Button(dialog, text=label, command=cmd, font=('Courier New', 8, 'bold'), fg=border_color,
                            bg='#07152B', relief='flat', cursor='hand2', activebackground='#0B2A4A',
                            activeforeground='#FFFFFF', highlightthickness=0, bd=0)
            cv.create_window(cx, cy, window=btn, width=w-2, height=h-2)

            def on_in(e):
                cv.itemconfig(bg_id, fill=_C['hot'], outline='#FFFFFF')
                btn.configure(bg=_C['hot'], fg='#FFFFFF')
            def on_out(e):
                cv.itemconfig(bg_id, fill='#07152B', outline=border_color)
                btn.configure(bg='#07152B', fg=border_color)

            btn.bind('<Enter>', on_in)
            btn.bind('<Leave>', on_out)

        make_btn(90, 160, "OPEN FILE", open_file, border_color=_C['cyan'])
        make_btn(210, 160, "OPEN FOLDER", open_folder, border_color=_C['cyan'])
        make_btn(330, 160, "OK", dialog.destroy, border_color=_C['green'])

    # ── Execute ────────────────────────────────────────────────────────
    def run_process(self):
        ip = self.input_file.get()
        od = self.output_dir.get()

        if not ip or not od:
            messagebox.showerror("Error", "Please specify both the input sales CSV file and output folder.")
            return
        if not os.path.exists(ip):
            messagebox.showerror("Error", "Input file does not exist.")
            return
        if not os.path.exists(od):
            messagebox.showerror("Error", "Output folder does not exist.")
            return

        # Get checked products
        selected_prods = [p for p in self.products_data if p['var_select'].get()]
        if not selected_prods:
            messagebox.showerror("No Selection", "Please check at least one product to generate report.")
            return

        self.exec_btn.configure(state='disabled')

        def task():
            try:
                self.root.after(0, lambda: self.set_progress(10, 'CONNECTING API'))

                # Spreadsheet ID and gid are read from credentials_master.json via the loader.
                from googleDrive.credentials_loader import get_sheet_service_account_credentials, get_spreadsheet_id
                scopes = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
                creds = get_sheet_service_account_credentials(scopes=scopes)
                client = gspread.authorize(creds)
                sheet = client.open_by_key(get_spreadsheet_id('master_field_force_sheet') or '1Q4utivZ5OpgDznqlqElYU-HWNnZYI71YYpcZKcSM3xY')

                # Get worksheet
                ws = None
                for w in sheet.worksheets():
                    if str(w.id) == '1918615875':
                        ws = w
                        break
                if not ws:
                    ws = sheet.get_worksheet(0)

                # Fetch GSheet values
                all_values = ws.get_all_values()
                if not all_values:
                    raise ValueError("Worksheet is empty!")

                self.root.after(0, lambda: self.set_progress(30, 'PARSING STRUCTURE'))

                # Cutoff for "FM (SELF APP CODE)" / "FM (SELF"
                cutoff_idx = len(all_values)
                for r_idx in range(1, len(all_values)):
                    row = all_values[r_idx]
                    row_str_joint = " ".join([str(c) for c in row if c is not None])
                    if "FM (SELF APP CODE)" in row_str_joint or "FM (SELF" in row_str_joint:
                        cutoff_idx = r_idx
                        break
                
                header = all_values[0]
                data_rows = all_values[1:cutoff_idx]
                df_sheet = pd.DataFrame(data_rows, columns=header)

                # Clean GSheet rows
                df_sheet = df_sheet[
                    (df_sheet['DEPOT'].str.strip() != '') &
                    (df_sheet['ZONE'].str.strip() != '') &
                    (df_sheet['MARKET'].str.strip() != '')
                ].copy()

                df_sheet['SM'] = df_sheet['SM'].str.strip().str.upper().fillna('VACANT')
                df_sheet['ZONE'] = df_sheet['ZONE'].str.strip().str.upper()
                df_sheet['FM/AM'] = df_sheet['FM/AM'].str.strip().str.upper().fillna('VACANT')
                df_sheet['MARKET'] = df_sheet['MARKET'].str.strip().str.upper()
                df_sheet['MPO NAME, JUN\'26'] = df_sheet["MPO NAME, JUN'26"].str.strip().str.upper().fillna('VACANT')
                
                # Check if there is a VACANT column and filter if checkbox is ticked
                if self.opt_exclude_vacant.get():
                    vacant_col = None
                    for col in df_sheet.columns:
                        if 'VACANT' in str(col).upper():
                            vacant_col = col
                            break
                    if vacant_col:
                        df_sheet = df_sheet[~df_sheet[vacant_col].astype(str).str.upper().isin(['Y', 'YES', 'TRUE', '1'])]

                # The user wants depot code instead of app code
                depot_col = 'DEPOTMPO CODE' if 'DEPOTMPO CODE' in df_sheet.columns else 'DEPOT MPO CODE' if 'DEPOT MPO CODE' in df_sheet.columns else 'DREAM APPS MPO CODE'
                df_sheet['DEPOT_CODE'] = df_sheet[depot_col].str.strip().str.upper()
                df_sheet['DREAM APPS MPO CODE'] = df_sheet['DREAM APPS MPO CODE'].str.strip().str.upper()
                df_sheet['DEPOT'] = df_sheet['DEPOT'].str.strip().str.upper()

                # Load sales CSV
                df_sales = pd.read_csv(ip)
                df_sales['Depot'] = df_sales['Depot'].str.strip().str.upper()
                df_sales['MPO_Code'] = df_sales['MPO_Code'].str.strip().str.upper()
                df_sales['Product_Code'] = df_sales['Product_Code'].str.strip().str.upper()
                df_sales['Month'] = df_sales['Month'].str.strip()

                # Get unique months and filter out June (2026-06) onwards
                all_months = sorted(df_sales['Month'].dropna().unique())
                unique_months = [m for m in all_months if m <= '2026-05']
                
                # Generate abbreviated month names and append Avg
                month_headers = []
                for m in unique_months:
                    try:
                        month_name = datetime.strptime(str(m).strip(), '%Y-%m').strftime('%b')
                    except Exception:
                        month_name = str(m)
                    month_headers.append(month_name)
                month_headers.append('Avg')

                # Map custom names and subgroup names from user inputs
                custom_names = {}
                subgroup_names = {}
                is_subgrouped = {}
                for p in selected_prods:
                    custom_names[p['code']] = p['var_name'].get().strip()
                    subgroup_names[p['code']] = p['var_subgroup'].get().strip()
                    is_subgrouped[p['code']] = p['var_is_subgrouped'].get()

                # Determine groups
                from collections import defaultdict
                groups = defaultdict(list)
                
                group_brand = self.opt_group_brand.get()
                group_subgroup = self.opt_group_subgroup.get()
                summary_end = self.opt_summary_end.get()
                only_summary = self.opt_only_summary.get()

                if group_brand:
                    for p in selected_prods:
                        grp = get_brand_name(custom_names[p['code']])
                        groups[grp].append(p['code'])
                elif group_subgroup:
                    for p in selected_prods:
                        if is_subgrouped[p['code']]:
                            grp = clean_subgroup_key(subgroup_names[p['code']])
                            if not grp:
                                grp = "UNNAMED SUBGROUP"
                        else:
                            grp = f"INDIVIDUAL_{p['code']}"
                        groups[grp].append(p['code'])
                else:
                    for p in selected_prods:
                        groups[p['code']].append(p['code'])

                # Build column structure
                column_items = [] # list of dict: {'type': 'individual'|'summary', 'code': str, 'name': str, 'group_key': str, 'group_label': str}
                for grp in sorted(groups.keys()):
                    codes_in_group = groups[grp]
                    
                    if grp.startswith("INDIVIDUAL_"):
                        grp_label = custom_names[codes_in_group[0]]
                    else:
                        grp_label = grp
                    
                    if (group_brand or group_subgroup) and only_summary:
                        column_items.append({
                            'type': 'summary',
                            'code': '',
                            'name': f"{grp_label} TOTAL",
                            'group_key': grp,
                            'group_label': grp_label
                        })
                    else:
                        for code in codes_in_group:
                            column_items.append({
                                'type': 'individual',
                                'code': code,
                                'name': custom_names[code],
                                'group_key': grp,
                                'group_label': grp_label
                            })
                        
                        if (group_brand or group_subgroup) and summary_end and len(codes_in_group) > 1:
                            column_items.append({
                                'type': 'summary',
                                'code': '',
                                'name': f"{grp_label} TOTAL",
                                'group_key': grp,
                                'group_label': grp_label
                            })

                self.root.after(0, lambda: self.set_progress(60, 'MAPPING SALES'))

                # Aggregate sales data by Product Code
                sales_grouped = df_sales.groupby(['Depot', 'MPO_Code', 'Product_Code', 'Month'])['ACTUAL_SALE_QTY'].sum().reset_index()
                sales_lookup = {}
                for _, row in sales_grouped.iterrows():
                    key = (row['Depot'], row['MPO_Code'], row['Product_Code'], row['Month'])
                    sales_lookup[key] = row['ACTUAL_SALE_QTY']

                # Sort Sheet rows
                df_sheet = df_sheet.sort_values(by=['SM', 'ZONE', 'FM/AM', 'MARKET']).reset_index(drop=True)

                # ── Generate Excel Workbook ──
                self.root.after(0, lambda: self.set_progress(80, 'WRITING EXCEL'))
                
                wb = Workbook()
                ws_out = wb.active
                ws_out.title = "Zone Wise Product Sales"
                ws_out.views.sheetView[0].showGridLines = True

                # Styles
                font_family = "Segoe UI"
                header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
                sub_header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
                total_fill = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
                
                font_title = Font(name=font_family, size=10, bold=True, color="FFFFFF")
                font_sub_title = Font(name=font_family, size=9, bold=True, color="1F497D")
                font_bold = Font(name=font_family, size=9, bold=True)
                font_regular = Font(name=font_family, size=9)

                border_thin = Side(border_style="thin", color="D9D9D9")
                border_double = Side(border_style="double", color="366092")
                grid_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
                bottom_total_border = Border(top=border_thin, bottom=border_double, left=border_thin, right=border_thin)

                metadata_cols = [
                    ('DEPOT', 'DEPOT'),
                    ('ZONE', 'ZONE'),
                    ('MARKET', 'MARKET'),
                    ('FM/AM', 'FM/AM, ZONE'),
                    ('VACANT_BOOL', "VACANT (JUN'26)?"),
                    ('DESIG', "DESIG (JUN'26)"),
                    ('MMO', 'MMO'),
                    ('SM', 'SM'),
                    ('DREAM APPS MPO CODE', 'DREAM APPS MPO CODE'),
                    ('DEPOT_CODE', 'DEPOTMPO CODE'),
                    ("MPO NAME, JUN'26", 'MPO NAME')
                ]
                num_metadata_cols = len(metadata_cols)

                # Write metadata column headers (merge vertically across 3 rows)
                for i, (col_key, col_label) in enumerate(metadata_cols, 1):
                    cell = ws_out.cell(row=1, column=i, value=col_label)
                    cell.font = font_title
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws_out.merge_cells(start_row=1, start_column=i, end_row=3, end_column=i)
                    ws_out.cell(row=2, column=i).fill = header_fill
                    ws_out.cell(row=3, column=i).fill = header_fill

                # Write product headers merged across months
                current_col = num_metadata_cols + 1
                for item in column_items:
                    start_col = current_col
                    end_col = start_col + len(unique_months)
                    ws_out.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                    ws_out.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
                    
                    # Row 1 value: Code or "SUM" label
                    row1_val = item['code'] if item['type'] == 'individual' else f"{item['group_label']} (SUM)"
                    cell_r1 = ws_out.cell(row=1, column=start_col, value=row1_val)
                    cell_r1.font = font_title
                    cell_r1.fill = header_fill
                    cell_r1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                    # Row 2 value: Custom Product Name
                    cell_r2 = ws_out.cell(row=2, column=start_col, value=item['name'])
                    cell_r2.font = font_title
                    cell_r2.fill = header_fill
                    cell_r2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                    for c in range(start_col + 1, end_col + 1):
                        ws_out.cell(row=1, column=c).fill = header_fill
                        ws_out.cell(row=2, column=c).fill = header_fill

                    # Row 3 values: Month names
                    for m_hdr in month_headers:
                        cell_r3 = ws_out.cell(row=3, column=current_col, value=m_hdr)
                        cell_r3.font = font_sub_title
                        cell_r3.fill = sub_header_fill
                        cell_r3.alignment = Alignment(horizontal="center", vertical="center")
                        current_col += 1

                ws_out.row_dimensions[1].height = 24
                ws_out.row_dimensions[2].height = 24
                ws_out.row_dimensions[3].height = 20

                # Populate data rows
                start_row = 4
                for idx, row in df_sheet.iterrows():
                    r_idx = start_row + idx
                    depot = row['DEPOT']
                    zone = row['ZONE']
                    market = row['MARKET']
                    fm = row['FM/AM']
                    
                    vacant_col_name = next((c for c in df_sheet.columns if 'VACANT' in str(c).upper()), None)
                    vacant_bool = row[vacant_col_name] if vacant_col_name and vacant_col_name in row else ''
                    
                    desig_col_name = next((c for c in df_sheet.columns if 'DESIG' in str(c).upper()), None)
                    desig = row[desig_col_name] if desig_col_name and desig_col_name in row else ''
                    
                    mmo = row.get('MMO', '')
                    sm = row['SM']
                    dream_code = row.get('DREAM APPS MPO CODE', '')
                    app_code = row.get('DEPOT_CODE', '')
                    mpo_name = row.get("MPO NAME, JUN'26", '')

                    # Write metadata in specific order
                    meta_vals = [depot, zone, market, fm, vacant_bool, desig, mmo, sm, dream_code, app_code, mpo_name]
                    for c_idx, val in enumerate(meta_vals, 1):
                        ws_out.cell(row=r_idx, column=c_idx, value=val).font = font_regular

                    for c_idx in range(1, num_metadata_cols + 1):
                        ws_out.cell(row=r_idx, column=c_idx).border = grid_border
                        ws_out.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal="left", vertical="center")

                    # Write product data quantities
                    current_col = num_metadata_cols + 1
                    for item in column_items:
                        for m in unique_months:
                            qty = 0
                            if item['type'] == 'individual':
                                # Lookup single code
                                c = item['code']
                                if app_code:
                                    qty = sales_lookup.get((depot, app_code, c, m), 0)
                                if qty == 0 and dream_code:
                                    qty = sales_lookup.get((depot, dream_code, c, m), 0)
                            else:
                                # Lookup group sum
                                for c in groups[item['group_key']]:
                                    val = 0
                                    if app_code:
                                        val = sales_lookup.get((depot, app_code, c, m), 0)
                                    if val == 0 and dream_code:
                                        val = sales_lookup.get((depot, dream_code, c, m), 0)
                                    qty += val

                            cell = ws_out.cell(row=r_idx, column=current_col, value=qty)
                            cell.font = font_regular
                            cell.border = grid_border
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                            cell.number_format = '#,##0'
                            current_col += 1

                        # Write Avg column formula (average of Jan-May columns for this row)
                        jan_col_letter = get_column_letter(current_col - len(unique_months))
                        may_col_letter = get_column_letter(current_col - 1)
                        avg_formula = f"=AVERAGE({jan_col_letter}{r_idx}:{may_col_letter}{r_idx})"
                        
                        cell = ws_out.cell(row=r_idx, column=current_col, value=avg_formula)
                        cell.font = font_regular
                        cell.border = grid_border
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = '#,##0'
                        current_col += 1

                    ws_out.row_dimensions[r_idx].height = 18

                # Add grand totals row
                total_row_idx = start_row + len(df_sheet)
                ws_out.cell(row=total_row_idx, column=1, value="GRAND TOTAL").font = font_bold
                ws_out.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=num_metadata_cols)
                
                for c_idx in range(1, num_metadata_cols + 1):
                    cell = ws_out.cell(row=total_row_idx, column=c_idx)
                    cell.fill = total_fill
                    cell.border = bottom_total_border
                
                ws_out.cell(row=total_row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
                ws_out.row_dimensions[total_row_idx].height = 22

                # Add grand total SUM formulas
                current_col = num_metadata_cols + 1
                for item in column_items:
                    for m in unique_months:
                        col_letter = get_column_letter(current_col)
                        formula = f"=SUM({col_letter}4:{col_letter}{total_row_idx-1})"
                        
                        cell = ws_out.cell(row=total_row_idx, column=current_col, value=formula)
                        cell.font = font_bold
                        cell.fill = total_fill
                        cell.border = bottom_total_border
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = '#,##0'
                        current_col += 1

                    # Add SUM for the Avg column as well
                    col_letter = get_column_letter(current_col)
                    formula = f"=SUM({col_letter}4:{col_letter}{total_row_idx-1})"
                    
                    cell = ws_out.cell(row=total_row_idx, column=current_col, value=formula)
                    cell.font = font_bold
                    cell.fill = total_fill
                    cell.border = bottom_total_border
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = '#,##0'
                    current_col += 1

                # Format column widths
                print("Formatting column widths...")
                for i in range(1, num_metadata_cols + 1):
                    col_letter = get_column_letter(i)
                    max_len = 0
                    for row in range(1, total_row_idx + 1):
                        val = ws_out.cell(row=row, column=i).value
                        if val:
                            max_len = max(max_len, len(str(val)))
                    ws_out.column_dimensions[col_letter].width = max(max_len + 4, 12)

                for i in range(num_metadata_cols + 1, current_col):
                    col_letter = get_column_letter(i)
                    ws_out.column_dimensions[col_letter].width = 8

                # Freeze panes below headers and after MPO name
                ws_out.freeze_panes = "H4"

                # Save file
                timestamp = datetime.now().strftime('%d_%b_%Y_%I.%M_%p')
                out_name = f"03_Zone_Wise_Sales_Grouped_Report_{timestamp}.xlsx"
                out_path = os.path.join(od, out_name)
                
                print(f"Saving workbook to {out_path}...")
                wb.save(out_path)
                wb.close()

                self.root.after(0, lambda: self.set_progress(100, 'COMPLETE'))
                self.root.after(200, lambda: self.show_success_dialog(out_path))
                self.root.after(3000, lambda: [self.set_progress(0), self.exec_btn.configure(state='normal')])

            except Exception as ex:
                self.root.after(0, lambda err=ex: [
                    self.set_progress(0, 'ERROR'),
                    self.lbl_status.configure(fg=_C['red'], text='◉ ERROR OCCURRED'),
                    messagebox.showerror("Execution Failed", f"An error occurred:\n{err}"),
                    self.exec_btn.configure(state='normal'),
                ])
                import traceback
                traceback.print_exc()

        threading.Thread(target=task, daemon=True).start()

# ══════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    root = tk.Tk()
    app = ZoneReportApp(root)
    root.mainloop()
