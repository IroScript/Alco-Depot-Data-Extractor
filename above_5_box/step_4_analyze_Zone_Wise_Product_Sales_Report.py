import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import math
import random
import threading
import glob

# ══════════════════════════════════════════════════════════════════
#  Design tokens (From SmallApp)
# ══════════════════════════════════════════════════════════════════
_C = {
    'void'   : '#03040A',
    'panel'  : '#060816',
    'panel2' : '#050912',
    'cyan'   : '#00F2FE',
    'mag'    : '#FF00AA',
    'violet' : '#9D00FF',
    'green'  : '#00FF88',
    'red'    : '#FF2244',
    'text'   : '#D8F0F8',
    'muted'  : '#5F98A7',
    'border' : '#102A45',
    'hot'    : '#0D2A3F',
}

W_WIN, H_WIN = 490, 315

class ZoneDataAnalyzerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("ALCO ZONE DATA ANALYZER")
        self.root.geometry(f"{W_WIN}x{H_WIN}")
        self.root.resizable(False, False)
        self.root.configure(bg=_C['void'])

        # State
        self.input_file = tk.StringVar()
        self.output_dir = tk.StringVar()
        self.opt_exclude_vacant = tk.BooleanVar(value=True)
        self.opt_exclude_zero_products = tk.BooleanVar(value=True)
        self.opt_exclude_mostly_zero_products = tk.BooleanVar(value=True)
        self._t          = 0.0
        self._scanline_y = 0.0
        self._stars      = []
        self._after_id   = None

        self._build()
        self._seed_stars()
        self._tick()
        self.auto_detect_latest_report()

    def auto_detect_latest_report(self):
        _project_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        reports = glob.glob(os.path.join(_project_dir, '03_Zone_Wise_Sales_Grouped_Report*.xlsx'))
        if reports:
            latest = max(reports, key=os.path.getmtime)
            self.input_file.set(latest)
            self.output_dir.set(os.path.dirname(latest))
            self.lbl_status.configure(text="✓ LATEST REPORT AUTO-SELECTED", fg=_C['green'])

    def _build(self):
        W, H = W_WIN, H_WIN

        self.cv = tk.Canvas(self.root, width=W, height=H, bg=_C['void'], highlightthickness=0, bd=0)
        self.cv.place(x=0, y=0)

        self._paint_gradient(W, H)
        self._paint_grid(W, H)

        self._sl = self.cv.create_line(0, 0, W, 0, fill='#00F2FE', width=1, stipple='gray12')

        self._paint_brackets(W, H)
        self._paint_title(W)

        # File Input
        self._draw_label(20, 55, "INPUT REPORT EXCEL FILE")
        self._draw_entry(20, 75, self.input_file, w=350)
        self._draw_button(385, 75, "BROWSE", self.browse_input, w=80)

        # Output Folder
        self._draw_label(20, 115, "OUTPUT FOLDER")
        self._draw_entry(20, 135, self.output_dir, w=350)
        self._draw_button(385, 135, "BROWSE", self.browse_output, w=80)

        # Options
        cb = tk.Checkbutton(self.root, text="Exclude Vacant MPOs (If still present)",
                            variable=self.opt_exclude_vacant, onvalue=True, offvalue=False,
                            font=('Courier New', 8), fg=_C['text'], bg=_C['void'],
                            activebackground=_C['void'], activeforeground=_C['cyan'], selectcolor=_C['panel'])
        cb.place(x=20, y=175)
        
        cb2 = tk.Checkbutton(self.root, text="Exclude products with zero total sales",
                             variable=self.opt_exclude_zero_products, onvalue=True, offvalue=False,
                             font=('Courier New', 8), fg=_C['text'], bg=_C['void'],
                             activebackground=_C['void'], activeforeground=_C['cyan'], selectcolor=_C['panel'])
        cb2.place(x=20, y=195)

        cb3 = tk.Checkbutton(self.root, text="Exclude products with >95% zero sales across months",
                             variable=self.opt_exclude_mostly_zero_products, onvalue=True, offvalue=False,
                             font=('Courier New', 8), fg=_C['text'], bg=_C['void'],
                             activebackground=_C['void'], activeforeground=_C['cyan'], selectcolor=_C['panel'])
        cb3.place(x=20, y=215)

        # Status
        self.lbl_status = tk.Label(self.root, text="◉ SYSTEM READY", font=('Courier New', 8), fg=_C['cyan'], bg=_C['void'])
        self.lbl_status.place(x=20, y=245)

        # Progress Bar
        self.progress_bar = self.cv.create_rectangle(20, 260, 20, 263, fill=_C['cyan'], outline='')

        # Execute
        self.exec_btn = tk.Button(self.root, text="⟨  RUN 10 PARAMETER ANALYSIS  ⟩", command=self.run_process,
                                  font=('Courier New', 10, 'bold'), fg=_C['cyan'], bg='#030A1C', relief='flat',
                                  cursor='hand2', activebackground='#061C34', activeforeground='#FFFFFF', bd=1,
                                  highlightbackground=_C['cyan'])
        self.exec_btn.place(x=20, y=275, width=450, height=24)

    def _paint_gradient(self, W, H):
        for y in range(H):
            t = y / H
            r = int(3  + t * 4)
            g = int(4  + t * 5)
            b = int(10 + t * 16)
            self.cv.create_line(0, y, W, y, fill=f'#{r:02x}{g:02x}{b:02x}', width=1)

    def _paint_grid(self, W, H):
        for x in range(0, W, 40):
            self.cv.create_line(x, 0, x, H, fill='#07091A', width=1)
        for y in range(0, H, 40):
            self.cv.create_line(0, y, W, y, fill='#07091A', width=1)

    def _paint_brackets(self, W, H, pad=10, sz=18):
        c = _C['cyan']
        self.cv.create_line(pad, pad, pad+sz, pad, fill=c, width=1)
        self.cv.create_line(pad, pad, pad, pad+sz, fill=c, width=1)
        self.cv.create_line(W-pad-sz, pad, W-pad, pad, fill=c, width=1)
        self.cv.create_line(W-pad, pad, W-pad, pad+sz, fill=c, width=1)
        self.cv.create_line(pad, H-pad, pad+sz, H-pad, fill=c, width=1)
        self.cv.create_line(pad, H-pad-sz, pad, H-pad, fill=c, width=1)
        self.cv.create_line(W-pad-sz, H-pad, W-pad, H-pad, fill=c, width=1)
        self.cv.create_line(W-pad, H-pad-sz, W-pad, H-pad, fill=c, width=1)

    def _paint_title(self, W):
        self.cv.create_text(W//2, 22, text='ZONE DATA ANALYZER', font=('Courier New', 12, 'bold'), fill=_C['cyan'])
        self.cv.create_text(W//2, 36, text='∞ 10 PARAMETER STATISTICAL GENERATOR ∞', font=('Courier New', 7), fill=_C['muted'])
        self.cv.create_line(40, 46, W-40, 46, fill=_C['border'], width=1)

    def _draw_label(self, x, y, text):
        lbl = tk.Label(self.root, text=text, font=('Courier New', 7, 'bold'), fg=_C['muted'], bg=_C['void'])
        lbl.place(x=x, y=y)

    def _draw_entry(self, x, y, var, w):
        ent = tk.Entry(self.root, textvariable=var, bg=_C['panel2'], fg=_C['text'],
                       insertbackground=_C['cyan'], relief='flat', font=('Segoe UI', 8), bd=1,
                       highlightthickness=1, highlightcolor=_C['cyan'], highlightbackground=_C['border'])
        ent.place(x=x, y=y, width=w, height=22)

    def _draw_button(self, x, y, text, cmd, w):
        btn = tk.Button(self.root, text=text, command=cmd, font=('Courier New', 7, 'bold'),
                        fg=_C['cyan'], bg='#07152B', relief='flat', cursor='hand2', activebackground='#0B2A4A')
        btn.place(x=x, y=y, width=w, height=22)

    def _tick(self):
        self._t += 0.05
        self._scanline_y = (self._scanline_y + 1.5) % H_WIN
        self.cv.coords(self._sl, 0, self._scanline_y, W_WIN, self._scanline_y)

        status_text = self.lbl_status.cget('text')
        if 'READY' in status_text:
            v = int(180 + 50 * math.sin(self._t))
            self.lbl_status.configure(fg=f'#00{v:02x}{v//2:02x}')

        self._after_id = self.root.after(50, self._tick)

    def _seed_stars(self):
        colors = [_C['cyan'], '#FFFFFF', _C['violet'], '#4FACFE']
        for _ in range(15):
            x = random.randint(15, W_WIN-15)
            y = random.randint(50, H_WIN-20)
            sz = random.choice([1, 1, 2])
            sid = self.cv.create_oval(x, y, x+sz, y+sz, fill=random.choice(colors), outline='', state='hidden')
            self._stars.append({'id': sid, 'ph': random.uniform(0, 2*math.pi), 'spd': random.uniform(0.025, 0.065)})

    def browse_input(self):
        f = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if f:
            self.input_file.set(f)
            if not self.output_dir.get():
                self.output_dir.set(os.path.dirname(f))

    def browse_output(self):
        d = filedialog.askdirectory()
        if d:
            self.output_dir.set(d)

    def update_progress(self, pct, status):
        self.lbl_status.configure(text=f'◉ {status}')
        w = 450 * (pct / 100.0)
        self.cv.coords(self.progress_bar, 20, 260, 20+w, 263)
        self.root.update_idletasks()

    def show_success_dialog(self, out_path):
        dialog = tk.Toplevel(self.root)
        dialog.title("Success")
        dialog.geometry("420x200")
        dialog.resizable(False, False)
        dialog.configure(bg=_C['void'])
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center relative to parent (W_WIN=490, H_WIN=315, Dialog=420x200)
        parent_x = self.root.winfo_rootx()
        parent_y = self.root.winfo_rooty()
        dialog.geometry(f"+{parent_x + 35}+{parent_y + 57}")

        W_D, H_D = 420, 200
        cv = tk.Canvas(dialog, width=W_D, height=H_D, bg=_C['void'], highlightthickness=0, bd=0)
        cv.place(x=0, y=0)

        # Background decoration
        for y in range(H_D):
            t = y / H_D
            r = int(3  + t * 4)
            g = int(4  + t * 5)
            b = int(10 + t * 16)
            cv.create_line(0, y, W_D, y, fill=f'#{r:02x}{g:02x}{b:02x}', width=1)

        # Brackets
        c = _C['cyan']
        pad, sz = 10, 12
        cv.create_line(pad, pad, pad+sz, pad, fill=c, width=1)
        cv.create_line(pad, pad, pad, pad+sz, fill=c, width=1)
        cv.create_line(W_D-pad-sz, pad, W_D-pad, pad, fill=c, width=1)
        cv.create_line(W_D-pad, pad, W_D-pad, pad+sz, fill=c, width=1)
        cv.create_line(pad, H_D-pad, pad+sz, H_D-pad, fill=c, width=1)
        cv.create_line(pad, H_D-pad-sz, pad, H_D-pad, fill=c, width=1)
        cv.create_line(W_D-pad-sz, H_D-pad, W_D-pad, H_D-pad, fill=c, width=1)
        cv.create_line(W_D-pad, H_D-pad-sz, W_D-pad, H_D-pad, fill=c, width=1)

        # Circle tick
        cv.create_oval(25, 30, 55, 60, fill='#051A18', outline=_C['green'], width=1.5)
        cv.create_text(40, 45, text='✓', font=('Courier New', 13, 'bold'), fill=_C['green'])

        # Title
        cv.create_text(70, 45, text="PROCESSING COMPLETE", font=('Courier New', 10, 'bold'), fill=_C['green'], anchor='w')

        norm_path = os.path.normpath(out_path)
        wrapped_path = norm_path
        if len(norm_path) > 45:
            chunks = []
            for i in range(0, len(norm_path), 45):
                chunks.append(norm_path[i:i+45])
            wrapped_path = "\n".join(chunks)

        cv.create_text(25, 80, text=f"Saved to:\n{wrapped_path}", font=('Courier New', 8), fill=_C['text'], anchor='nw')

        # Action handlers
        def open_file():
            try:
                os.startfile(norm_path)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to open file:\n{e}", parent=dialog)

        def open_folder():
            try:
                os.startfile(os.path.dirname(norm_path))
            except Exception as e:
                messagebox.showerror("Error", f"Failed to open folder:\n{e}", parent=dialog)

        # Draw buttons using Canvas items and Tkinter Buttons
        def make_btn(cx, cy, label, cmd, w=100, h=24, border_color=_C['cyan']):
            x1, y1, x2, y2 = cx-w//2, cy-h//2, cx+w//2, cy+h//2
            r = 4
            poly = [x1+r, y1, x2-r, y1, x2, y1, x2, y1+r, x2, y2-r, x2, y2, x2-r, y2, x1+r, y2, x1, y2, x1, y2-r, x1, y1+r, x1, y1]
            bg_id = cv.create_polygon(poly, smooth=True, fill='#07152B', outline=border_color, width=1)
            
            btn = tk.Button(dialog, text=label, command=cmd, font=('Courier New', 8, 'bold'), fg=border_color,
                            bg='#07152B', relief='flat', cursor='hand2', activebackground='#0B2A4A',
                            activeforeground='#FFFFFF', highlightthickness=0, bd=0)
            cv.create_window(cx, cy, window=btn, width=w-2, height=h-2)

            def on_in(e):
                cv.itemconfig(bg_id, fill=_C['hot'], outline='#FFFFFF')
                btn.configure(bg=_C['hot'], fg='#FFFFFF')
            def on_out(e):
                cv.itemconfig(bg_id, fill='#07152B', outline=border_color)
                btn.configure(bg='#07152B', fg=border_color)

            btn.bind('<Enter>', on_in)
            btn.bind('<Leave>', on_out)

        make_btn(90, 160, "OPEN FILE", open_file, border_color=_C['cyan'])
        make_btn(210, 160, "OPEN FOLDER", open_folder, border_color=_C['cyan'])
        make_btn(330, 160, "OK", dialog.destroy, border_color=_C['green'])

    def run_process(self):
        ip = self.input_file.get()
        od = self.output_dir.get()
        
        if not ip or not od or not os.path.exists(ip) or not os.path.exists(od):
            messagebox.showerror("Error", "Valid input file and output directory required.")
            return
            
        self.exec_btn.configure(state='disabled')
        self.update_progress(10, 'LOADING EXCEL REPORT...')

        def task():
            try:
                wb = load_workbook(ip)
                ws = wb.active
                
                # Find start and end rows
                start_row = 4
                end_row = ws.max_row
                grand_total_row = None
                
                for r in range(end_row, 0, -1):
                    val = ws.cell(row=r, column=1).value
                    if val == "GRAND TOTAL":
                        grand_total_row = r
                        break
                
                if grand_total_row:
                    end_row = grand_total_row - 1
                
                self.update_progress(30, 'FILTERING DATA...')
                if self.opt_exclude_vacant.get():
                    for r in range(end_row, start_row - 1, -1):
                        mpo_val = str(ws.cell(row=r, column=11).value or "").strip().upper()
                        fm_val = str(ws.cell(row=r, column=4).value or "").strip().upper()
                        sm_val = str(ws.cell(row=r, column=8).value or "").strip().upper()
                        vacant_val = str(ws.cell(row=r, column=5).value or "").strip().upper()
                        
                        if mpo_val == 'VACANT' or fm_val == 'VACANT' or sm_val == 'VACANT' or vacant_val in ['Y', 'YES', 'TRUE', '1']:
                            ws.delete_rows(r, 1)
                            end_row -= 1
                            if grand_total_row:
                                grand_total_row -= 1
                                
                max_col = ws.max_column

                # Exclude products based on zero/mostly zero conditions
                exclude_zero = self.opt_exclude_zero_products.get()
                exclude_mostly_zero = self.opt_exclude_mostly_zero_products.get()
                
                if (exclude_zero or exclude_mostly_zero) and grand_total_row:
                    self.update_progress(40, 'EXCLUDING UNUSED PRODUCTS...')
                    
                    product_boundaries = []
                    current_prod_start = 12
                    
                    for col in range(13, max_col + 1):
                        val1 = ws.cell(row=1, column=col).value
                        if val1 is not None and str(val1).strip() != "":
                            product_boundaries.append((current_prod_start, col - 1))
                            current_prod_start = col
                    
                    if current_prod_start <= max_col:
                        product_boundaries.append((current_prod_start, max_col))
                        
                    for start_c, end_c in reversed(product_boundaries):
                        prod_sum = 0
                        zero_count = 0
                        total_cells = (grand_total_row - start_row) * (end_c - start_c + 1)
                        
                        for c in range(start_c, end_c + 1):
                            for r in range(start_row, grand_total_row):
                                val = ws.cell(row=r, column=c).value
                                if isinstance(val, (int, float)):
                                    prod_sum += val
                                elif val:
                                    try:
                                        prod_sum += float(val)
                                    except ValueError:
                                        pass
                                
                                is_zero = False
                                if val is None:
                                    is_zero = True
                                else:
                                    try:
                                        fval = float(val)
                                        if fval == 0:
                                            is_zero = True
                                    except ValueError:
                                        sval = str(val).strip()
                                        if sval in ['', '0', '-', 'None']:
                                            is_zero = True
                                if is_zero:
                                    zero_count += 1
                        
                        should_exclude = False
                        if exclude_zero and prod_sum == 0:
                            should_exclude = True
                        elif exclude_mostly_zero and total_cells > 0 and (zero_count / total_cells) > 0.95:
                            should_exclude = True
                            
                        if should_exclude:
                            ws.delete_cols(start_c, end_c - start_c + 1)
                            
                    max_col = ws.max_column

                self.update_progress(50, 'FINDING ZONES & INSERTING 10 PARAMETERS...')
                
                # Identify zone boundaries
                zone_boundaries = []
                last_zone = None
                for r in range(start_row, end_row + 1):
                    z_val = str(ws.cell(row=r, column=2).value or "").strip()
                    if last_zone is not None and z_val != last_zone:
                        zone_boundaries.append((r, last_zone))
                    last_zone = z_val
                    
                if end_row >= start_row:
                    zone_boundaries.append((end_row + 1, last_zone))
                
                max_col = ws.max_column
                
                # Define 10 parameters
                param_names = [
                    "COUNT", "SUM", "MEAN", "MEDIAN", "VARIANCE",
                    "SD", "MIN", "MAX", "RANGE", "CV %"
                ]
                
                # Curated premium color formatting for parameter rows
                font_family = "Arial Narrow"
                font_param_label = Font(name=font_family, size=11, bold=True, color="1F497D")
                font_param_val_bold = Font(name=font_family, size=11, bold=True)
                font_param_val_reg = Font(name=font_family, size=11)
                
                fill_sum_mean = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft light green
                fill_cv = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Soft light yellow
                fill_other = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Very light gray
                
                border_thin = Side(border_style="thin", color="D9D9D9")
                param_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

                # Process backwards to avoid messing up row indices
                for idx, (r_boundary, z_name) in enumerate(reversed(zone_boundaries)):
                    i = len(zone_boundaries) - 1 - idx
                    offset = i * 13
                    
                    # We need to find where this zone started to create the formulas
                    z_start = start_row
                    for r in range(r_boundary - 1, start_row - 1, -1):
                        if str(ws.cell(row=r, column=2).value or "").strip() != z_name:
                            z_start = r + 1
                            break
                    z_end = r_boundary - 1
                    
                    final_z_start = z_start + offset
                    final_z_end = z_end + offset
                    
                    # Insert 13 rows (10 for parameters, 3 for gap)
                    ws.insert_rows(r_boundary, amount=13)
                    if grand_total_row:
                        grand_total_row += 13
                        
                    # Write parameter labels in ZONE column (2)
                    for idx_p, p_name in enumerate(param_names):
                        r_curr = r_boundary + idx_p
                        ws.cell(row=r_curr, column=2, value=p_name)
                        
                    # Write formulas for all data columns (12 to max_col)
                    for col in range(12, max_col + 1):
                        col_l = get_column_letter(col)
                        rng = f"{col_l}{final_z_start}:{col_l}{final_z_end}"
                        
                        # Row indices for the inserted rows
                        r_count = r_boundary
                        r_sum = r_boundary + 1
                        r_mean = r_boundary + 2
                        r_median = r_boundary + 3
                        r_var = r_boundary + 4
                        r_sd = r_boundary + 5
                        r_min = r_boundary + 6
                        r_max = r_boundary + 7
                        r_range = r_boundary + 8
                        r_cv = r_boundary + 9
                        
                        def dyn_round(expr):
                            return f"=IFERROR(IF({expr}=0, 0, IF({expr}<0.5, ROUND({expr}, 2), IF({expr}<1, ROUND({expr}, 1), ROUND({expr}, 0)))), \"-\")"
 
                        # Formulas
                        ws.cell(row=r_count, column=col, value=f"=COUNTIF({rng}, \">5\")")
                        ws.cell(row=r_sum, column=col, value=dyn_round(f"SUMIF({rng}, \">5\")"))
                        ws.cell(row=r_mean, column=col, value=dyn_round(f"AVERAGEIF({rng}, \">5\")"))
                        ws.cell(row=r_median, column=col, value=dyn_round(f"MEDIAN(_xlfn.FILTER({rng}, {rng}>5))"))
                        ws.cell(row=r_var, column=col, value=dyn_round(f"VARP(_xlfn.FILTER({rng}, {rng}>5))"))
                        ws.cell(row=r_sd, column=col, value=dyn_round(f"STDEVP(_xlfn.FILTER({rng}, {rng}>5))"))
                        ws.cell(row=r_min, column=col, value=dyn_round(f"MIN(_xlfn.FILTER({rng}, {rng}>5))"))
                        ws.cell(row=r_max, column=col, value=dyn_round(f"MAX(_xlfn.FILTER({rng}, {rng}>5))"))
                        ws.cell(row=r_range, column=col, value=dyn_round(f"(MAX(_xlfn.FILTER({rng}, {rng}>5))-MIN(_xlfn.FILTER({rng}, {rng}>5)))"))
                        
                        # CV %: use raw decimal division, format as whole percentage (e.g., 100%)
                        final_r_sd = r_sd + offset
                        final_r_mean = r_mean + offset
                        cv_cell = ws.cell(row=r_cv, column=col, value=f"=IFERROR({col_l}{final_r_sd}/{col_l}{final_r_mean}, \"-\")")
                        cv_cell.number_format = '0%'
                    
                    # Format & style the parameter and gap rows
                    for idx_p, p_name in enumerate(param_names):
                        r_curr = r_boundary + idx_p
                        ws.row_dimensions[r_curr].height = 18
                        
                        for col in range(1, max_col + 1):
                            cell = ws.cell(row=r_curr, column=col)
                            cell.border = param_border
                            
                            if col == 2:
                                cell.font = font_param_label
                                cell.alignment = Alignment(horizontal="left", vertical="center")
                            elif col >= 12:
                                if p_name in ["SUM", "MEAN", "CV %"]:
                                    cell.font = font_param_val_bold
                                else:
                                    cell.font = font_param_val_reg
                                cell.alignment = Alignment(horizontal="right", vertical="center")
                            else:
                                cell.font = font_param_val_reg
                                cell.alignment = Alignment(horizontal="left", vertical="center")
                                
                            if p_name in ["SUM", "MEAN"]:
                                cell.fill = fill_sum_mean
                            elif p_name == "CV %":
                                cell.fill = fill_cv
                            else:
                                cell.fill = fill_other

                    for idx_g in range(10, 13):
                        r_gap = r_boundary + idx_g
                        ws.row_dimensions[r_gap].height = 10
                        for col in range(1, max_col + 1):
                            cell = ws.cell(row=r_gap, column=col)
                            cell.border = Border()
                            cell.fill = PatternFill(fill_type=None)

                    # Add row page break after the gap rows (except for the last zone)
                    if idx > 0:
                        ws.row_breaks.append(Break(id=r_boundary + offset + 12))
                            
                # Correct the average formulas in shifted rows
                self.update_progress(75, 'CORRECTING SHIFTED AVERAGE FORMULAS...')
                avg_cols = []
                for col in range(12, max_col + 1):
                    val_r3 = ws.cell(row=3, column=col).value
                    if val_r3 and str(val_r3).strip().upper() == "AVG":
                        avg_cols.append(col)
                
                for r in range(start_row, grand_total_row):
                    val_c2 = str(ws.cell(row=r, column=2).value or "").strip().upper()
                    if val_c2 in ["COUNT", "SUM", "MEAN", "MEDIAN", "VARIANCE", "SD", "MIN", "MAX", "RANGE", "CV %"]:
                        continue
                    
                    is_empty = True
                    for col in range(1, max_col + 1):
                        if ws.cell(row=r, column=col).value is not None:
                            is_empty = False
                            break
                    if is_empty:
                        continue
                    
                    # Rewrite Avg formulas
                    for col_idx in avg_cols:
                        jan_col_letter = get_column_letter(col_idx - 5)
                        may_col_letter = get_column_letter(col_idx - 1)
                        ws.cell(row=r, column=col_idx, value=f"=AVERAGE({jan_col_letter}{r}:{may_col_letter}{r})")

                self.update_progress(80, 'UPDATING GRAND TOTAL FORMULAS...')
                if grand_total_row:
                    for col in range(12, max_col + 1):
                        val = ws.cell(row=grand_total_row, column=col).value
                        if str(val).startswith("=SUM("):
                            col_letter = get_column_letter(col)
                            sum_cells = []
                            for r in range(start_row, grand_total_row):
                                if str(ws.cell(row=r, column=2).value).strip().upper() == "SUM":
                                    sum_cells.append(f"{col_letter}{r}")
                            if sum_cells:
                                new_formula = f"=SUM({','.join(sum_cells)})"
                                ws.cell(row=grand_total_row, column=col, value=new_formula)
                
                # Apply Arial Narrow font (size 11) to the entire sheet
                for r in range(1, ws.max_row + 1):
                    for col in range(1, max_col + 1):
                        cell = ws.cell(row=r, column=col)
                        is_bold = False
                        if cell.font and cell.font.bold:
                            is_bold = True
                        elif r in [1, 2, 3]:
                            is_bold = True
                        
                        color = cell.font.color if cell.font else None
                        cell.font = Font(name="Arial Narrow", size=11, bold=is_bold, color=color)

                # Set row heights for rotated headers (Row 2 for product names, Row 3 for months)
                ws.row_dimensions[2].height = 120
                ws.row_dimensions[3].height = 35

                # Re-format all column widths to prevent clipped text
                for col in range(1, max_col + 1):
                    col_letter = get_column_letter(col)
                    max_len = 0
                    for row in range(1, ws.max_row + 1):
                        val = ws.cell(row=row, column=col).value
                        if val:
                            val_str = str(val)
                            if not val_str.startswith('='):
                                max_len = max(max_len, len(val_str))
                    if col <= 11:
                        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
                    else:
                        # 17 pixels is approx width 1.71 in Excel
                        ws.column_dimensions[col_letter].width = 1.71

                # Rotate header text for product columns
                for col in range(12, max_col + 1):
                    for r in [2, 3]:
                        cell = ws.cell(row=r, column=col)
                        cur_align = cell.alignment
                        cell.alignment = Alignment(
                            horizontal=cur_align.horizontal if cur_align else "center",
                            vertical=cur_align.vertical if cur_align else "center",
                            wrap_text=cur_align.wrap_text if cur_align else True,
                            textRotation=90
                        )

                # Identify product boundaries in the final sheet
                product_boundaries = []
                current_prod_start = 12
                for col in range(13, max_col + 1):
                    val1 = ws.cell(row=1, column=col).value
                    if val1 is not None and str(val1).strip() != "":
                        product_boundaries.append((current_prod_start, col - 1))
                        current_prod_start = col
                if current_prod_start <= max_col:
                    product_boundaries.append((current_prod_start, max_col))

                # Apply vertical borders to box each product group
                border_medium_vertical = Side(border_style="medium", color="000000")
                for start_c, end_c in product_boundaries:
                    for r in range(2, ws.max_row + 1):
                        cell = ws.cell(row=r, column=end_c)
                        cur_border = cell.border
                        cell.border = Border(
                            left=cur_border.left if cur_border else None,
                            right=border_medium_vertical,
                            top=cur_border.top if cur_border else None,
                            bottom=cur_border.bottom if cur_border else None
                        )
                
                # Apply left border to the very first product column (column 12)
                if max_col >= 12:
                    for r in range(2, ws.max_row + 1):
                        cell = ws.cell(row=r, column=12)
                        cur_border = cell.border
                        cell.border = Border(
                            left=border_medium_vertical,
                            right=cur_border.right if cur_border else None,
                            top=cur_border.top if cur_border else None,
                            bottom=cur_border.bottom if cur_border else None
                        )

                # Set Print Layout Settings for Legal Landscape & Fitting
                ws.page_setup.paperSize = 5 # Legal size
                ws.page_setup.orientation = 'landscape'
                ws.sheet_properties.pageSetUpPr.fitToPage = True
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0
                
                ws.page_margins.left = 0.25
                ws.page_margins.right = 0.25
                ws.page_margins.top = 0.25
                ws.page_margins.bottom = 0.25
                
                ws.print_title_rows = '1:3' # Repeat header rows on every page
                ws.views.sheetView[0].showGridLines = True
                
                # Hide columns for clean print layout (DEPOT, and VACANT to MPO NAME)
                hide_cols = [1, 5, 6, 7, 8, 9, 10, 11]
                for col in hide_cols:
                    col_letter = get_column_letter(col)
                    ws.column_dimensions[col_letter].hidden = True
                
                self.update_progress(90, 'SAVING FILE...')
                original_name = os.path.basename(ip)
                # Strip the 03_ prefix if it exists to keep it clean
                if original_name.startswith("03_"):
                    original_name = original_name[3:]
                out_name = "04_Analyzed_10_Param_" + original_name
                out_path = os.path.join(od, out_name)
                wb.save(out_path)
                wb.close()
                
                self.root.after(0, lambda: [
                    self.update_progress(100, 'COMPLETED SUCCESSFULLY'),
                    self.lbl_status.configure(fg=_C['green']),
                    self.show_success_dialog(out_path),
                    self.exec_btn.configure(state='normal')
                ])
                
            except Exception as e:
                self.root.after(0, lambda err=e: [
                    self.update_progress(0, 'ERROR OCCURRED'),
                    self.lbl_status.configure(fg=_C['red']),
                    messagebox.showerror("Error", f"Failed to process file:\n{err}"),
                    self.exec_btn.configure(state='normal')
                ])

        threading.Thread(target=task, daemon=True).start()

if __name__ == "__main__":
    root = tk.Tk()
    app = ZoneDataAnalyzerApp(root)
    root.mainloop()
