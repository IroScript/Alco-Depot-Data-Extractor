"""
Create Target vs Achievement with Alternating Columns
====================================================
Structure: For each product, create two columns:
  - Column 1: Product Target
  - Column 2: Product Achievement (ACH.)

Example:
  G: Mokast-10 TARGET
  H: Mokast-10 ACH.
  I: Alagra 120 TARGET
  J: Alagra 120 ACH.

Author: Auto-generated
Date: 2026-05-23
"""

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# ============================================================================
# CONFIGURATION
# ============================================================================

TARGET_FILE = r"C:\Users\Irak\Desktop\Barishal April Data\MPO_Field_With_Targets_20260523_102519.xlsx"
ACHIEVEMENT_FILE = r"C:\Users\Irak\Desktop\Barishal April Data\Achievement_Pivot_Fixed_20260523_124434.xlsx"

print("=" * 80)
print("CREATE TARGET VS ACHIEVEMENT WITH ALTERNATING COLUMNS")
print("=" * 80)
print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

# ============================================================================
# STEP 1: LOAD FILES
# ============================================================================

print("STEP 1: Loading files...")

df_target = pd.read_excel(TARGET_FILE, sheet_name='MPO_Field_Targets')
df_achievement = pd.read_excel(ACHIEVEMENT_FILE, sheet_name='April_Achievement')

print(f"  - Target records: {len(df_target)}")
print(f"  - Achievement records: {len(df_achievement)}")

# ============================================================================
# STEP 2: CREATE PRODUCT CODE MAPPING
# ============================================================================

print("\nSTEP 2: Creating product code mapping...")

# Get unique product codes and names from achievement file
product_mapping = df_achievement[['Product_Code', 'Product_Name']].drop_duplicates()
product_code_dict = dict(zip(product_mapping['Product_Name'], product_mapping['Product_Code']))

print(f"  - Unique products in achievement: {len(product_code_dict)}")

# ============================================================================
# STEP 3: MATCH TARGET PRODUCTS TO CODES
# ============================================================================

print("\nSTEP 3: Matching target products to product codes...")

# Get product columns from target
base_cols = ['DEPOT', 'ZONE', 'FM/AM, ZONE', 'MARKET', 'MPO CODE', 'DEPOT_MPO_CODE']
product_cols = [col for col in df_target.columns if col not in base_cols + ['Total_Target']]

# Match target products to achievement products
from difflib import SequenceMatcher
import re

def normalize_product_name(name):
    if pd.isna(name):
        return ""
    name = str(name).upper().strip()
    name = re.sub(r'\s*(TAB|TABLET|CAP|CAPSULE|SYP|SYRUP|SUSP|SUSPENSION)\.?\s*', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\([^)]*\)', '', name)
    name = re.sub(r'\d+\s*MG', '', name, flags=re.IGNORECASE)
    name = name.replace('-', '').replace('.', '').replace(',', '').replace('+', '')
    name = ' '.join(name.split())
    return name.replace(' ', '')

def find_product_code(target_product, product_code_dict):
    target_norm = normalize_product_name(target_product)
    
    best_match = None
    best_score = 0
    
    for ach_product, code in product_code_dict.items():
        ach_norm = normalize_product_name(ach_product)
        score = SequenceMatcher(None, target_norm, ach_norm).ratio()
        
        if score > best_score:
            best_score = score
            best_match = code
    
    if best_score >= 0.75:
        return best_match
    else:
        return None

# Create mapping of target product to product code
target_to_code = {}
for prod in product_cols:
    code = find_product_code(prod, product_code_dict)
    target_to_code[prod] = code if code else ""

matched = sum(1 for v in target_to_code.values() if v)
print(f"  - Matched: {matched}/{len(product_cols)}")

# ============================================================================
# STEP 4: CREATE NEW DATAFRAME WITH ALTERNATING COLUMNS
# ============================================================================

print("\nSTEP 4: Creating new dataframe with alternating columns...")

# Create new column structure: base columns + alternating (target, ach) for each product
new_columns = base_cols.copy()
new_col_mapping = {}  # Maps new column name to (original_col, type)

for prod in product_cols:
    # Only add if product has a code match
    if target_to_code.get(prod):
        target_col = f"{prod}"
        ach_col = f"{prod}_ACH"
        new_columns.append(target_col)
        new_columns.append(ach_col)
        new_col_mapping[target_col] = (prod, 'target')
        new_col_mapping[ach_col] = (prod, 'ach')

# Create new dataframe
df_new = pd.DataFrame(columns=new_columns)

# Copy base columns
for col in base_cols:
    df_new[col] = df_target[col]

# Copy target values
for new_col, (orig_col, col_type) in new_col_mapping.items():
    if col_type == 'target':
        df_new[new_col] = df_target[orig_col]
    else:
        # Initialize achievement columns with 0
        df_new[new_col] = 0

print(f"  - New shape: {df_new.shape}")
print(f"  - Total columns: {len(new_columns)} ({len(base_cols)} base + {len(new_col_mapping)} product columns)")

# ============================================================================
# STEP 5: ADD PRODUCT CODES ROW AND ACH LABELS ROW
# ============================================================================

print("\nSTEP 5: Adding product codes and ACH labels rows...")

# Create row 1: Product codes
row1_data = {}
for col in base_cols:
    row1_data[col] = ""

for new_col, (orig_col, col_type) in new_col_mapping.items():
    if col_type == 'target':
        row1_data[new_col] = ""  # Empty for target columns
    else:
        # Product code for ACH columns
        row1_data[new_col] = target_to_code.get(orig_col, "")

# Create row 2: Column labels
row2_data = {}
for col in base_cols:
    row2_data[col] = ""

for new_col, (orig_col, col_type) in new_col_mapping.items():
    if col_type == 'target':
        row2_data[new_col] = "TARGET"
    else:
        row2_data[new_col] = "ACH."

# Insert these rows at the top
df_row1 = pd.DataFrame([row1_data])
df_row2 = pd.DataFrame([row2_data])

df_final = pd.concat([df_row1, df_row2, df_new], ignore_index=True)

print(f"  - Final shape: {df_final.shape}")
print(f"  - Total rows: {len(df_final)} (2 header rows + {len(df_new)} data rows)")

# ============================================================================
# STEP 6: SAVE TO EXCEL WITH VLOOKUP FORMULAS
# ============================================================================

print("\nSTEP 6: Saving to Excel with VLOOKUP formulas...")

timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
output_file = f"MPO_Target_vs_Achievement_{timestamp}.xlsx"

# Save initial dataframe
df_final.to_excel(output_file, sheet_name='Target_vs_Achievement', index=False)

# Now add VLOOKUP formulas using openpyxl
print("  - Adding VLOOKUP formulas...")

wb = load_workbook(output_file)
ws = wb['Target_vs_Achievement']

# Find column indices for ACH columns
# Excel columns start at 1, pandas adds header row at row 1
# So our row 1 (codes) becomes Excel row 2, row 2 (labels) becomes Excel row 3, etc.

for excel_col_idx, col_name in enumerate(df_final.columns, start=1):
    # Check if this is an ACH column
    if col_name in new_col_mapping:
        orig_col, col_type = new_col_mapping[col_name]
        
        if col_type == 'ach':
            col_letter = get_column_letter(excel_col_idx)
            
            # Get product code from row 2 (Excel row 2, which is our dataframe row 0)
            product_code_cell = f"{col_letter}$2"
            
            # DEPOT_MPO_CODE is in column F
            depot_mpo_cell = "$F"
            
            # Add formulas starting from row 4 (Excel row 4 = dataframe row 2 = first data row)
            # Excel rows: 1=header, 2=codes, 3=labels, 4+=data
            for row_idx in range(4, len(df_final) + 2):  # +2 because Excel is 1-indexed and we have header
                # VLOOKUP formula using UPPER() to match the uppercase lookup column
                # Column A has LOOKUP_KEY_UPPER, Column H (8th column) has April_Achievement
                formula = f'=IFERROR(VLOOKUP(UPPER({depot_mpo_cell}{row_idx}&"_"&{product_code_cell}),[{ACHIEVEMENT_FILE}]April_Achievement!$A:$H,8,FALSE),0)'
                ws[f"{col_letter}{row_idx}"] = formula

print(f"  - Formulas added for {matched} products")

# Save workbook
wb.save(output_file)
wb.close()

print(f"\n✓ Output saved: {output_file}")

# ============================================================================
# STEP 7: CREATE VERSION WITH CALCULATED VALUES
# ============================================================================

print("\nSTEP 7: Creating version with calculated values...")

# Create achievement lookup
achievement_lookup = {}
for idx, row in df_achievement.iterrows():
    key = row['LOOKUP_KEY_UPPER']  # Use the uppercase key
    achievement_lookup[key] = row['April_Achievement']

# Load the file we just created
df_with_values = pd.read_excel(output_file, sheet_name='Target_vs_Achievement')

# Convert all columns to object type to allow mixed data
for col in df_with_values.columns:
    df_with_values[col] = df_with_values[col].astype(object)

# Add achievement values
col_idx_map = {col: i for i, col in enumerate(df_with_values.columns)}

for new_col, (orig_col, col_type) in new_col_mapping.items():
    if col_type == 'ach':
        code = target_to_code.get(orig_col)
        if code:
            # Calculate achievements for each row (starting from row 2, which is index 2 in dataframe)
            for row_idx in range(2, len(df_with_values)):
                depot_mpo = df_with_values.iloc[row_idx]['DEPOT_MPO_CODE']
                if pd.notna(depot_mpo):
                    # Add underscore to match the concatenation in Achievement_Pivot
                    key = f"{depot_mpo}_{code}".upper()
                    ach_value = achievement_lookup.get(key, 0)
                    # Convert to int if it's a valid number
                    if isinstance(ach_value, (int, float)) and not pd.isna(ach_value):
                        df_with_values.at[row_idx, new_col] = int(ach_value) if ach_value == int(ach_value) else ach_value
                    else:
                        df_with_values.at[row_idx, new_col] = 0

# Save values version
values_file = f"MPO_Target_vs_Achievement_Values_{timestamp}.xlsx"
df_with_values.to_excel(values_file, sheet_name='Target_vs_Achievement', index=False)

print(f"✓ Values version saved: {values_file}")

# ============================================================================
# SUMMARY
# ============================================================================

print("\n" + "=" * 80)
print("SUMMARY")
print("=" * 80)
print(f"Files created:")
print(f"  1. {output_file} (with VLOOKUP formulas)")
print(f"  2. {values_file} (with calculated values)")
print(f"\nStructure:")
print(f"  - Row 1: Column headers")
print(f"  - Row 2: Product codes (for ACH columns only)")
print(f"  - Row 3: 'TARGET' and 'ACH.' labels")
print(f"  - Row 4+: MPO data with alternating target and achievement columns")
print(f"\nProduct codes matched: {matched}/{len(product_cols)}")
print(f"\nColumn structure:")
print(f"  - Base columns: {len(base_cols)}")
print(f"  - Product columns: {len(new_col_mapping)} ({matched} products × 2 columns each)")

print("\n" + "=" * 80)
print(f"Completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print("=" * 80)
