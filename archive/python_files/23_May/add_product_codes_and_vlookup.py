"""
Add Product Codes Row and VLOOKUP Formulas
==========================================
1. Load MPO target file
2. Load achievement file to get product codes
3. Insert row 1 with product codes
4. Insert row 2 with "ACH." labels
5. Add VLOOKUP formulas for achievements

Author: Auto-generated
Date: 2026-05-23
"""

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# ============================================================================
# CONFIGURATION
# ============================================================================

TARGET_FILE = r"C:\Users\Irak\Desktop\Barishal April Data\MPO_Field_With_Targets_20260523_102519.xlsx"
ACHIEVEMENT_FILE = r"C:\Users\Irak\Desktop\Barishal April Data\Achievement_Pivot_20260523_104633.xlsx"

print("=" * 80)
print("ADD PRODUCT CODES AND VLOOKUP FORMULAS")
print("=" * 80)
print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

# ============================================================================
# STEP 1: LOAD FILES
# ============================================================================

print("STEP 1: Loading files...")

df_target = pd.read_excel(TARGET_FILE, sheet_name='MPO_Field_Targets')
df_achievement = pd.read_excel(ACHIEVEMENT_FILE, sheet_name='April_Achievement')

print(f"  - Target records: {len(df_target)}")
print(f"  - Achievement records: {len(df_achievement)}")

# ============================================================================
# STEP 2: CREATE PRODUCT CODE MAPPING
# ============================================================================

print("\nSTEP 2: Creating product code mapping...")

# Get unique product codes and names from achievement file
product_mapping = df_achievement[['Product_Code', 'Product_Name']].drop_duplicates()
product_code_dict = dict(zip(product_mapping['Product_Name'], product_mapping['Product_Code']))

print(f"  - Unique products in achievement: {len(product_code_dict)}")
print(f"  - Sample mappings:")
for i, (name, code) in enumerate(list(product_code_dict.items())[:5]):
    print(f"    {name} → {code}")

# ============================================================================
# STEP 3: MATCH TARGET PRODUCTS TO CODES
# ============================================================================

print("\nSTEP 3: Matching target products to product codes...")

# Get product columns from target
base_cols = ['DEPOT', 'ZONE', 'FM/AM, ZONE', 'MARKET', 'MPO CODE', 'DEPOT_MPO_CODE']
product_cols = [col for col in df_target.columns if col not in base_cols + ['Total_Target']]

# Match target products to achievement products (from previous matching)
from difflib import SequenceMatcher
import re

def normalize_product_name(name):
    if pd.isna(name):
        return ""
    name = str(name).upper().strip()
    name = re.sub(r'\s*(TAB|TABLET|CAP|CAPSULE|SYP|SYRUP|SUSP|SUSPENSION)\.?\s*', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\([^)]*\)', '', name)
    name = re.sub(r'\d+\s*MG', '', name, flags=re.IGNORECASE)
    name = name.replace('-', '').replace('.', '').replace(',', '').replace('+', '')
    name = ' '.join(name.split())
    return name.replace(' ', '')

def find_product_code(target_product, product_code_dict):
    target_norm = normalize_product_name(target_product)
    
    best_match = None
    best_score = 0
    
    for ach_product, code in product_code_dict.items():
        ach_norm = normalize_product_name(ach_product)
        score = SequenceMatcher(None, target_norm, ach_norm).ratio()
        
        if score > best_score:
            best_score = score
            best_match = code
    
    if best_score >= 0.75:
        return best_match
    else:
        return None

# Create mapping of target product to product code
target_to_code = {}
for prod in product_cols:
    code = find_product_code(prod, product_code_dict)
    target_to_code[prod] = code if code else ""

matched = sum(1 for v in target_to_code.values() if v)
print(f"  - Matched: {matched}/{len(product_cols)}")

# ============================================================================
# STEP 4: CREATE NEW DATAFRAME WITH PRODUCT CODES
# ============================================================================

print("\nSTEP 4: Creating new dataframe with product codes...")

# Create row 1: Product codes
row1_data = {}
for col in base_cols:
    row1_data[col] = ""
for prod in product_cols:
    row1_data[prod] = target_to_code.get(prod, "")

# Create row 2: "ACH." labels
row2_data = {}
for col in base_cols:
    row2_data[col] = ""
for prod in product_cols:
    row2_data[prod] = "ACH."

# Create new dataframe with these rows at the top
df_row1 = pd.DataFrame([row1_data])
df_row2 = pd.DataFrame([row2_data])

# Concatenate: row1, row2, then original data
df_final = pd.concat([df_row1, df_row2, df_target], ignore_index=True)

print(f"  - Final shape: {df_final.shape}")
print(f"  - Total rows: {len(df_final)} (2 header rows + {len(df_target)} data rows)")

# ============================================================================
# STEP 5: SAVE TO EXCEL WITH VLOOKUP FORMULAS
# ============================================================================

print("\nSTEP 5: Saving to Excel with VLOOKUP formulas...")

timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
output_file = f"MPO_Target_With_ProductCodes_{timestamp}.xlsx"

# Save initial dataframe
df_final.to_excel(output_file, sheet_name='Target_with_Codes', index=False)

# Now add VLOOKUP formulas using openpyxl
print("  - Adding VLOOKUP formulas...")

wb = load_workbook(output_file)
ws = wb['Target_with_Codes']

# Find column indices for products (after base columns)
base_col_count = len(base_cols)

# Add formulas starting from row 4 (row 1 = headers, row 2 = codes, row 3 = ACH, row 4+ = data)
for row_idx in range(4, len(df_final) + 2):  # +2 because Excel is 1-indexed and we have 2 extra rows
    for col_idx, prod in enumerate(product_cols, start=base_col_count + 1):
        col_letter = get_column_letter(col_idx)
        
        # Get product code from row 2 (Excel row 2)
        product_code_cell = f"{col_letter}$2"
        
        # DEPOT_MPO_CODE is in column F
        depot_mpo_cell = f"$F{row_idx}"
        
        # VLOOKUP formula
        # =VLOOKUP($F4&G$2,[Achievement_Pivot_20260523_104633.xlsx]April_Achievement!$A:$G,7,FALSE)
        formula = f'=IFERROR(VLOOKUP({depot_mpo_cell}&{product_code_cell},[{ACHIEVEMENT_FILE}]April_Achievement!$A:$G,7,FALSE),0)'
        
        # Only add formula if product code exists
        if target_to_code.get(prod):
            ws[f"{col_letter}{row_idx}"] = formula

print(f"  - Formulas added for rows 4 to {len(df_final) + 1}")

# Save workbook
wb.save(output_file)
wb.close()

print(f"\n✓ Output saved: {output_file}")

# ============================================================================
# STEP 6: CREATE VERSION WITHOUT FORMULAS (VALUES ONLY)
# ============================================================================

print("\nSTEP 6: Creating version with calculated values...")

# Load the file we just created to calculate formulas
df_with_formulas = pd.read_excel(output_file, sheet_name='Target_with_Codes')

# Create achievement lookup
achievement_lookup = {}
for idx, row in df_achievement.iterrows():
    key = row['DepotMPO_CodeProduct_Code']  # This is already concatenated
    achievement_lookup[key] = row['April_Achievement']

# Convert product columns to object type to allow mixed data
for prod in product_cols:
    df_with_formulas[prod] = df_with_formulas[prod].astype(object)

# Add achievement values
for col_idx, prod in enumerate(product_cols):
    if target_to_code.get(prod):
        code = target_to_code[prod]
        
        # Calculate achievements for each row (starting from row 2, which is index 2 in dataframe)
        for row_idx in range(2, len(df_with_formulas)):
            depot_mpo = df_with_formulas.iloc[row_idx]['DEPOT_MPO_CODE']
            if pd.notna(depot_mpo):
                key = f"{depot_mpo}_{code}".upper()
                ach_value = achievement_lookup.get(key, 0)
                # Convert to int if it's a valid number, otherwise keep as is
                if isinstance(ach_value, (int, float)) and not pd.isna(ach_value):
                    df_with_formulas.at[row_idx, prod] = int(ach_value) if ach_value == int(ach_value) else ach_value
                else:
                    df_with_formulas.at[row_idx, prod] = 0

# Save values version
values_file = f"MPO_Target_With_Achievement_Values_{timestamp}.xlsx"
df_with_formulas.to_excel(values_file, sheet_name='Target_vs_Achievement', index=False)

print(f"✓ Values version saved: {values_file}")

# ============================================================================
# SUMMARY
# ============================================================================

print("\n" + "=" * 80)
print("SUMMARY")
print("=" * 80)
print(f"Files created:")
print(f"  1. {output_file} (with VLOOKUP formulas)")
print(f"  2. {values_file} (with calculated values)")
print(f"\nStructure:")
print(f"  - Row 1: Column headers")
print(f"  - Row 2: Product codes (for VLOOKUP)")
print(f"  - Row 3: 'ACH.' labels")
print(f"  - Row 4+: MPO data with targets and achievements")
print(f"\nProduct codes matched: {matched}/{len(product_cols)}")

print("\n" + "=" * 80)
print(f"Completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print("=" * 80)
