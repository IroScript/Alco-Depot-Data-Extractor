import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

df = pd.read_excel('MPO_Target_vs_Achievement_20260523_122119.xlsx', sheet_name='Target_vs_Achievement')

print("Column order in dataframe:")
for i, col in enumerate(df.columns):
    excel_col = get_column_letter(i + 1)
    print(f"{excel_col} (idx {i}): {col}")
    if i >= 12:
        break

wb = load_workbook('MPO_Target_vs_Achievement_20260523_122119.xlsx')
ws = wb['Target_vs_Achievement']

print("\n\nExcel cell values:")
for col_letter in ['G', 'H', 'I', 'J', 'K', 'L']:
    print(f"\n{col_letter}:")
    print(f"  Row 2: {ws[f'{col_letter}2'].value}")
    print(f"  Row 3: {ws[f'{col_letter}3'].value}")
    val = ws[f'{col_letter}4'].value
    if isinstance(val, str) and len(val) > 80:
        print(f"  Row 4: {val[:80]}...")
    else:
        print(f"  Row 4: {val}")

wb.close()
