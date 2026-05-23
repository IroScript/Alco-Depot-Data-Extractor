from openpyxl import load_workbook
import pandas as pd

# Check the structure
df = pd.read_excel('MPO_Target_vs_Achievement_20260523_121822.xlsx', sheet_name='Target_vs_Achievement')

print("=" * 80)
print("STRUCTURE VERIFICATION")
print("=" * 80)

print(f"\nShape: {df.shape}")
print(f"\nFirst 10 columns:")
for i, col in enumerate(df.columns[:10]):
    print(f"  {i+1}. {col}")

print(f"\nRow 1 (Product codes for ACH columns):")
print(df.iloc[0, 6:12].to_dict())

print(f"\nRow 2 (TARGET/ACH labels):")
print(df.iloc[1, 6:12].to_dict())

print(f"\nRow 3 (First data row):")
print(df.iloc[2, :8].to_dict())

# Check formulas
wb = load_workbook('MPO_Target_vs_Achievement_20260523_121822.xlsx')
ws = wb['Target_vs_Achievement']

print(f"\n\nFORMULA VERIFICATION")
print("=" * 80)

print(f"\nColumn G (should be TARGET - no formula):")
print(f"  G2: {ws['G2'].value}")
print(f"  G3: {ws['G3'].value}")
print(f"  G4: {ws['G4'].value}")

print(f"\nColumn H (should be ACH - with formula):")
print(f"  H2: {ws['H2'].value}")
print(f"  H3: {ws['H3'].value}")
print(f"  H4: {ws['H4'].value}")

print(f"\nColumn I (should be TARGET - no formula):")
print(f"  I2: {ws['I2'].value}")
print(f"  I3: {ws['I3'].value}")
print(f"  I4: {ws['I4'].value}")

print(f"\nColumn J (should be ACH - with formula):")
print(f"  J2: {ws['J2'].value}")
print(f"  J3: {ws['J3'].value}")
print(f"  J4: {ws['J4'].value}")

wb.close()
