from openpyxl import load_workbook

wb = load_workbook('MPO_Target_With_ProductCodes_20260523_121519.xlsx')
ws = wb['Target_with_Codes']

print("Sample formulas:")
print(f"G4: {ws['G4'].value}")
print(f"H4: {ws['H4'].value}")
print(f"I4: {ws['I4'].value}")

print("\nProduct codes in row 2:")
print(f"G2: {ws['G2'].value}")
print(f"H2: {ws['H2'].value}")
print(f"I2: {ws['I2'].value}")

print("\nDEPOT_MPO_CODE in F column:")
print(f"F4: {ws['F4'].value}")
print(f"F5: {ws['F5'].value}")

wb.close()
