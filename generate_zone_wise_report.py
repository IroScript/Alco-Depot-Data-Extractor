import os
import json
import gspread
import pandas as pd
import glob
from datetime import datetime
from google.oauth2.service_account import Credentials
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def main():
    print("=" * 80)
    print("ZONE-WISE PRODUCT SALES REPORT GENERATOR")
    print("=" * 80)

    # 1. Load configuration and credentials
    config_path = r'c:\Users\Irak\Desktop\Barishal April Data\FieldEdit\config.json'
    creds_path = r'c:\Users\Irak\Desktop\Barishal April Data\FieldEdit\alco-pharma-cf4b49e394bb.json'

    if not os.path.exists(config_path):
        print(f"ERROR: Configuration not found at {config_path}")
        return
    if not os.path.exists(creds_path):
        print(f"ERROR: Credentials not found at {creds_path}")
        return

    with open(config_path, 'r') as f:
        config = json.load(f)

    print("Connecting to Google Sheets API...")
    scopes = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
    creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
    client = gspread.authorize(creds)
    sheet = client.open_by_key(config['spreadsheet_id'])

    # Get the worksheet by GID
    ws = None
    for w in sheet.worksheets():
        if str(w.id) == str(config.get('gid', '1918615875')):
            ws = w
            break
    if not ws:
        ws = sheet.get_worksheet(0)

    print(f"Connected to sheet: '{sheet.title}' | Worksheet: '{ws.title}'")

    # Fetch all values from Google Sheet
    all_values = ws.get_all_values()
    if not all_values:
        print("ERROR: Worksheet is empty!")
        return

    # 2. Parse and clean Google Sheet rows
    # Apply cutoff for "FM (SELF APP CODE)" / "FM (SELF"
    cutoff_idx = len(all_values)
    for r_idx in range(1, len(all_values)):
        row = all_values[r_idx]
        row_str_joint = " ".join([str(c) for c in row if c is not None])
        if "FM (SELF APP CODE)" in row_str_joint or "FM (SELF" in row_str_joint:
            cutoff_idx = r_idx
            break
    
    header = all_values[0]
    data_rows = all_values[1:cutoff_idx]
    df_sheet = pd.DataFrame(data_rows, columns=header)

    # Clean and filter master sheet rows
    # We require DEPOT, ZONE, and MARKET to be non-empty
    df_sheet = df_sheet[
        (df_sheet['DEPOT'].str.strip() != '') &
        (df_sheet['ZONE'].str.strip() != '') &
        (df_sheet['MARKET'].str.strip() != '')
    ].copy()

    # Normalize names for grouping/sorting
    df_sheet['SM'] = df_sheet['SM'].str.strip().str.upper().fillna('VACANT')
    df_sheet['ZONE'] = df_sheet['ZONE'].str.strip().str.upper()
    df_sheet['FM/AM'] = df_sheet['FM/AM'].str.strip().str.upper().fillna('VACANT')
    df_sheet['MARKET'] = df_sheet['MARKET'].str.strip().str.upper()
    df_sheet['MPO NAME, JUN\'26'] = df_sheet["MPO NAME, JUN'26"].str.strip().str.upper().fillna('VACANT')
    df_sheet['APP CODE (FINAL)'] = df_sheet['APP CODE (FINAL)'].str.strip().str.upper()
    df_sheet['DREAM APPS MPO CODE'] = df_sheet['DREAM APPS MPO CODE'].str.strip().str.upper()
    df_sheet['DEPOT'] = df_sheet['DEPOT'].str.strip().str.upper()

    print(f"Loaded {len(df_sheet)} active field force records from Google Sheet.")

    # 3. Load latest sales CSV
    csv_files = glob.glob(r'c:\Users\Irak\Desktop\Barishal April Data\Product_Level_Net_Sales_*.csv')
    if not csv_files:
        print("ERROR: No Product_Level_Net_Sales CSV found!")
        return

    latest_csv = max(csv_files, key=os.path.getctime)
    print(f"Loading sales data from: {os.path.basename(latest_csv)}")
    df_sales = pd.read_csv(latest_csv)

    # Clean CSV values
    df_sales['Depot'] = df_sales['Depot'].str.strip().str.upper()
    df_sales['MPO_Code'] = df_sales['MPO_Code'].str.strip().str.upper()
    df_sales['Product_Code'] = df_sales['Product_Code'].str.strip().str.upper()
    df_sales['Product_Name'] = df_sales['Product_Name'].str.strip()
    df_sales['Month'] = df_sales['Month'].str.strip()

    # Determine the standard product name for each product code (most frequent name in CSV)
    print("Determining standard product names for each product code...")
    prod_name_counts = df_sales.groupby(['Product_Code', 'Product_Name']).size().reset_index(name='count')
    idx_max = prod_name_counts.groupby('Product_Code')['count'].idxmax()
    standard_names = prod_name_counts.loc[idx_max, ['Product_Code', 'Product_Name']]
    code_to_name = dict(zip(standard_names['Product_Code'], standard_names['Product_Name']))

    # Get unique product codes (sorted by their standard name) and months
    unique_product_codes = sorted(df_sales['Product_Code'].dropna().unique(), key=lambda code: code_to_name.get(code, code).upper())
    unique_months = sorted(df_sales['Month'].dropna().unique())

    # Map month keys for display (e.g. 2026-01 -> Jan, 2026-02 -> Feb)
    month_mapping = {
        '2026-01': 'Jan',
        '2026-02': 'Feb',
        '2026-03': 'Mar',
        '2026-04': 'Apr',
        '2026-05': 'May',
        '2026-06': 'Jun'
    }
    month_headers = [month_mapping.get(m, m) for m in unique_months]

    print(f"Found {len(unique_product_codes)} unique product codes and {len(unique_months)} months: {unique_months}")

    # Build sales lookup dictionary for fast lookup (aggregated by Product_Code)
    print("Building sales aggregation lookup...")
    sales_grouped = df_sales.groupby(['Depot', 'MPO_Code', 'Product_Code', 'Month'])['ACTUAL_SALE_QTY'].sum().reset_index()
    
    sales_lookup = {}
    for _, row in sales_grouped.iterrows():
        key = (row['Depot'], row['MPO_Code'], row['Product_Code'], row['Month'])
        sales_lookup[key] = row['ACTUAL_SALE_QTY']

    # 4. Sort sheet rows hierarchically
    print("Sorting field forces hierarchically by SM -> ZONE -> FM/AM -> MARKET...")
    # Sorting order
    df_sheet = df_sheet.sort_values(by=['SM', 'ZONE', 'FM/AM', 'MARKET']).reset_index(drop=True)

    # 5. Create Excel Workbook
    wb = Workbook()
    ws_out = wb.active
    ws_out.title = "Zone Wise Product Sales"
    ws_out.views.sheetView[0].showGridLines = True

    # Colors and styles
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Blue
    sub_header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid") # Light Blue-Gray
    total_fill = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid") # Soft Purple
    
    font_title = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    font_sub_title = Font(name=font_family, size=9, bold=True, color="1F497D")
    font_bold = Font(name=font_family, size=9, bold=True)
    font_regular = Font(name=font_family, size=9)

    border_thin = Side(border_style="thin", color="D9D9D9")
    border_double = Side(border_style="double", color="366092")
    grid_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    bottom_total_border = Border(top=border_thin, bottom=border_double, left=border_thin, right=border_thin)

    # Define metadata columns
    metadata_cols = [
        ('SM', 'SM'),
        ('ZONE', 'ZONE'),
        ('FM/AM', 'FM/AM'),
        ('DEPOT', 'DEPOT'),
        ('MARKET', 'MARKET'),
        ('APP CODE (FINAL)', 'APP CODE'),
        ("MPO NAME, JUN'26", 'MPO NAME')
    ]
    num_metadata_cols = len(metadata_cols)

    # Write header rows (3-row layout)
    # Metadata columns will have their labels and will be merged vertically across 3 rows
    for i, (col_key, col_label) in enumerate(metadata_cols, 1):
        cell = ws_out.cell(row=1, column=i, value=col_label)
        cell.font = font_title
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_out.merge_cells(start_row=1, start_column=i, end_row=3, end_column=i)
        
        # Apply header fill to merged rows for styling consistency
        ws_out.cell(row=2, column=i).fill = header_fill
        ws_out.cell(row=3, column=i).fill = header_fill

    # Write Product Codes on Row 1 (merged across months)
    current_col = num_metadata_cols + 1
    for prod_code in unique_product_codes:
        start_col = current_col
        end_col = start_col + len(unique_months) - 1
        ws_out.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        
        cell = ws_out.cell(row=1, column=start_col, value=prod_code)
        cell.font = font_title
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for c in range(start_col + 1, end_col + 1):
            ws_out.cell(row=1, column=c).fill = header_fill

        current_col = end_col + 1

    # Write Product Names on Row 2 (merged across months)
    current_col = num_metadata_cols + 1
    for prod_code in unique_product_codes:
        start_col = current_col
        end_col = start_col + len(unique_months) - 1
        ws_out.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        
        prod_name = code_to_name.get(prod_code, prod_code)
        cell = ws_out.cell(row=2, column=start_col, value=prod_name)
        cell.font = font_title
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for c in range(start_col + 1, end_col + 1):
            ws_out.cell(row=2, column=c).fill = header_fill

        current_col = end_col + 1

    # Write Month Names on Row 3 (not merged)
    current_col = num_metadata_cols + 1
    for prod_code in unique_product_codes:
        for m_hdr in month_headers:
            cell = ws_out.cell(row=3, column=current_col, value=m_hdr)
            cell.font = font_sub_title
            cell.fill = sub_header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            current_col += 1

    ws_out.row_dimensions[1].height = 24
    ws_out.row_dimensions[2].height = 24
    ws_out.row_dimensions[3].height = 20

    # Write data rows
    start_row = 4
    print("Populating data cells...")
    for idx, row in df_sheet.iterrows():
        r_idx = start_row + idx
        depot = row['DEPOT']
        zone = row['ZONE']
        sm = row['SM']
        fm = row['FM/AM']
        market = row['MARKET']
        app_code = row['APP CODE (FINAL)']
        dream_code = row['DREAM APPS MPO CODE']
        mpo_name = row["MPO NAME, JUN'26"]

        # Write metadata
        ws_out.cell(row=r_idx, column=1, value=sm).font = font_regular
        ws_out.cell(row=r_idx, column=2, value=zone).font = font_regular
        ws_out.cell(row=r_idx, column=3, value=fm).font = font_regular
        ws_out.cell(row=r_idx, column=4, value=depot).font = font_regular
        ws_out.cell(row=r_idx, column=5, value=market).font = font_regular
        ws_out.cell(row=r_idx, column=6, value=app_code).font = font_regular
        ws_out.cell(row=r_idx, column=7, value=mpo_name).font = font_regular

        for c_idx in range(1, num_metadata_cols + 1):
            ws_out.cell(row=r_idx, column=c_idx).border = grid_border
            ws_out.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal="left", vertical="center")

        # Write product sales quantities
        current_col = num_metadata_cols + 1
        for prod_code in unique_product_codes:
            for m in unique_months:
                # Look up sales using both possible codes for maximum coverage
                qty = 0
                if app_code:
                    qty = sales_lookup.get((depot, app_code, prod_code, m), 0)
                if qty == 0 and dream_code:
                    qty = sales_lookup.get((depot, dream_code, prod_code, m), 0)

                cell = ws_out.cell(row=r_idx, column=current_col, value=qty)
                cell.font = font_regular
                cell.border = grid_border
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0'
                current_col += 1

        ws_out.row_dimensions[r_idx].height = 18

    # 6. Add Total Row at the bottom
    total_row_idx = start_row + len(df_sheet)
    ws_out.cell(row=total_row_idx, column=1, value="GRAND TOTAL").font = font_bold
    ws_out.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=num_metadata_cols)
    
    # Apply styling to merged total cells
    for c_idx in range(1, num_metadata_cols + 1):
        cell = ws_out.cell(row=total_row_idx, column=c_idx)
        cell.fill = total_fill
        cell.border = bottom_total_border
    
    ws_out.cell(row=total_row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws_out.row_dimensions[total_row_idx].height = 22

    # Write SUM formulas for each product month column (starts from row 4 now)
    current_col = num_metadata_cols + 1
    for prod_code in unique_product_codes:
        for m in unique_months:
            col_letter = get_column_letter(current_col)
            formula = f"=SUM({col_letter}4:{col_letter}{total_row_idx-1})"
            
            cell = ws_out.cell(row=total_row_idx, column=current_col, value=formula)
            cell.font = font_bold
            cell.fill = total_fill
            cell.border = bottom_total_border
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '#,##0'
            current_col += 1

    # 7. Format columns width and freeze panes
    print("Formatting column widths...")
    # Metadata columns auto-fit
    for i in range(1, num_metadata_cols + 1):
        col_letter = get_column_letter(i)
        max_len = 0
        for row in range(1, total_row_idx + 1):
            val = ws_out.cell(row=row, column=i).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws_out.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Product monthly columns width
    for i in range(num_metadata_cols + 1, current_col):
        col_letter = get_column_letter(i)
        ws_out.column_dimensions[col_letter].width = 8

    # Freeze panes at Row 4 (below triple header) and Column 8 (after MPO NAME)
    ws_out.freeze_panes = "H4"

    # Save output workbook
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_filename = f"Zone_Wise_Product_Sales_Report_{timestamp}.xlsx"
    output_filepath = os.path.join(r"C:\Users\Irak\Desktop\Barishal April Data", output_filename)
    
    print(f"Saving report to: {output_filepath}...")
    wb.save(output_filepath)
    wb.close()
    
    print(f"\n[OK] Report created successfully!")
    print(f"Output File: {output_filename}")
    print("=" * 80)

if __name__ == "__main__":
    main()
