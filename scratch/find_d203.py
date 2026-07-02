import sqlite3
import os
import openpyxl

print("--- Checking sales.db ---")
if os.path.exists("sales.db"):
    conn = sqlite3.connect("sales.db")
    tables = [t[0] for t in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
    print("Tables in sales.db:", tables)
    for table in tables:
        cols = [c[1] for c in conn.execute(f"PRAGMA table_info({table})").fetchall()]
        for col in cols:
            rows = conn.execute(f"SELECT * FROM {table} WHERE CAST({col} AS TEXT) LIKE '%203%' OR CAST({col} AS TEXT) LIKE '%DHAKA-1%' LIMIT 5").fetchall()
            for r in rows:
                print(f"[sales.db -> {table} -> {col}]: {r}")

print("\n--- Checking all excel files in archive/recent and FieldEdit ---")
for root, dirs, files in os.walk("."):
    for f in files:
        if f.endswith(".xlsx") or f.endswith(".xls"):
            path = os.path.join(root, f)
            try:
                wb = openpyxl.load_workbook(path, data_only=True)
                for sname in wb.sheetnames:
                    ws = wb[sname]
                    for r_idx, row in enumerate(ws.iter_rows(max_row=500), 1):
                        for c_idx, cell in enumerate(row, 1):
                            val = str(cell.value) if cell.value is not None else ""
                            if "203" in val or "DHAKA-1" in val.upper():
                                print(f"[{path} -> {sname} -> row {r_idx} col {c_idx} ({openpyxl.utils.get_column_letter(c_idx)})]: {val}")
            except Exception as e:
                pass
