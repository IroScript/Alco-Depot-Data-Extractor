import openpyxl
import glob
import os

def main():
    base_dir = r"c:\Users\Irak\Desktop\Barishal April Data"
    files = glob.glob(os.path.join(base_dir, "04_Analyzed_10_Param_Zone_Wise_Sales_Grouped_Report_*.xlsx"))
    files.sort(key=os.path.getmtime, reverse=True)
    latest_file = files[0]
    print(f"Inspecting file: {os.path.basename(latest_file)}")
    
    wb = openpyxl.load_workbook(latest_file, data_only=False)
    ws = wb.active
    
    # We want to check row 474 (CV %) for columns:
    # RL (col 246 approx), RV (col 256 approx)
    # Let's find columns by letters
    cols_to_check = ['RL', 'RM', 'RN', 'RO', 'RP', 'RQ', 'RR', 'RS', 'RT', 'RU', 'RV']
    
    for col_let in cols_to_check:
        cell_mean = ws[f"{col_let}467"]
        cell_sd = ws[f"{col_let}470"]
        cell_cv = ws[f"{col_let}474"]
        print(f"Col {col_let} | MEAN: {cell_mean.value} | SD: {cell_sd.value} | CV%: {cell_cv.value}")

if __name__ == "__main__":
    main()
