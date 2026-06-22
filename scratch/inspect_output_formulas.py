import openpyxl
import glob
import os

def main():
    base_dir = r"c:\Users\Irak\Desktop\Barishal April Data"
    files = glob.glob(os.path.join(base_dir, "04_Analyzed_10_Param_Zone_Wise_Sales_Grouped_Report_*.xlsx"))
    if not files:
        print("No output file found!")
        return
    
    files.sort(key=os.path.getmtime, reverse=True)
    latest_file = files[0]
    print(f"Inspecting file: {os.path.basename(latest_file)}")
    
    wb = openpyxl.load_workbook(latest_file, data_only=False)
    ws = wb.active
    
    # Let's inspect columns to find Avg columns
    avg_cols = []
    for col in range(1, ws.max_column + 1):
        val_r3 = ws.cell(row=3, column=col).value
        if val_r3 and str(val_r3).strip().upper() == "AVG":
            avg_cols.append(col)
            
    print(f"Found average columns at indices: {avg_cols}")
    
    # Inspect row 67 specifically (or rows around 64 to 74)
    print("\n--- Rows 64 to 74 in first Avg column ---")
    first_avg_col = avg_cols[0] if avg_cols else 17
    for r in range(64, 75):
        cell_val = ws.cell(row=r, column=first_avg_col).value
        label = ws.cell(row=r, column=2).value
        print(f"Row {r} | Label: {label} | Value/Formula: {cell_val}")
        
    # Inspect a parameter block, e.g. for MEDIAN, VARP, STDEVP
    print("\n--- Parameter formulas for a column ---")
    median_val = ws.cell(row=67, column=12).value  # Let's see what is in row 67 column 12
    print(f"Row 67 Col 12 Value: {median_val}")
    
    # Let's inspect some rows to check where MEDIAN is written
    for r in range(60, 85):
        label = str(ws.cell(row=r, column=2).value or "").strip().upper()
        if label in ["MEDIAN", "VAR", "MEAN", "COUNT"]:
            print(f"Row {r} | Parameter: {label} | Col 12 Formula: {ws.cell(row=r, column=12).value}")

if __name__ == "__main__":
    main()
