import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.formula import ArrayFormula

def main():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 10
    ws['A2'] = 20
    ws['A3'] = 2
    ws['A4'] = 30
    
    # Write array formula
    ws['B1'] = ArrayFormula('B1', '=MEDIAN(IF(A1:A4>5, A1:A4))')
    
    out_path = r"c:\Users\Irak\Desktop\Barishal April Data\scratch\test_array_formula.xlsx"
    wb.save(out_path)
    print("Workbook saved successfully with ArrayFormula!")

if __name__ == "__main__":
    main()
