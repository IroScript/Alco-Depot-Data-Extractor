import openpyxl
from openpyxl import Workbook

def main():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 10
    ws['A2'] = 20
    ws['A3'] = 2
    ws['A4'] = 30
    
    # Write formula with prefix
    ws['B1'] = "=MEDIAN(_xlfn.FILTER(A1:A4, A1:A4>5))"
    
    out_path = r"c:\Users\Irak\Desktop\Barishal April Data\scratch\test_filter_prefix.xlsx"
    wb.save(out_path)
    print("Workbook saved successfully with _xlfn.FILTER!")

if __name__ == "__main__":
    main()
