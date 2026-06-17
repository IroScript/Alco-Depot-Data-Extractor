import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import math
import random
import threading

# ══════════════════════════════════════════════════════════════════
#  Design tokens (From SmallApp)
# ══════════════════════════════════════════════════════════════════
_C = {
    'void'   : '#03040A',
    'panel'  : '#060816',
    'panel2' : '#050912',
    'cyan'   : '#00F2FE',
    'mag'    : '#FF00AA',
    'violet' : '#9D00FF',
    'green'  : '#00FF88',
    'red'    : '#FF2244',
    'text'   : '#D8F0F8',
    'muted'  : '#5F98A7',
    'border' : '#102A45',
    'hot'    : '#0D2A3F',
}

W_WIN, H_WIN = 490, 295

class ZoneDataAnalyzerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("ALCO ZONE DATA ANALYZER")
        self.root.geometry(f"{W_WIN}x{H_WIN}")
        self.root.resizable(False, False)
        self.root.configure(bg=_C['void'])

        # State
        self.input_file = tk.StringVar()
        self.output_dir = tk.StringVar()
        self.opt_exclude_vacant = tk.BooleanVar(value=True)
        self._t          = 0.0
        self._scanline_y = 0.0
        self._stars      = []
        self._after_id   = None

        self._build()
        self._seed_stars()
        self._tick()

    def _build(self):
        W, H = W_WIN, H_WIN

        self.cv = tk.Canvas(self.root, width=W, height=H, bg=_C['void'], highlightthickness=0, bd=0)
        self.cv.place(x=0, y=0)

        self._paint_gradient(W, H)
        self._paint_grid(W, H)

        self._sl = self.cv.create_line(0, 0, W, 0, fill='#00F2FE', width=1, stipple='gray12')

        self._paint_brackets(W, H)
        self._paint_title(W)

        # File Input
        self._draw_label(20, 55, "INPUT REPORT EXCEL FILE")
        self._draw_entry(20, 75, self.input_file, w=350)
        self._draw_button(385, 75, "BROWSE", self.browse_input, w=80)

        # Output Folder
        self._draw_label(20, 115, "OUTPUT FOLDER")
        self._draw_entry(20, 135, self.output_dir, w=350)
        self._draw_button(385, 135, "BROWSE", self.browse_output, w=80)

        # Options
        cb = tk.Checkbutton(self.root, text="Exclude Vacant MPOs (If still present)",
                            variable=self.opt_exclude_vacant, onvalue=True, offvalue=False,
                            font=('Courier New', 8), fg=_C['text'], bg=_C['void'],
                            activebackground=_C['void'], activeforeground=_C['cyan'], selectcolor=_C['panel'])
        cb.place(x=20, y=175)

        # Status
        self.lbl_status = tk.Label(self.root, text="◉ SYSTEM READY", font=('Courier New', 8), fg=_C['cyan'], bg=_C['void'])
        self.lbl_status.place(x=20, y=215)

        # Progress Bar
        self.progress_bar = self.cv.create_rectangle(20, 245, 20, 248, fill=_C['cyan'], outline='')

        # Execute
        self.exec_btn = tk.Button(self.root, text="⟨  RUN 10 PARAMETER ANALYSIS  ⟩", command=self.run_process,
                                  font=('Courier New', 10, 'bold'), fg=_C['cyan'], bg='#030A1C', relief='flat',
                                  cursor='hand2', activebackground='#061C34', activeforeground='#FFFFFF', bd=1,
                                  highlightbackground=_C['cyan'])
        self.exec_btn.place(x=20, y=260, width=450, height=24)

    def _paint_gradient(self, W, H):
        for y in range(H):
            t = y / H
            r = int(3  + t * 4)
            g = int(4  + t * 5)
            b = int(10 + t * 16)
            self.cv.create_line(0, y, W, y, fill=f'#{r:02x}{g:02x}{b:02x}', width=1)

    def _paint_grid(self, W, H):
        for x in range(0, W, 40):
            self.cv.create_line(x, 0, x, H, fill='#07091A', width=1)
        for y in range(0, H, 40):
            self.cv.create_line(0, y, W, y, fill='#07091A', width=1)

    def _paint_brackets(self, W, H, pad=10, sz=18):
        c = _C['cyan']
        self.cv.create_line(pad, pad, pad+sz, pad, fill=c, width=1)
        self.cv.create_line(pad, pad, pad, pad+sz, fill=c, width=1)
        self.cv.create_line(W-pad-sz, pad, W-pad, pad, fill=c, width=1)
        self.cv.create_line(W-pad, pad, W-pad, pad+sz, fill=c, width=1)
        self.cv.create_line(pad, H-pad, pad+sz, H-pad, fill=c, width=1)
        self.cv.create_line(pad, H-pad-sz, pad, H-pad, fill=c, width=1)
        self.cv.create_line(W-pad-sz, H-pad, W-pad, H-pad, fill=c, width=1)
        self.cv.create_line(W-pad, H-pad-sz, W-pad, H-pad, fill=c, width=1)

    def _paint_title(self, W):
        self.cv.create_text(W//2, 22, text='ZONE DATA ANALYZER', font=('Courier New', 12, 'bold'), fill=_C['cyan'])
        self.cv.create_text(W//2, 36, text='∞ 10 PARAMETER STATISTICAL GENERATOR ∞', font=('Courier New', 7), fill=_C['muted'])
        self.cv.create_line(40, 46, W-40, 46, fill=_C['border'], width=1)

    def _draw_label(self, x, y, text):
        lbl = tk.Label(self.root, text=text, font=('Courier New', 7, 'bold'), fg=_C['muted'], bg=_C['void'])
        lbl.place(x=x, y=y)

    def _draw_entry(self, x, y, var, w):
        ent = tk.Entry(self.root, textvariable=var, bg=_C['panel2'], fg=_C['text'],
                       insertbackground=_C['cyan'], relief='flat', font=('Segoe UI', 8), bd=1,
                       highlightthickness=1, highlightcolor=_C['cyan'], highlightbackground=_C['border'])
        ent.place(x=x, y=y, width=w, height=22)

    def _draw_button(self, x, y, text, cmd, w):
        btn = tk.Button(self.root, text=text, command=cmd, font=('Courier New', 7, 'bold'),
                        fg=_C['cyan'], bg='#07152B', relief='flat', cursor='hand2', activebackground='#0B2A4A')
        btn.place(x=x, y=y, width=w, height=22)

    def _tick(self):
        self._t += 0.05
        self._scanline_y = (self._scanline_y + 1.5) % H_WIN
        self.cv.coords(self._sl, 0, self._scanline_y, W_WIN, self._scanline_y)

        status_text = self.lbl_status.cget('text')
        if 'READY' in status_text:
            v = int(180 + 50 * math.sin(self._t))
            self.lbl_status.configure(fg=f'#00{v:02x}{v//2:02x}')

        self._after_id = self.root.after(50, self._tick)

    def _seed_stars(self):
        colors = [_C['cyan'], '#FFFFFF', _C['violet'], '#4FACFE']
        for _ in range(15):
            x = random.randint(15, W_WIN-15)
            y = random.randint(50, H_WIN-20)
            sz = random.choice([1, 1, 2])
            sid = self.cv.create_oval(x, y, x+sz, y+sz, fill=random.choice(colors), outline='', state='hidden')
            self._stars.append({'id': sid, 'ph': random.uniform(0, 2*math.pi), 'spd': random.uniform(0.025, 0.065)})

    def browse_input(self):
        f = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if f:
            self.input_file.set(f)
            if not self.output_dir.get():
                self.output_dir.set(os.path.dirname(f))

    def browse_output(self):
        d = filedialog.askdirectory()
        if d:
            self.output_dir.set(d)

    def update_progress(self, pct, status):
        self.lbl_status.configure(text=f'◉ {status}')
        w = 450 * (pct / 100.0)
        self.cv.coords(self.progress_bar, 20, 245, 20+w, 248)
        self.root.update_idletasks()

    def run_process(self):
        ip = self.input_file.get()
        od = self.output_dir.get()
        
        if not ip or not od or not os.path.exists(ip) or not os.path.exists(od):
            messagebox.showerror("Error", "Valid input file and output directory required.")
            return
            
        self.exec_btn.configure(state='disabled')
        self.update_progress(10, 'LOADING EXCEL REPORT...')

        def task():
            try:
                wb = load_workbook(ip)
                ws = wb.active
                
                # Find start and end rows
                start_row = 4
                end_row = ws.max_row
                grand_total_row = None
                
                for r in range(end_row, 0, -1):
                    val = ws.cell(row=r, column=1).value
                    if val == "GRAND TOTAL":
                        grand_total_row = r
                        break
                
                if grand_total_row:
                    end_row = grand_total_row - 1
                
                self.update_progress(30, 'FILTERING DATA...')
                if self.opt_exclude_vacant.get():
                    for r in range(end_row, start_row - 1, -1):
                        mpo_val = str(ws.cell(row=r, column=7).value or "").strip().upper()
                        fm_val = str(ws.cell(row=r, column=3).value or "").strip().upper()
                        sm_val = str(ws.cell(row=r, column=1).value or "").strip().upper()
                        
                        if mpo_val == 'VACANT' or fm_val == 'VACANT' or sm_val == 'VACANT':
                            ws.delete_rows(r, 1)
                            end_row -= 1
                            if grand_total_row:
                                grand_total_row -= 1

                self.update_progress(50, 'FINDING ZONES & INSERTING 10 PARAMETERS...')
                
                # Identify zone boundaries
                zone_boundaries = []
                last_zone = None
                for r in range(start_row, end_row + 1):
                    z_val = str(ws.cell(row=r, column=2).value or "").strip()
                    if last_zone is not None and z_val != last_zone:
                        zone_boundaries.append((r, last_zone))
                    last_zone = z_val
                    
                if end_row >= start_row:
                    zone_boundaries.append((end_row + 1, last_zone))
                
                max_col = ws.max_column
                
                # Define 10 parameters
                param_names = [
                    "Count", "Sum", "Mean", "Median", "variance",
                    "sd", "Min", "Max", "Range", "CV %"
                ]
                
                # Process backwards to avoid messing up row indices
                for r_boundary, z_name in reversed(zone_boundaries):
                    # We need to find where this zone started to create the formulas
                    z_start = start_row
                    for r in range(r_boundary - 1, start_row - 1, -1):
                        if str(ws.cell(row=r, column=2).value or "").strip() != z_name:
                            z_start = r + 1
                            break
                    z_end = r_boundary - 1
                    
                    # Insert 10 rows
                    ws.insert_rows(r_boundary, amount=10)
                    if grand_total_row:
                        grand_total_row += 10
                        
                    # Write parameter labels in MPO NAME column (7)
                    for i, p_name in enumerate(param_names):
                        c = ws.cell(row=r_boundary + i, column=7, value=p_name)
                        # Optionally right align or style
                        
                    # Write formulas for all data columns (8 to max_col)
                    for col in range(8, max_col + 1):
                        col_l = get_column_letter(col)
                        rng = f"{col_l}{z_start}:{col_l}{z_end}"
                        
                        # Row indices for the inserted rows
                        r_count = r_boundary
                        r_sum = r_boundary + 1
                        r_mean = r_boundary + 2
                        r_median = r_boundary + 3
                        r_var = r_boundary + 4
                        r_sd = r_boundary + 5
                        r_min = r_boundary + 6
                        r_max = r_boundary + 7
                        r_range = r_boundary + 8
                        r_cv = r_boundary + 9
                        
                        # Formulas
                        ws.cell(row=r_count, column=col, value=f"=COUNT({rng})")
                        ws.cell(row=r_sum, column=col, value=f"=SUM({rng})")
                        ws.cell(row=r_mean, column=col, value=f"=AVERAGE({rng})")
                        ws.cell(row=r_median, column=col, value=f"=MEDIAN({rng})")
                        ws.cell(row=r_var, column=col, value=f"=VAR.S({rng})")
                        ws.cell(row=r_sd, column=col, value=f"=STDEV.S({rng})")
                        ws.cell(row=r_min, column=col, value=f"=MIN({rng})")
                        ws.cell(row=r_max, column=col, value=f"=MAX({rng})")
                        ws.cell(row=r_range, column=col, value=f"={col_l}{r_max}-{col_l}{r_min}")
                        ws.cell(row=r_cv, column=col, value=f"=IFERROR({col_l}{r_sd}/{col_l}{r_mean}*100, \"-\")")
                            
                self.update_progress(80, 'UPDATING GRAND TOTAL FORMULAS...')
                if grand_total_row:
                    for col in range(8, max_col + 1):
                        val = ws.cell(row=grand_total_row, column=col).value
                        if str(val).startswith("=SUM("):
                            col_letter = get_column_letter(col)
                            # Create a sum formula that ignores the intermediate parameter rows
                            # Excel's SUM ignores text, but our parameters have numbers!
                            # Since we inserted them, a simple SUM(A4:A{grand_total-1}) will double count or count the sums and counts!
                            # Better approach: SUM() for just the individual cells? Or let Excel's SUBTOTAL handle it?
                            # Or since we only want the sum of the original data rows, we can just use SUMIF on MPO NAME != these parameter names?
                            # No, SUMIF would be complicated.
                            # Since we already computed "Sum" in the parameter rows, we can just sum those "Sum" cells!
                            # Let's write a formula summing the "Sum" cells.
                            sum_cells = []
                            for r in range(start_row, grand_total_row):
                                if str(ws.cell(row=r, column=7).value) == "Sum":
                                    sum_cells.append(f"{col_letter}{r}")
                            if sum_cells:
                                new_formula = f"=SUM({','.join(sum_cells)})"
                                ws.cell(row=grand_total_row, column=col, value=new_formula)

                self.update_progress(90, 'SAVING FILE...')
                out_name = "Analysis_" + os.path.basename(ip)
                out_path = os.path.join(od, out_name)
                wb.save(out_path)
                wb.close()
                
                self.root.after(0, lambda: [
                    self.update_progress(100, 'COMPLETED SUCCESSFULLY'),
                    self.lbl_status.configure(fg=_C['green']),
                    messagebox.showinfo("Success", f"Data analysis added and saved to:\n{out_path}"),
                    self.exec_btn.configure(state='normal')
                ])
                
            except Exception as e:
                self.root.after(0, lambda err=e: [
                    self.update_progress(0, 'ERROR OCCURRED'),
                    self.lbl_status.configure(fg=_C['red']),
                    messagebox.showerror("Error", f"Failed to process file:\n{err}"),
                    self.exec_btn.configure(state='normal')
                ])

        threading.Thread(target=task, daemon=True).start()

if __name__ == "__main__":
    root = tk.Tk()
    app = ZoneDataAnalyzerApp(root)
    root.mainloop()
